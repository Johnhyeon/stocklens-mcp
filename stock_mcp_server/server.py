"""한국 주식 데이터 MCP 서버.

네이버 증권에서 차트, 수급, 재무 데이터를 가져와
Claude에서 자연어로 분석할 수 있게 해줍니다.
"""

import functools
import re
from datetime import datetime
from pathlib import Path

# TLS 신뢰 기준(OS 인증서 저장소)을 다른 무엇보다 먼저 세운다. 순서가 중요하다 —
# yfinance(curl_cffi)는 첫 요청 때 환경변수를 읽고, httpx 클라이언트도 이 기준으로
# 만들어진다. 아래 import 들보다 뒤로 밀면 미국 주식 도구만 조용히 옛 기준으로 남는다.
from stock_mcp_server import _tls as _tls_bootstrap

_tls_bootstrap.apply()

import httpx  # noqa: E402

from mcp.server.fastmcp import FastMCP  # noqa: E402
from stock_mcp_server import _pdf as pdfx  # noqa: E402
from stock_mcp_server.naver import (
    search_stock as naver_search_stock,
    NaverParseError,
    PARSE_MISS_KEY,
    get_ohlcv,
    get_current_price,
    get_investor_flow,
    get_financials,
    get_market_index,
    list_themes as naver_list_themes,
    get_theme_stocks as naver_get_theme_stocks,
    list_sectors as naver_list_sectors,
    get_sector_stocks as naver_get_sector_stocks,
    get_stock_sector as naver_get_stock_sector,
    get_volume_ranking as naver_get_volume_ranking,
    get_change_ranking as naver_get_change_ranking,
    get_alert_codes as naver_get_alert_codes,
    get_market_cap_ranking as naver_get_market_cap_ranking,
    get_multi_stocks as naver_get_multi_stocks,
    get_multi_chart_stats as naver_get_multi_chart_stats,
    scan_stocks_to_snapshot as naver_scan_snapshot,
    get_etf_list as naver_get_etf_list,
    get_etf_detail as naver_get_etf_detail,
    get_consensus as naver_get_consensus,
    get_reports as naver_get_reports,
    get_report_detail as naver_get_report_detail,
    REPORT_READ_URL as naver_report_read_url,
    get_disclosure_list as naver_get_disclosure_list,
    _ttm_eps as _ttm_eps_for,
    _latest_confirmed_annual as _latest_fin,
)
from stock_mcp_server._excel import (
    get_snapshot_dir,
    generate_filename,
    save_dataframe_to_excel,
    load_excel,
    apply_filters,
    unknown_filter_columns,
    COLUMN_LABELS_KO,
)
from stock_mcp_server import _metrics as _metrics_mod  # noqa: E402
from stock_mcp_server._metrics import (
    track_metrics,
    load_metrics,
    summarize_metrics,
    get_metrics_file,
)
from stock_mcp_server._indicators import (
    split_valid_bars,
    compute_indicators,
    AVAILABLE_INDICATORS,
)
from stock_mcp_server._chart_html import render_chart_html, render_multi_chart_html
from stock_mcp_server.market_clock import (  # noqa: E402
    format_market_clock,
    get_market_clock as build_market_clock,
    krx_calendar,
    us_calendar,
    KST,
)
from stock_mcp_server import _result_meta as rmeta
from stock_mcp_server import _watchlist as wl
from stock_mcp_server.event_reaction import (
    CORPORATE_ACTION_NOTE,
    INSUFFICIENT_POST_EVENT_HISTORY,
    PROVIDER_ERROR,
    build_event_reaction,
    format_event_reaction,
    normalize_date,
    price_adjustment_meta,
    reaction_meta_fields,
)
from stock_mcp_server.status import build_status, format_status
from stock_mcp_server import sec_edgar as sec
from stock_mcp_server import yfinance_source as us
from stock_mcp_server.licensing import is_licensed, locked_message
from stock_mcp_server._update_check import get_update_notice
import asyncio
import json
import datetime as _dt
import sys
import pandas as pd


def _support_hint() -> str:
    """safe_tool 의 '예상 못한 오류' 버킷에서만 붙이는 자가진단 안내.

    이미 원인이 명확한 예외(라이선스/타임아웃/연결오류 등)엔 안 붙인다 — 사소한 것까지
    문의로 유도하면 노이즈만 늘어난다. sys.platform 은 이 프로세스가 실제 도는 OS라
    Mac/Win 명령을 헷갈릴 일 없이 바로 골라 보여줄 수 있다.
    """
    if sys.platform == "darwin":
        log_cmd = "cd ~/Library/Logs/Claude && zip -r ~/Desktop/claude_logs.zip . && open ~/Desktop"
    elif sys.platform == "win32":
        log_cmd = (
            'powershell -c "Compress-Archive $env:APPDATA\\Claude\\logs\\* '
            '$env:USERPROFILE\\Desktop\\claude_logs.zip -Force; explorer $env:USERPROFILE\\Desktop"'
        )
    else:
        log_cmd = None
    hint = "\n\n계속되면:\n1) Claude 완전 종료 후 재시작 → 다시 시도"
    if log_cmd:
        hint += (
            "\n2) 그래도 안 되면 아래 명령으로 로그를 모아서 osy980315@gmail.com 으로 "
            f"보내주세요\n   {log_cmd}"
        )
    return hint


_QUERY_RE = re.compile(r"\?[^\s'\"]*")
_CAUSE_MAX = 160


def _cause_text(exc: Exception) -> str:
    """예외 메시지를 사용자에게 보여줄 수 있는 형태로. 없으면 예외 타입 이름.

    쿼리스트링은 통째로 지운다 — httpx 예외 문자열에는 요청 URL이 그대로 들어가고,
    거기엔 크리덴셜이 실린다(DartLens의 `?crtfc_key=<40자리>`가 확인된 사례다).
    이 문자열은 화면에 뜨고 로그에도 남고 지원 번들로 밖에 나간다.
    """
    msg = str(exc).strip()
    if not msg:
        return type(exc).__name__
    msg = _QUERY_RE.sub("?…", msg)
    if len(msg) > _CAUSE_MAX:
        msg = msg[:_CAUSE_MAX] + "…"
    return f"{type(exc).__name__}: {msg}"


def safe_tool(func):
    """MCP 도구 함수의 예외를 사용자 친화적 메시지로 변환합니다.

    아울러 라이선스 게이트를 적용한다 — 모든 도구가 이 래퍼를 거치므로,
    활성화되지 않은 경우 데이터 조회 없이 안내 메시지를 반환한다.
    """

    @functools.wraps(func)
    async def wrapper(*args, **kwargs):
        if not is_licensed():
            return locked_message()
        try:
            result = await func(*args, **kwargs)
        except httpx.TimeoutException as e:
            # ConnectError/HTTPError에는 원인을 붙였는데 이 분기만 빠뜨렸었다 —
            # 어떤 요청이 얼마나 걸리다 잘렸는지 알 수 없으면 똑같이 손을 못 쓴다.
            return (
                "⚠️ 네이버 증권 응답이 지연되고 있습니다. 잠시 후 다시 시도해주세요.\n"
                f"(원인: {_cause_text(e)})"
            )
        except httpx.ConnectError as e:
            # 원인을 같이 적는다. 예전엔 이 한 줄로 뭉갰는데, 실제 문의(2026-08-13,
            # 뉴질랜드 사용자)에서 연결 실패가 3시간 내내 100% 재현되는데도 원인을
            # 좁힐 수가 없었다 — 사용자가 볼 수 있는 것도, 로그에 남는 것도
            # "ConnectError" 뿐이라 이름 조회 실패인지·거부당한 건지·프록시인지
            # 구분이 안 됐다. 그 구분에 따라 사용자가 할 일이 완전히 다르다.
            return (
                "⚠️ 네이버 증권에 연결할 수 없습니다. 인터넷 연결을 확인해주세요.\n"
                f"(원인: {_cause_text(e)})"
            )
        except httpx.HTTPError as e:
            return f"⚠️ 네트워크 오류가 발생했습니다: {type(e).__name__}\n(원인: {_cause_text(e)})"
        except NaverParseError as e:
            # 아래 일반 분기는 "종목코드가 올바른지 확인해주세요"라고 안내하는데,
            # 파싱 실패에는 틀린 안내다 — 코드는 멀쩡하고 우리가 못 읽은 것이다.
            # 사용자가 종목명을 바꿔가며 재시도하게 만들면 안 된다.
            return (
                "⚠️ 네이버 증권 페이지를 읽지 못했습니다 "
                "(데이터가 없는 것이 아니라 **파싱 실패**입니다).\n"
                f"(원인: {e})\n"
                "페이지 구조가 바뀌었을 수 있습니다. StockLens 업데이트를 확인해 주세요."
                + _support_hint()
            )
        except Exception as e:
            return (
                f"⚠️ 데이터 처리 중 오류가 발생했습니다: {type(e).__name__}\n"
                f"종목코드가 올바른지, 상장된 종목인지 확인해주세요."
                + _support_hint()
            )
        try:
            notice = await get_update_notice()
        except Exception:
            notice = ""
        # 체험 만료가 다가오면 한 줄 덧붙인다. 사람들은 LeetKit Manager를 잘 안 열어서,
        # 그 화면에만 두면 기간이 끝나고 도구가 잠긴 뒤에야 알게 된다(trial_notice 참고).
        # pending_notice 는 표시 이력을 소진하므로 **한 번만** 부른다.
        try:
            import stock_mcp_server.trial_notice as _tn

            trial = _tn.pending_notice() or ""
        except Exception:
            trial = ""
        return _deliver_notices(result, [notice, trial])

    return wrapper

mcp = FastMCP(
    "StockLens",
    instructions="""StockLens — 한국 주식 데이터를 네이버 증권에서 실시간 조회합니다.
종목코드(예: 005930)나 종목명(예: 삼성전자)으로 검색할 수 있습니다.
차트 데이터, 투자자 수급, 재무지표, 시장 지수를 제공합니다.

## 🚨 종목코드 규칙 (절대 원칙)

사용자가 6자리 코드가 아닌 **종목명**을 주면, **반드시 아래 순서를 지켜라**:

1. 먼저 `search` 또는 `search_stock` 도구로 종목명 조회
2. 검색 결과가 여러 개면 사용자에게 확인 요청
3. 검색 결과 없으면 "종목명을 찾을 수 없습니다. 6자리 코드를 알려주세요" 요청
4. **종목 코드 추측 절대 금지** — 학습 지식으로 "아마 XXXXXX일 것" 식 추론 금지
5. 종목 코드는 상장폐지·합병·재할당으로 바뀔 수 있음. 반드시 실시간 조회로 검증.

이 규칙을 어기면 잘못된 종목 분석으로 사용자를 오도할 수 있다.

## 답변 형태 (출력 규칙)

도구가 돌려준 표·숫자는 **이미 사용자에게 보여줄 최종 형태**다. 그대로 인용하라.

1. **표/수치 재서술 금지.** 반환된 markdown 표를 산문으로 다시 풀어 설명하지 마라.
   ("RSI 72로, 보통 70 이상은 과매수를 뜻하며…" 식 X). 핵심 1~3줄 코멘트 + 표 인용으로 끝낸다.
2. **차트·SVG를 직접 그리지 마라.** OHLCV(`get_chart`)를 받아 손으로 SVG·ASCII·그래프를
   생성하는 것은 토큰 낭비이자 부정확하다. 수치는 반환된 표로 충분하다. 사용자가 굳이
   시각 차트를 원하면 임의로 그리지 말고 어떤 형태를 원하는지 먼저 확인하라.
3. 데이터에 없는 해석·전망을 학습지식으로 채우지 마라. 반환에 담긴 판정 라벨·수치 안에서만 말한다.
4. **단위·통화·기준을 바꿔 쓰지 마라.** 컬럼명과 `※` 주석에 단위(억원/원/주/%),
   통화(USD/TWD), 재무기준(IFRS연결/별도)이 명시돼 있다. 그 라벨을 떼고 숫자만
   옮기거나, 명시 안 된 단위를 추측해 붙이지 마라.
5. **`데이터 없음`·`추정`·`⚠️`는 그대로 전달하라.** "데이터 없음"을 0으로 바꾸지 말고,
   `단위: 억원 **추정**`이 붙은 값을 확정 수치로 쓰지 말고, `⚠️통화혼합`이 붙은 비율
   (P/B·P/S 등)을 정상 밸류에이션처럼 해석하지 마라.

## 🕐 결과 메타 봉투 (RESULT_META_JSON / `_meta`) — 시간축의 진실

대부분의 데이터 도구는 응답 끝에 `RESULT_META_JSON_START…END` 블록을(JSON 도구는
payload 안 `_meta` 키로) 붙인다. **날짜를 말하기 전에 반드시 이걸 먼저 보라.**

- `as_of` = **조회 시각**(내가 언제 물어봤나). `data_as_of` = **데이터 기준일**(이 숫자가
  언제 것인가). 토요일에 조회하면 as_of는 토요일, data_as_of는 직전 거래일이다.
  **사용자에게 말할 날짜는 언제나 `data_as_of`다.** as_of를 "오늘 시세"라고 쓰지 마라.
- `data_basis`가 판정의 확정도를 말한다:
  - `last_close` 확정치 — 그대로 서술 가능
  - `realtime` 장중 스냅샷 — 종가 아님. "현재" 라고만 하고 "종가"라 하지 마라
  - **`in_progress_bar` 장중 미완성 봉** — 이 봉으로 계산된 RSI·MACD·골든크로스·
    신고가 판정은 **마감 시 달라질 수 있다.** 확정 신호로 단정하지 말고 "장중 기준
    잠정"임을 반드시 함께 말하라
  - `filing` 공시 기반 / `aggregate` 수집 구간 집계
- `data_completeness`가 `partial`/`none`이면 없는 부분을 추정으로 메우지 마라.
- `warnings`는 요약에 반영하라. 비어 있지 않은데 무시하면 안 된다.
- `entity`의 `stock_code`·`corp_code`·`name`은 **다음 도구 호출에 그대로 재사용**하라
  (DartLens `corp_code`, StockLens 6자리 `code`). 코드를 다시 검색하거나 추측하지 마라.
- 메타 블록 자체는 내부용이다. **사용자에게 JSON을 그대로 보여주지 마라** — 필요한
  사실(기준일·경고)만 문장으로 옮긴다.
- 메타가 아예 없는 응답은 구버전 도구다. 그때는 기준일을 **미상**으로 다루고 추측하지 마라.

### 🌏 해외 사용자 (`viewer_tz`)

메타에 `viewer_tz`가 있으면 **사용자는 한국 밖에 있다.** 없으면 국내 사용자다.

- `viewer_tz.local_date`가 사용자의 '오늘'이다. `data_as_of`(한국 시장 기준)와
  다를 수 있다. LA 저녁이면 현지는 8/15인데 한국은 이미 8/16이다.
  이때 **"오늘 시세"라고 말하지 마라** — "한국 기준 8/14 종가"처럼 날짜를 박아 말한다.
- 시세·재무 날짜를 사용자 현지 날짜로 바꾸지 마라. 거래일은 한국 시장의 날짜다.
  바꾸지 말고, 사용자의 '오늘'과 다르다는 사실만 필요할 때 덧붙인다.
- 개장 시각을 묻거든 `get_market_clock`을 써라. 남은 시간(`opens_in`/`closes_in`)은
  어디서 보든 같고, `next_open_viewer`가 사용자 현지 시각으로 환산된 개장 시점이다.
  **날짜가 하루 다를 수 있으니** 현지 환산을 같이 알려주는 게 친절하다.

### ⏰ 서머타임 (`dst`)

미국장 개장은 한국시간 기준 **여름 22:30 / 겨울 23:30**으로 1시간 움직인다.
개장 시각을 학습지식으로 말하지 말고 `get_market_clock`의 값을 그대로 써라.

- `dst.us` / `dst.krx`의 `imminent`가 true면 전환이 2주 안이다. 개장 시각을
  안내할 때 **바뀔 예정이라는 사실을 함께 말하라** — `open_in_viewer_now` →
  `open_in_viewer_after`가 그 값이다.
- `cause`가 `viewer`면 시장이 아니라 **사용자 지역**이 전환하는 것이다. 한국장은
  서머타임이 없지만 시드니·런던 사용자는 자기 지역 전환으로 현지 개장 시각이 바뀐다.
- `krx.has_dst`가 false면 한국은 서머타임이 없다는 뜻이다. 있다고 가정하지 마라.

## ⭐ 관심종목 (`watchlist`)

사용자가 지정한 '내 종목' 목록. **DartLens·TelegramLens와 같은 목록을 공유**한다.

- "내 종목", "관심종목", "○○ 넣어줘/빼줘" → `watchlist` 도구를 쓴다.
- 조회 결과에 ⭐가 붙어 있으면 그 종목은 사용자의 관심종목이다. 언급해주면 좋다.
- **"내 종목 어때?" 류 요청**: `watchlist(action="list")`로 codes를 받아
  `get_multi_stocks`·`get_flow_batch`·`get_indicators_bulk`에 **한 번에** 넘긴다.
  종목마다 개별 호출하지 마라 (토큰·시간 낭비).
- 사용자가 특정 종목을 반복해서 물어보면 관심종목 등록을 한 번 권해도 좋다.
  다만 매번 권하지는 마라.

## 🔁 여러 종목을 다룰 때 — 같은 도구를 반복하지 마라

**같은 도구를 2번 이상 연달아 부르게 되면 멈추고 배치 도구를 쓴다.**
실측에서 한 사용자가 재무를 26종목·수급을 71종목에 대해 하나씩 불렀다.
느리고, 토큰을 몇 배로 쓰고, 화면이 도구 호출로 도배되고, 중간에 끊기면 처음부터다.

| 하나씩 부르지 말고 | 이걸로 |
|---|---|
| `get_price` × N | `get_multi_stocks` |
| `get_chart` × N | `get_multi_chart_stats` (기간 통계만) |
| `get_flow` × N | `get_flow_batch` |
| `get_financial` × N | `get_financial_batch` |
| `get_indicators` × N | `get_indicators_bulk` |
| `get_us_price` × N | `get_us_multi_price` |

랭킹·테마·스크리닝 결과로 종목 목록을 얻었다면 **그 목록을 통째로** 배치에 넘긴다.
5종목 비교 실측: 단건 15회 7,367토큰·3.1초 → 배치 3회 1,441토큰·1.0초.

## 도구 역할 구분
- `get_chart`: **OHLCV 시계열 데이터** (수치 분석·요약용, count 최소 120)
- `get_indicators`: **기술지표 판정값** (RSI/MACD/Phase 등 숫자+라벨)
- `get_price`: **현재가 스냅샷** (단일 시점, 시계열 아님)
- `get_event_reaction`: **event_date 기준 주가·수급 반응**. DartLens 공시 접수일을
  event_date로 넘겨 공시 전후 거래일 반응을 확인할 때 사용.
  응답의 **검증 상태**를 먼저 보라 — `unavailable`이면 숫자가 아예 없다(보유 데이터 구간
  밖 사건·전 구간 거래정지). 이때 "반응 없음/변동 없음"으로 서술하지 말고 측정 불가라고
  말하라. `—` 표기와 `해당 사건창 수급 데이터 없음`은 0이 아니라 데이터 없음이다.

## ETF 도구
- `get_etf_list`: ETF 목록 조회 + 카테고리 필터 + 이름 키워드 필터 + 정렬
- `get_etf_info`: ETF 상세 정보 (기초지수, 보수율, 구성종목, 수익률)
- ETF도 일반 종목 도구(`search`, `get_price`, `get_chart`)로 조회 가능

## 🔍 테마·컨셉으로 종목/ETF 찾기 (발견 vs 검증 분리)

`search`/`get_etf_list`는 **이름을 아는 대상을 정확히 조회**하는 도구이지, "이런 조건에
맞는 걸 찾아줘" 같은 발견(discovery)에는 한계가 있다. 특정 회사(특히 비상장사) 지분을
보유한 ETF를 찾는 것처럼, 도구 검색만으로 후보가 안 나오는 질문은:

1. 학습 지식이나 웹서치로 후보 종목·ETF **이름**을 먼저 좁혀라(추측 아님 — 실제 조사).
2. 후보가 나오면 **반드시** `get_etf_info`/`get_price`/`get_financial` 등으로 넘어가서
   수수료·구성종목·가격 같은 실제 숫자를 stocklens로 재검증하고 그 숫자로 답하라.
   웹서치에서 본 숫자를 그대로 답에 쓰지 마라 — 오래됐거나 부정확할 수 있다.
3. `get_etf_list(keyword="...")`로 먼저 시도해볼 것 — 한국 ETF는 이름 자체가 테마를
   직설적으로 담는 관행이라(예: "TIGER 미국우주테크") 카테고리를 몰라도 키워드 하나로
   꽤 많은 경우가 해결된다.
4. keyword는 이름 부분일치라, 1~2글자처럼 짧거나 흔한 키워드는 무관한 결과가 섞인다
   (예: "금" → 금현물 ETF 몇 개 말고 "CD금리액티브"·"금융채" 같은 무관한 상품이 다수
   섞여 나옴 — "금리"·"금융"에도 "금"이 들어있어서). 결과 이름들을 훑어서 의도한
   테마와 안 맞는 게 섞여 있으면 그대로 쓰지 말고, 종목명 검색 결과가 여러 개일 때와
   같은 원칙으로 사용자에게 "이 중 어떤 뜻으로 물으신 건가요?" 하고 되물어라.

## 분석 도구
- `get_consensus`: 컨센서스 — 증권사 투자의견·목표주가 + **어닝 서프라이즈**
  (영업이익·당기순이익의 컨센서스 vs 잠정치). **매출액은 이 도구에 없다** — 매출은
  `get_financial`(네이버 실적표) 또는 DartLens `get_major_accounts`(공시 원본)로.
- `get_reports`: 증권사 리포트 — 목록 + 본문 요약 + PDF 링크 (조회수 포함)
- `get_sector_valuation`: 업종·테마의 PER·PBR·ROE **집계**와 종목별 할증/할인 위치.
  개별 지표만으로는 알 수 없는 **비교 기준**을 만든다. 중앙값이 기준이며
  적자(PER 음수)는 집계에서 빼고 건수를 따로 알린다.
- `get_disclosure`: 공시 목록 — DART 전자공시 제목/날짜
- `get_event_reactions`: **공시 반응 이력** — 최근 공시들에 이 종목이 어떻게 반응해 왔는지
  한 번에. 일봉·수급을 1회만 받아 여러 공시일을 계산한다(get_event_reaction 을 N번
  부르지 말 것). 반응이 최근 약해지는지도 함께 판정한다.
- DartLens가 함께 설치되어 있으면: StockLens 주가·수급 이상 구간 → DartLens 해당 기간 공시 확인,
  DartLens 실적/공시 결과 → StockLens `get_event_reaction(code, event_date=공시일)` 순서로 이어가기

## 🇺🇸 US 주식 도구 (NYSE / NASDAQ · yfinance)

**티커 구분 규칙:**
- 6자리 영숫자 (예: `005930`) → 한국 주식 → `get_price`, `get_chart` 등
- 1~5자 알파벳 (예: `AAPL`, `TSLA`, `BRK.B`) → 미국 주식 → **`get_us_*` 도구 사용**

사용자가 "애플", "테슬라" 같은 한국어·회사명만 말하면 **추측 금지**. `get_us_search`로
먼저 검색해 티커를 확정한 뒤 다른 tool을 호출하세요.

### 탐색
- `get_us_search`: 종목명·티커 검색 (예: "Apple" → AAPL)
- `get_us_market`: 주요 지수 (S&P 500, Dow, Nasdaq, Russell 2000, VIX)
- `get_us_screener`: 프리셋 스크리너 (day_gainers/losers/most_actives/undervalued 등)
- `get_us_sector`: 섹터별 overview + top 20 기업

### 기본 데이터
- `get_us_price`: 현재가 + 전일대비 + 52주 고저 + 베타 + 시총
- `get_us_info`: 기업정보 (섹터·산업·사업요약)
- `get_us_chart`: OHLCV 시계열 (period/interval/prepost)
- `get_us_financials`: Valuation (P/E, PEG, P/B) + Profitability + Dividend 비율
- `get_us_financial_statement`: 재무제표 (income/balance/cash_flow, annual/quarterly)
- `get_us_multi_price`: 여러 티커 일괄 가격 스냅샷

### US 고유 정보
- `get_us_earnings`: 다음 실적발표일 + 최근 서프라이즈 이력
- `get_us_analyst`: 목표주가 + buy/hold/sell + 업·다운그레이드 + EPS/매출 추정치
- `get_us_dividends`: 배당 이력 + ex-date + yield + payout ratio
- `get_us_options`: 옵션 체인 (calls/puts, IV, OI) — Greeks 미포함
- `get_us_insider`: Form 4 내부자 거래 + 현재 내부자 명단
- `get_us_holders`: 기관 (13F) + 뮤추얼펀드 + 주주 비중 요약
- `get_us_short`: 공매도 % of float + days to cover (2~4주 stale 경고)
- `get_us_filings`: SEC 공시 목록 (10-K, 10-Q, 8-K) + EDGAR URL
- `get_us_news`: 최근 뉴스 헤드라인
- `get_us_etf_info`: ETF 전용 상세 (top holdings, 섹터 비중, 자산 배분)

⚠️ **데이터 제약:** Yahoo Finance는 최대 15분 지연. 실시간 호가창·다크풀 미지원.
프리/포스트 마켓은 `get_us_chart(prepost=True)` 사용.
""",
)

@mcp.tool()
@safe_tool
@track_metrics("get_market_clock")
async def get_market_clock() -> str:
    """한국장/미국장 현재 상태, 휴장 여부, 최근/다음 거래일을 한 번에 조회합니다.

    종목 분석 전 데이터 기준시각을 확인할 때 사용합니다. KRX와 NYSE/NASDAQ의
    주말, 정규 휴장일, 장전/정규장/시간외/장마감 상태를 함께 반환합니다.
    """
    return format_market_clock(build_market_clock())


@mcp.tool()
@track_metrics("stocklens_status")
async def stocklens_status() -> str:
    """상태요약 — 버전/라이선스/국내·미국 시장 상태/최근 성공·실패/캐시 쓰기 가능 여부를 한 번에.

    문제가 있는지 대화 중 가볍게 확인할 때 사용합니다. 이미 기록된 최근 호출 로그와
    업데이트 확인 캐시만 읽으므로 네트워크를 새로 호출하지 않습니다(빠름).
    라이선스 미활성 상태에서도 원인 파악용으로 동작해야 하므로 다른 도구와 달리
    라이선스 게이트를 걸지 않습니다. 더 깊은 진단(오프라인 재현 등)은 터미널의
    stocklens-doctor 커맨드를 안내하세요.
    """
    try:
        return format_status(build_status())
    except Exception as e:
        return f"⚠️ 상태 조회 중 오류: {type(e).__name__}. 터미널에서 stocklens-doctor로 더 자세히 진단해보세요."


def _normalize_date(raw, is_intraday: bool) -> str:
    """네이버 YYYYMMDD, yfinance ISO 등 이종 포맷을 YYYY-MM-DD(·HH:MM)로 통일."""
    s = str(raw or "")
    if len(s) == 8 and s.isdigit():  # 네이버 KR: "20260407"
        return f"{s[:4]}-{s[4:6]}-{s[6:8]}"
    return s[:16] if is_intraday else s[:10]


def _format_ohlcv_rows(
    rows: list[dict],
    *,
    is_intraday: bool,
    decimals: int,
) -> list[str]:
    """OHLCV 시계열을 markdown table 행으로 포맷.

    토큰 절감 포맷: 파이프 주변 공백 제거, 시간부 생략(일봉+), KR 날짜 대시화,
    거래량 쉼표 제거(정밀도 100% 보존). 첫 줄(header)은 호출자가 앞뒤로 붙이는 것을 가정.
    """
    price_fmt = f"{{:.{decimals}f}}"

    def _fmt(r: dict) -> tuple[str, str, str, str, str, str]:
        date = _normalize_date(r.get("date") or r.get("datetime"), is_intraday)
        o = price_fmt.format(float(r.get("open") or 0))
        h = price_fmt.format(float(r.get("high") or 0))
        lo = price_fmt.format(float(r.get("low") or 0))
        c = price_fmt.format(float(r.get("close") or 0))
        v = str(int(r.get("volume") or 0))
        return (date, o, h, lo, c, v)

    out = [
        "날짜|시가|고가|저가|종가|거래량",
        "---|---|---|---|---|---",
    ]
    out.extend("|".join(_fmt(r)) for r in rows)
    return out


async def _search_impl(query: str) -> str:
    """공통 구현 — search와 search_stock 도구에서 공유."""
    results = await naver_search_stock(query)
    if not results:
        return f"'{query}'에 대한 검색 결과가 없습니다. 6자리 코드를 직접 알려주세요."

    lines = [f"검색 결과 ({len(results)}건):"]
    for r in results:
        market = r.get("market", "")
        suffix = f" [{market}]" if market else ""
        lines.append(f"  - {r['name']} ({r['code']}){suffix}")
    return "\n".join(lines)


@mcp.tool()
@safe_tool
@track_metrics("search")
async def search(query: str) -> str:
    """종목검색 — 종목명 또는 종목코드로 한국 주식 종목을 조회합니다.
    "삼성전자 종목코드", "반도체 관련주", "005930 뭐야" 같은 질문에 사용합니다.

    ⚠️ 사용자가 종목명만 주고 코드를 모를 때 **반드시 이 도구를 먼저 호출**.
    종목 코드를 추측하지 말 것. 상장폐지·재할당으로 코드가 바뀔 수 있음.

    search_stock으로도 동일하게 호출 가능 (별명).

    Args:
        query: 검색할 종목명 또는 코드 (예: "삼성전자", "005930", "알멕")
    """
    return await _search_impl(query)


@mcp.tool()
@safe_tool
@track_metrics("search_stock")
async def search_stock(query: str) -> str:
    """종목코드조회 (stock lookup) — 한국 주식 종목명/코드 조회 전용 도구.

    `search`와 동일 기능. 도구 디스커버리에서 "stock"/"ticker"/"종목" 키워드로
    빠르게 매칭되도록 명확한 이름을 갖는 별명입니다.

    ⚠️ 종목명만 있고 6자리 코드를 모를 때 **이 도구를 먼저 호출**해야 합니다.
    코드 추측(guessing) 금지. 다른 도구(get_price, get_chart 등)에 잘못된 코드를
    넣으면 엉뚱한 종목이 조회됩니다.

    Args:
        query: 종목명(한/영) 또는 6자리 코드. 예: "알멕", "Samsung", "005930"

    Returns:
        매칭된 종목 리스트. 여러 개면 사용자에게 확인 요청 필요.
    """
    return await _search_impl(query)


@mcp.tool()
@safe_tool
@track_metrics("watchlist")
async def watchlist(action: str = "list", query: str = "") -> str:
    """관심종목 — '내 종목' 목록 보기/추가/삭제 (LeetKit 세 Lens 공용).

    "내 종목", "관심종목", "○○ 관심종목에 넣어줘", "내 종목 시세 보여줘" 같은
    요청에 사용합니다. DartLens·TelegramLens와 **같은 목록**을 공유합니다.

    `list` 결과의 `codes`를 그대로 `get_multi_stocks` / `get_flow_batch` /
    `get_indicators_bulk`에 넘기면 관심종목 전체를 한 번에 볼 수 있습니다.
    종목을 하나씩 다시 조회하지 마세요.

    Args:
        action: "list"(기본) / "add" / "remove" / "clear"
        query: add·remove에 쓸 종목명 또는 6자리 코드 (예: "디오", "039840")
    """
    action = (action or "list").strip().lower()
    if action not in ("list", "add", "remove", "clear"):
        return f"⚠️ action은 list/add/remove/clear 중 하나여야 합니다 (받음: '{action}')."

    if action == "add":
        if not query.strip():
            return "⚠️ 추가할 종목명이나 코드를 알려주세요. 예: watchlist(action='add', query='디오')"
        # 코드를 추측하지 않는다 — 검색으로 확정한다.
        q = query.strip()
        found = await naver_search_stock(q)
        if not found:
            return f"'{q}' 종목을 찾을 수 없습니다. 6자리 코드로 다시 시도해주세요."

        # 네이버 자동완성은 부분일치도 함께 준다("디오" → 디오·디와이파워·디와이…).
        # **정확히 일치하는 이름이나 코드가 있으면 그건 추측이 아니라 확정이다.**
        # 그때까지 되물으면 매번 한 번 더 묻게 되어 쓸모가 떨어진다.
        exact = [r for r in found if r["code"] == q.upper() or r["name"] == q]
        if len(exact) == 1:
            hit = exact[0]
        elif len(found) == 1:
            hit = found[0]
        else:
            cands = exact if len(exact) > 1 else found
            names = ", ".join(f"{r['name']}({r['code']})" for r in cands[:5])
            return f"'{q}'에 해당하는 종목이 여러 개입니다. 어느 것인가요? — {names}"
        ok, why = wl.add(hit["code"], hit["name"])
        head = (f"✅ 관심종목에 추가했습니다: {hit['name']} ({hit['code']})"
                if ok else f"ℹ️ {hit['name']} ({hit['code']}) — {why}")
    elif action == "remove":
        if not query.strip():
            return "⚠️ 삭제할 종목명이나 코드를 알려주세요."
        gone = wl.remove(query.strip())
        head = (f"🗑️ 관심종목에서 뺐습니다: {gone['name']} ({gone['code']})"
                if gone else f"'{query}'는 관심종목에 없습니다.")
    elif action == "clear":
        head = f"🗑️ 관심종목 {wl.clear()}개를 모두 비웠습니다."
    else:
        head = None

    stocks = wl.load()
    lines = [head] if head else []
    if not stocks:
        lines += ["", "관심종목이 비어 있습니다.",
                  "예: \"디오를 관심종목에 넣어줘\" 또는 watchlist(action='add', query='디오')"]
        return _append_result_meta("\n".join(lines).strip(),
                                   _kr_meta(kind="snapshot", data_completeness=rmeta.NONE))

    lines += ["", f"## ⭐ 내 관심종목 ({len(stocks)}개)", ""]
    lines.append("| # | 종목명 | 코드 |")
    lines.append("|---:|---|---|")
    for i, s in enumerate(stocks, 1):
        lines.append(f"| {i} | {s['name']} | {s['code']} |")
    lines.append("")
    lines.append("codes: " + ", ".join(s["code"] for s in stocks))
    lines.append("_시세·수급을 한 번에 보려면 위 codes를 get_multi_stocks / "
                 "get_flow_batch 에 그대로 넘기세요._")
    return _append_result_meta("\n".join(lines).strip(), _kr_meta(kind="snapshot"))


@mcp.tool()
@safe_tool
@track_metrics("get_chart")
async def get_chart(
    code: str,
    timeframe: str = "day",
    count: int = 120,
) -> str:
    """캔들차트 OHLCV — 종목의 시계열 캔들 데이터(시가/고가/저가/종가/거래량, candlestick OHLCV).

    ⚠️ **여러 종목의 기간 통계만 필요하면 `get_multi_chart_stats`를 쓰세요.**
    "삼성전자 일봉", "3개월 주봉", "월봉 데이터", "price history" 같은 질문에 사용. 시계열 진입점(US는 get_us_chart).

    보조지표(이평선·RSI·MACD 등)는 사용자가 명시 요청할 때만 get_indicators로 숫자만 받아 요약.

    Args:
        code: 종목코드 6자리 (예: "005930")
        timeframe: "day"(일봉), "week"(주봉), "month"(월봉)
        count: 가져올 봉 개수 (기본 120 ≈ 6개월, 최대 500)
    """
    count = min(count, 500)
    raw = await get_ohlcv(code, timeframe, count)
    data, _excluded = split_valid_bars(raw)
    exclusion_info, exclusion_warns = _bar_exclusion_info(_excluded)
    if not data:
        meta = _kr_meta(
            kind="bars", code=code, data_completeness=rmeta.NONE,
            warnings=["차트 데이터를 가져오지 못했습니다."],
        )
        return _append_result_meta(f"종목코드 {code}의 차트 데이터를 가져올 수 없습니다.", meta)

    tf_name = {"day": "일봉", "week": "주봉", "month": "월봉"}.get(timeframe, timeframe)
    header = f"종목 {code} {tf_name} OHLCV ({len(data)}개 봉)"
    lines = [header, ""]
    lines.extend(_format_ohlcv_rows(data, is_intraday=False, decimals=0))

    if exclusion_info:
        lines.append("")
        lines.append(
            f"※ 거래정지 placeholder 등 비정상 봉 {exclusion_info['count']}개 제외"
            f" ({', '.join(b['date'] for b in exclusion_info['bars'][:5])})."
        )
    warnings = list(exclusion_warns)
    completeness = rmeta.COMPLETE
    if len(data) < count:
        completeness = rmeta.PARTIAL
        warnings.append(f"요청한 {count}개 중 {len(data)}개만 조회됨(상장일 등으로 이력이 짧을 수 있음).")
    # 기준일은 시장 캘린더 역산이 아니라 **실제 마지막 봉**에서 뽑는다.
    meta = _kr_meta(
        kind="bars", code=code, data_as_of=data[-1].get("date"),
        data_completeness=completeness, warnings=warnings,
        bar_state=_bar_state(
            timeframe=timeframe, rows=data, market_calendar=krx_calendar()
        ),
        price_adjustment=price_adjustment_meta(),
        extra={"bar_exclusions": exclusion_info} if exclusion_info else None,
    )
    return _append_result_meta("\n".join(lines), meta)


# --- get_chart_html 비활성화 (v0.2.5~) ---
# 차트 HTML 렌더링 도구. 필요 시 주석 해제.
#
# @mcp.tool()
# @safe_tool
# @track_metrics("get_chart_html")
# async def get_chart_html(
#     code: str,
#     timeframe: str = "day",
#     count: int = 120,
#     timeframes: list[str] | None = None,
#     show_sr: bool = True,
#     custom_sr: list[dict] | None = None,
# ) -> str:
#     """차트HTML — 완성된 캔들+거래량 차트 HTML 반환."""
#     info = await get_current_price(code)
#     name = info.get("name", code) if info else code
#     if timeframes:
#         frame_defaults = {"day": 120, "week": 52, "month": 24}
#         frames = []
#         for tf in timeframes:
#             c = frame_defaults.get(tf, 120)
#             ohlcv = await get_ohlcv(code, tf, c)
#             if ohlcv:
#                 frames.append({"timeframe": tf, "ohlcv": ohlcv})
#         if not frames:
#             return f"종목코드 {code}의 차트 데이터를 가져올 수 없습니다."
#         html = render_multi_chart_html(code, name, frames, show_sr=show_sr, custom_sr=custom_sr)
#     else:
#         count = min(count, 500)
#         ohlcv = await get_ohlcv(code, timeframe, count)
#         if not ohlcv:
#             return f"종목코드 {code}의 차트 데이터를 가져올 수 없습니다."
#         html = render_chart_html(
#             code, name, ohlcv,
#             timeframe=timeframe, show_sr=show_sr, custom_sr=custom_sr,
#         )
#     MAX_INLINE = 60_000
#     if len(html) <= MAX_INLINE:
#         return html
#     chart_dir = get_snapshot_dir() / "charts"
#     chart_dir.mkdir(parents=True, exist_ok=True)
#     tf_label = "-".join(timeframes) if timeframes else timeframe
#     fname = chart_dir / f"{code}_{tf_label}.html"
#     fname.write_text(html, encoding="utf-8")
#     size_kb = len(html) / 1024
#     return (
#         f"✓ 차트 HTML 생성 완료 ({size_kb:.0f}KB — 멀티프레임이라 파일로 저장)\n"
#         f"경로: {fname}\n\n"
#         f"사용자에게 이 경로를 안내해주세요. 브라우저에서 직접 열면 "
#         f"캔들차트 + 거래량 + S/R 오버레이가 표시됩니다."
#     )


_append_result_meta = rmeta.append_meta


def _us_meta(
    *,
    kind: str,
    ticker: str | None = None,
    data_as_of=None,
    data_completeness: str = rmeta.COMPLETE,
    coverage: dict | None = None,
    bar_state: dict | None = None,
    price_adjustment: dict | None = None,
    extra: dict | None = None,
    warnings: list[str] | None = None,
) -> dict:
    """US 도구용 메타. yfinance는 15분 지연 공지가 있어 is_delayed를 세운다."""
    us_state = build_market_clock()["us"]
    if kind == "filing":
        basis = rmeta.BASIS_FILING
    elif us_state.get("is_open"):
        basis = rmeta.BASIS_IN_PROGRESS_BAR if kind == "bars" else rmeta.BASIS_REALTIME
    else:
        basis = rmeta.BASIS_LAST_CLOSE

    warns = list(warnings or [])
    if kind != "filing" and not us_state.get("is_open"):
        warns.insert(
            0,
            f"미국장 마감 상태 — 표시된 값은 최근 거래일({us_state.get('last_trading_day')}) 기준입니다.",
        )

    data_completeness, coverage, warns = _bar_state_effects(
        bar_state, data_completeness, coverage, warns
    )

    meta = rmeta.build_meta(
        lens="stocklens",
        data_basis=basis,
        data_as_of=data_as_of or (None if kind == "filing" else us_state.get("last_trading_day")),
        market="US",
        session=us_state.get("status"),
        is_delayed=bool(us_state.get("is_open")),  # yfinance 실시간은 지연 시세
        data_completeness=data_completeness,
        coverage=coverage,
        entity_info=rmeta.entity(stock_code=ticker),
        warnings=warns,
    )
    if bar_state:
        meta["bar_state"] = bar_state
    if price_adjustment:
        meta["price_adjustment"] = price_adjustment
    for key, value in (extra or {}).items():
        meta[key] = value
    return meta


def _deliver_notices(result, notices: list[str]) -> str:
    """업데이트·체험 안내를 결과에 싣는다.

    JSON 도구(get_indicators 등)는 응답 전체가 json.loads 대상이라, 뒤에 텍스트를
    붙이면 Extra data 로 깨진다(실측 2026-08-27, get_indicators_bulk). JSON 이면
    _meta.warnings 안에 넣고, 텍스트면 예전처럼 뒤에 덧붙인다.
    """
    if not isinstance(result, str):
        return result
    notes = [n.strip() for n in notices if n and n.strip()]
    if not notes:
        return result

    if result.lstrip().startswith("{"):
        try:
            payload = json.loads(result)
        except ValueError:
            payload = None
        if isinstance(payload, dict):
            meta = payload.get("_meta")
            if not isinstance(meta, dict):
                meta = {}
                payload["_meta"] = meta
            warns = meta.get("warnings")
            if not isinstance(warns, list):
                warns = []
                meta["warnings"] = warns
            for n in notes:
                # 텍스트용 구분선("---")은 warnings 안에서는 노이즈다.
                lines = [ln for ln in n.splitlines() if ln.strip() and ln.strip() != "---"]
                warns.append(" ".join(ln.strip() for ln in lines))
            return json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
        # JSON 처럼 생겼는데 파싱이 안 되면 원문을 건드리지 않는다.
        return result

    out = result
    for n in notes:
        out = f"{out}\n\n{n}"
    return out


def _kr_market_note(krx: dict | None = None) -> list[str]:
    """KRX가 지금 닫혀 있으면, 보여주는 값이 실시간이 아니라 최근 종가 기준임을 명시.

    Naver 소스는 공식적으로 지연을 공지하지 않으므로(US Yahoo와 달리) is_delayed는
    건드리지 않고, '휴장 중 최근 종가' 사실만 warning으로 남긴다.
    """
    krx = krx or build_market_clock()["krx"]
    if krx["is_open"]:
        return []
    return [f"KRX 장마감 상태 — 표시된 값은 최근 거래일({krx['last_trading_day']}) 기준입니다."]


def _emit_coverage_counters(lens: str, meta: dict) -> None:
    """메타에 이미 담긴 한계 사실을 카운터로 옮긴다.

    조건을 다시 계산하지 않는다. 응답에 실린 것과 세는 것이 어긋나면 지표를
    믿을 수 없다. 라벨은 고정 집합이라 종목코드·검색어는 들어가지 않는다.
    """
    tool = _metrics_mod.current_tool()
    if not tool:
        return
    try:
        cov = meta.get("coverage") or {}
        if cov.get("truncated") is True:
            _metrics_mod.count_limitation(
                "lens_coverage_truncated_total",
                lens=lens, tool=tool, reason=str(cov.get("reason") or "unknown"),
            )
        bar = meta.get("bar_state") or {}
        if bar.get("calculation_includes_incomplete") is True:
            _metrics_mod.count_limitation(
                "lens_incomplete_bar_total",
                tool=tool, timeframe=str(bar.get("timeframe") or "unknown"),
            )
        if (meta.get("period_coverage") or {}).get("consistency") == "mixed":
            _metrics_mod.count_limitation("lens_mixed_period_total", tool=tool)
        if (meta.get("price_adjustment") or {}).get("status") == "unknown":
            _metrics_mod.count_limitation("lens_unknown_adjustment_total", tool=tool)
    except Exception:
        # 지표 때문에 도구 응답이 죽으면 안 된다.
        pass


_NEWS_COMPARISON_RE = re.compile(
    r"\b(vs\.?|versus|better buy|compared to|comparison)\b", re.I)
_NEWS_SUPPLY_RE = re.compile(
    r"\b(supplier|supply chain|customer of|partner of|contract manufacturer)\b", re.I)
_NEWS_LIST_RE = re.compile(
    r"\b(etf|index fund|\d+\s+(best|top)\b|(best|top)\s+\d+|stocks to (buy|watch)|"
    r"our list|watchlist)\b", re.I)


def _classify_news_relevance(item: dict, names: tuple[str, ...]) -> dict:
    """기사에서 요청 종목의 역할을 평가한다(SL-10).

    실측(AAPL 피드): InterDigital·NVDA·Sandisk 가 중심인 기사가 그대로 섞여
    온다. yfinance 새 포맷에는 relatedTickers 도 없어 제목·요약 텍스트로
    판정한다. 점수는 확률이 아니라 "제목에 있나 / 요약에만 있나 / 없나"의
    순서 척도이고, 근거(basis)를 항상 같이 돌려준다.
    """
    title = (item.get("title") or "")
    summary = (item.get("summary") or "")
    hay_title = title.lower()
    hay_sum = summary.lower()
    keys = [n.lower() for n in names if n]

    in_title = any(k in hay_title for k in keys)
    in_summary = any(k in hay_sum for k in keys)

    if not in_title and not in_summary:
        return {"category": "unrelated", "relevance_score": 10,
                "basis": "제목·요약 어디에도 종목명이 없음"}

    combined = f"{hay_title} {hay_sum}"
    if _NEWS_COMPARISON_RE.search(hay_title) or (
            in_title and _NEWS_COMPARISON_RE.search(combined)):
        return {"category": "comparison", "relevance_score": 55,
                "basis": "비교 기사 - 이 종목이 단독 주제가 아님"}
    if _NEWS_SUPPLY_RE.search(combined) and in_title:
        return {"category": "supply_chain", "relevance_score": 50,
                "basis": "공급망 문맥 - 주체는 거래 상대일 수 있음"}
    if _NEWS_LIST_RE.search(combined):
        return {"category": "list_or_etf", "relevance_score": 40,
                "basis": "종목 나열·ETF 문맥 - 개별 재료가 아님"}
    if in_title:
        return {"category": "direct", "relevance_score": 90,
                "basis": "제목에 종목명 - 기사 주제"}
    return {"category": "mention", "relevance_score": 35,
            "basis": "요약에만 언급 - 주변 문맥"}


def _earnings_session(timestamp: str) -> str:
    """발표 ET 시각 -> 장전/장중/장후. 자정 정각은 날짜만 있는 것으로 본다.

    반응창 선택이 여기 걸린다: 장전 발표는 당일 봉이, 장후 발표는 다음
    거래일 봉이 첫 반응이다(SL-05 요구 4).
    """
    try:
        ts = _dt.datetime.fromisoformat(timestamp)
    except (TypeError, ValueError):
        return "unknown"
    t = ts.time()
    if t == _dt.time(0, 0):
        return "unknown"
    if t < _dt.time(9, 30):
        return "pre_market"
    if t >= _dt.time(16, 0):
        return "post_market"
    return "during_market"


def _match_fiscal_period(event_date: str, periods: list[str]) -> str | None:
    """발표일 직전 120일 안에서 가장 가까운 분기말을 회계기간으로 잡는다."""
    try:
        ed = _dt.date.fromisoformat(event_date)
    except (TypeError, ValueError):
        return None
    best = None
    for pd_ in periods:
        try:
            d = _dt.date.fromisoformat(pd_)
        except ValueError:
            continue
        gap = (ed - d).days
        if 0 <= gap <= 120 and (best is None or gap < best[0]):
            best = (gap, pd_)
    return best[1] if best else None


def _classify_insider_tx(row: dict) -> str:
    """Form 4 거래 한 건의 경제적 유형(SL-15).

    실측(XOM): 무상 부여·증여가 제공자 요약의 Purchases 에 섞여, 실제 현금
    매수 0건·매도 6건인 종목이 "내부자 대량 순매수"로 읽혔다. 대금이 실제로
    오간 거래와 주식만 움직인 거래를 같은 이름으로 합치지 않는다.
    """
    action = ((row.get("transaction") or "") or (row.get("text") or "")).lower()
    value = row.get("value")
    has_cash = isinstance(value, (int, float)) and value > 0
    if "gift" in action:
        return "gift"
    if "award" in action or "grant" in action:
        return "grant"
    if "tax" in action or "payment of exercise" in action:
        return "tax_withholding"
    if "exercise" in action or "conversion" in action:
        return "exercise"
    if "purchase" in action or "buy" in action:
        # 대금 없는 "Purchase" 는 공개시장 매수가 아니다 - 부여·오기재일 수 있다.
        return "open_market_buy" if has_cash else "unknown"
    if "sale" in action or "sell" in action:
        return "sale"
    return "unknown"


def _classify_us_security(q: dict) -> str:
    """스크리너 행의 증권 유형 판정(SL-12).

    워런트도 Yahoo quoteType 은 EQUITY 라 유형 필드만으로는 못 거른다
    (실측: GRABW·KLXER 가 small_cap_gainers 에 보통주처럼 섞여 나옴).
    이름 키워드가 가장 확실하고, 없으면 나스닥 5글자 접미사 규약
    (W=워런트, R=라이츠, U=유닛)을 쓴다. 아무 근거도 없으면 unknown 이다 -
    모르는 것을 보통주로 승격하지 않는다.
    """
    name = (q.get("name") or "").lower()
    if "warrant" in name:
        return "warrant"
    if "right" in name and "copyright" not in name:
        return "right"
    if name.endswith(" unit") or name.endswith(" units") or " units " in name:
        return "unit"
    if "american depositary" in name or " adr" in f" {name}" or name.endswith(" adr"):
        return "adr"

    qt = (q.get("quote_type") or "").upper()
    if qt == "ETF":
        return "etf"
    if qt in ("MUTUALFUND", "MONEYMARKET"):
        return "fund"
    if qt in ("INDEX", "CRYPTOCURRENCY", "FUTURE", "CURRENCY"):
        return qt.lower()

    symbol = (q.get("symbol") or "").upper()
    if qt == "EQUITY" and len(symbol) == 5 and symbol[-1] in ("W", "R", "U")             and symbol.isalpha():
        return {"W": "warrant", "R": "right", "U": "unit"}[symbol[-1]]
    if qt == "EQUITY":
        return "common_stock"
    return "unknown"


# 업종 밸류에이션에서 한 번에 재무를 긁을 최대 종목 수. KRX 업종은 대부분
# 이 안에 들어온다(최대 반도체 172개). 넘으면 표본 집계임을 명시한다(SL-11).
_SECTOR_AGG_CAP = 300


def _period_coverage_of(periods_by_entity: dict[str, str]) -> dict:
    """종목별 최신 기간에서 uniform/mixed/unknown 을 판정한다.

    KR 재무 배치가 쓰는 판정이고, 미국 재무 배치가 생기면 같은 헬퍼를 쓴다
    (SL-03). 서로 다른 분기의 값을 나란히 놓는 순간 비교가 아니게 되는 것은
    시장이 어디든 같다.
    """
    clean = {k: v for k, v in (periods_by_entity or {}).items() if v}
    uniq = sorted(set(clean.values()))
    return {
        "consistency": "unknown" if not uniq else ("uniform" if len(uniq) == 1 else "mixed"),
        "periods": clean,
        "oldest": uniq[0] if uniq else None,
        "newest": uniq[-1] if uniq else None,
    }


def _bar_state_effects(
    bar_state: dict | None,
    data_completeness: str,
    coverage: dict | None,
    warns: list[str],
) -> tuple[str, dict | None, list[str]]:
    """봉 상태를 완전성·coverage·경고에 반영한다. KR·US 메타가 같이 쓴다.

    봉 구간이 안 끝났다는 사실은 세션 상태와 별개다. 수요일 저녁이면 장은
    닫혔지만(basis=last_close) 그 주 주봉은 금요일까지 남아 있다.
    """
    if not bar_state:
        return data_completeness, coverage, warns
    tf_label = {"day": "일봉", "week": "주봉", "month": "월봉"}.get(
        bar_state.get("timeframe"), bar_state.get("timeframe") or "봉"
    )
    if bar_state.get("calculation_includes_incomplete") is True:
        done = bar_state.get("last_completed_bar_date")
        warns.append(
            f"마지막 {tf_label}({bar_state.get('last_bar_date')})이 아직 진행 중입니다"
            + (f". 확정된 마지막 {tf_label}은 {done}입니다." if done else ".")
            + " 이 봉이 포함된 지표·낙폭·신고가 판정은 구간이 끝나면 달라질 수 있습니다."
        )
        data_completeness = rmeta.PARTIAL
        coverage = dict(coverage or {})
        coverage.setdefault("truncated", False)
        coverage["coverage_complete"] = False
        coverage["reason"] = coverage.get("reason") or "incomplete_tail"
    elif bar_state.get("last_bar_complete") is None:
        warns.append(
            f"마지막 {tf_label}의 마감 여부를 확인하지 못했습니다. "
            "이 봉이 확정치인지 진행 중인지 알 수 없습니다."
        )
        # 모르는 것을 확정치로 내보내지 않는다. 판별 실패는 partial 이다.
        data_completeness = rmeta.PARTIAL
        coverage = dict(coverage or {})
        coverage.setdefault("truncated", False)
        coverage["coverage_complete"] = False
        coverage["reason"] = coverage.get("reason") or "unknown"
    return data_completeness, coverage, warns


def _kr_meta(
    *,
    kind: str,
    data_as_of=None,
    data_period: str | None = None,
    code: str | None = None,
    name: str | None = None,
    data_completeness: str = rmeta.COMPLETE,
    coverage: dict | None = None,
    bar_state: dict | None = None,
    price_adjustment: dict | None = None,
    extra: dict | None = None,
    warnings: list[str] | None = None,
) -> dict:
    """KRX 도구용 메타. 세션 상태에서 data_basis를 자동 판정한다.

    kind:
      "snapshot"    현재가·랭킹·ETF 시세 — 장중이면 realtime, 아니면 last_close
      "bars"        차트·지표·수급 — 장중이면 **in_progress_bar**(마지막 봉 미마감)
      "filing"      재무·컨센서스 — 세션과 무관하게 공시 기반

    data_as_of는 실제 데이터에서 뽑은 날짜를 넘기는 게 원칙이고(차트 마지막 봉,
    공시 발표일), 없을 때만 시장 캘린더의 최근 거래일로 대체한다.
    """
    krx = build_market_clock()["krx"]
    if kind == "filing":
        basis = rmeta.BASIS_FILING
    elif krx.get("is_open"):
        basis = rmeta.BASIS_IN_PROGRESS_BAR if kind == "bars" else rmeta.BASIS_REALTIME
    else:
        basis = rmeta.BASIS_LAST_CLOSE

    warns = list(warnings or [])
    if kind != "filing":
        warns = _kr_market_note(krx) + warns

    data_completeness, coverage, warns = _bar_state_effects(
        bar_state, data_completeness, coverage, warns
    )

    meta = rmeta.build_meta(
        lens="stocklens",
        data_basis=basis,
        data_as_of=data_as_of or (None if kind == "filing" else krx.get("last_trading_day")),
        data_period=data_period,
        market="KR",
        session=krx.get("status"),
        data_completeness=data_completeness,
        coverage=coverage,
        entity_info=rmeta.entity(stock_code=code, name=name),
        warnings=warns,
    )
    if bar_state:
        meta["bar_state"] = bar_state
    if price_adjustment:
        meta["price_adjustment"] = price_adjustment
    # 판정에 영향을 주지 않는 v3 부가 필드(period_coverage 등). 도구가 완전성을
    # 직접 정한 뒤 사실만 실어 보낸다.
    for key, value in (extra or {}).items():
        meta[key] = value
    _emit_coverage_counters("stocklens", meta)
    return meta


@mcp.tool()
@safe_tool
@track_metrics("get_price")
async def get_price(code: str) -> str:
    """현재가 — 종목의 현재 시세 스냅샷 (오늘 하루치 OHLC + 거래량).

    ⚠️ **종목이 2개 이상이면 `get_multi_stocks`를 쓰세요.**
    "삼성전자 지금 얼마", "현재가", "오늘 시세", "주가 알려줘" 같은 질문에 사용합니다.

    ⚠️ **단일 시점 스냅샷**. 과거 시계열 아님. 차트/히스토리 필요 시 get_chart 사용.
    코스피200/코스닥150 등 NXT 대상 종목은 KRX 정규장을 기본값으로 보여주고,
    NXT(대체거래소) 시세는 별도 블록으로 덧붙입니다(두 시장 수치를 섞지 않음).

    Args:
        code: 종목코드 6자리 (예: "005930")
    """
    data = await get_current_price(code)
    # 파서가 실어 보낸 결측 목록과, 실제로 비어 있는 키를 합친다. 파서를 거치지 않고
    # 만들어진 dict(테스트·다른 경로)에서도 '빠진 건 빠졌다'고 말할 수 있어야 한다.
    _required = ("price", "change", "open", "high", "low", "volume")
    missing = sorted(
        set(data.get(PARSE_MISS_KEY) or []) | {k for k in _required if k not in data}
    ) if data else list(_required)
    if not data or "price" not in data:
        # 종목명은 읽혔는데 시세만 비었다면 '데이터 없음'이 아니라 '페이지를 못 읽음'이다.
        # 종목명조차 없으면 애초에 종목 페이지가 아닐 가능성이 높다(잘못된 코드 등).
        if data and data.get("name"):
            text = (
                f"{data['name']}({code})의 시세를 읽지 못했습니다 "
                f"(데이터가 없는 것이 아니라 **파싱 실패**입니다).\n"
                f"못 읽은 항목: {', '.join(missing) or '현재가'}\n"
                f"네이버 증권 페이지 구조가 바뀌었을 수 있습니다. StockLens 업데이트를 확인해 주세요."
            )
            warns = [f"현재가 파싱 실패(구조 변경 의심): {', '.join(missing) or 'price'}"]
        else:
            text = f"종목코드 {code}의 현재가를 가져올 수 없습니다."
            warns = ["현재가 데이터를 가져오지 못했습니다."]
        meta = _kr_meta(
            kind="snapshot", code=code, data_completeness=rmeta.NONE, warnings=warns,
        )
        return _append_result_meta(text, meta)

    has_nxt = "nxt_price" in data
    price_label = "현재가 (KRX 정규장)" if has_nxt else "현재가"

    # 관심종목이면 조회할 때마다 눈에 띄게. 안 보이면 등록해둔 걸 잊는다.
    star = "⭐ " if code in wl.codes() else ""
    lines = [f"{star}종목: {data.get('name', code)} ({code})"]
    # 관리종목·투자경고를 정상 종목과 똑같이 보여주면 시세만 보고 판단하게 된다.
    # 값이 아니라 종목의 성격이 다르므로 가격보다 위에 둔다.
    flags = data.get("status_flags") or []
    if flags:
        lines.append(f"⚠️ **시장경보/지정: {' · '.join(flags)}** — 일반 종목과 매매 조건·위험이 다릅니다")
    lines.append(f"{price_label}: {data['price']:,}원")
    if "change" in data:
        sign = "+" if data["change"] > 0 else ""
        lines.append(f"전일대비: {sign}{data['change']:,}원")
    else:
        lines.append("전일대비: 데이터 없음")
    # 못 읽은 항목은 줄을 빼지 않고 '데이터 없음'으로 남긴다. 조용히 빠지면
    # 사용자는 그 항목이 원래 없는 건지 우리가 못 읽은 건지 알 수 없다.
    for key, label, unit in (
        ("open", "시가", "원"), ("high", "고가", "원"),
        ("low", "저가", "원"), ("volume", "거래량", ""),
    ):
        if key in data:
            lines.append(f"{label}: {data[key]:,}{unit}")
        else:
            lines.append(f"{label}: 데이터 없음")

    if has_nxt:
        lines.append("")
        lines.append("─ NXT(대체거래소) ─")
        lines.append(f"현재가: {data['nxt_price']:,}원")
        if "nxt_change" in data:
            sign = "+" if data["nxt_change"] > 0 else ""
            lines.append(f"전일대비: {sign}{data['nxt_change']:,}원")
        if "nxt_volume" in data:
            lines.append(f"거래량: {data['nxt_volume']:,}")
        lines.append("※ NXT는 KRX와 별도 체결 시장으로, 위 정규장 수치에 포함되지 않습니다")

    completeness = rmeta.COMPLETE if not missing else rmeta.PARTIAL
    warns = [f"시장경보/지정 종목: {' · '.join(flags)}"] if flags else []
    if missing:
        # 무엇이 빠졌는지 이름을 대지 않으면 'partial'은 아무 정보도 주지 못한다.
        warns.append(
            f"읽지 못한 항목: {', '.join(missing)} — 값이 0인 것이 아니라 결측입니다."
        )
    meta = _kr_meta(
        kind="snapshot", code=code, name=data.get("name"),
        data_as_of=data.get("quote_date"), data_completeness=completeness,
        warnings=warns or None,
    )
    return _append_result_meta("\n".join(lines), meta)


@mcp.tool()
@safe_tool
@track_metrics("get_flow")
async def get_flow(code: str, days: int = 20) -> str:
    """투자자수급 — 투자자별 매매동향 (기관/외국인 순매매 주식 수)을 가져옵니다.

    ⚠️ **종목이 2개 이상이면 이 도구를 반복하지 말고 `get_flow_batch`를 쓰세요.**
    네이버 증권 소스 특성상 **개인 순매매는 제공되지 않습니다** (기관·외국인만).
    "외국인 수급", "기관 순매수", "수급 분석", "누가 사고 있어" 같은 질문에 사용합니다.

    Args:
        code: 종목코드 6자리 (예: "005930")
        days: 조회할 일수 (기본 20일)
    """
    days = min(days, 60)
    try:
        data = await get_investor_flow(code, days)
    except NaverParseError as e:
        # 구조 변경은 '데이터 없음'과 다르다 — 없는 척하면 사용자가 조용히 오해한다.
        return _append_result_meta(
            f"수급 표를 읽지 못했습니다 (데이터가 없는 것이 아니라 **파싱 실패**입니다).\n"
            f"원인: {e}\n"
            f"네이버 증권 페이지 구조가 바뀌었을 수 있습니다. StockLens 업데이트를 확인해 주세요.",
            _kr_meta(kind="bars", code=code, data_completeness=rmeta.NONE,
                     warnings=[f"수급 표 파싱 실패(구조 변경 의심): {e}"]),
        )
    if not data:
        return _append_result_meta(
            f"종목코드 {code}의 수급 데이터를 가져올 수 없습니다.",
            _kr_meta(kind="bars", code=code, data_completeness=rmeta.NONE,
                     warnings=["수급 데이터를 가져오지 못했습니다."]),
        )

    lines = [f"종목 {code} 투자자별 매매동향 ({len(data)}일):", ""]
    lines.append("날짜 | [주] 기관 순매매 | [주] 외국인 순매매 | [참고] 종가 | [참고] 거래량")
    lines.append("---|---|---|---|---")
    for row in data:
        lines.append(
            f"{row['date']} | {row['institutional']:,} | {row['foreign']:,} | "
            f"{row['close']:,} | {row['volume']:,}"
        )

    # 합계
    total_inst = sum(r["institutional"] for r in data)
    total_frgn = sum(r["foreign"] for r in data)
    lines.append("")
    lines.append(f"합계 | {total_inst:,} | {total_frgn:,} | - | -")
    lines.append("")
    lines.append(
        "※ [주] 필드는 이 도구의 주 목적 (수급 분석). "
        "[참고] 종가·거래량은 편의 제공이며, **가격 차트·시계열 분석 소스로 사용 금지**. "
        "차트는 get_chart, 현재가는 get_price 사용."
    )

    # 기준일은 표의 최신 행에서 — 캘린더 역산보다 정확하다(거래정지 등).
    return _append_result_meta(
        "\n".join(lines),
        _kr_meta(kind="bars", code=code, data_as_of=data[0].get("date"),
                 data_completeness=rmeta.COMPLETE if len(data) >= min(days, 5) else rmeta.PARTIAL),
    )


_CODE_RE = re.compile(r"^[A-Za-z0-9]{6}$")
_EVENT_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


@mcp.tool()
@safe_tool
@track_metrics("get_event_reaction")
async def get_event_reaction(
    code: str,
    event_date: str,
    before: int = 5,
    after: int = 20,
) -> str:
    """이벤트반응 — 특정 날짜(event_date) 기준 전후 주가·거래량·수급 반응을 정렬합니다.

    DartLens의 scan_earnings_season·list_disclosures에 나온 **공시 접수일**을 event_date로 넘김
    (휴장일이면 다음 거래일 기준). "공시→주가/수급" 또는 "주가/수급 이상→해당 기간 공시 확인"
    시간축 정렬용 — 원인 단정·매수/매도 판단 아님.

    분석 불가한 사건창(보유 데이터 구간 밖·전 구간 거래정지·수급 결측)은 숫자를 만들지 않고
    `validation` 상태와 코드로 되돌린다. 데이터 없음은 0%·순매매 0으로 표기되지 않는다.

    Args:
        code: 종목코드 6자리 (예: "005930")
        event_date: 기준 날짜 YYYY-MM-DD 또는 YYYYMMDD. DartLens 공시 접수일 권장
        before: event_date 전 비교 거래일 수 (기본 5, 최대 60)
        after: event_date 후 비교 거래일 수 (기본 20, 최대 60)
    """
    if not _CODE_RE.match(code or ""):
        return "종목코드는 6자리 영숫자여야 합니다. 종목명이라면 먼저 search 도구로 코드를 확정하세요."

    if not _EVENT_DATE_RE.match(normalize_date(event_date)):
        return "event_date는 YYYY-MM-DD 또는 YYYYMMDD 형식이어야 합니다."

    before = max(0, min(int(before), 60))
    after = max(0, min(int(after), 60))

    # 공급자 실패는 '사건 무반응'이 아니라 조회 실패로 구분해 되돌린다.
    try:
        ohlcv = await get_ohlcv(code, "day", 500)
    except Exception as e:
        return (
            f"⚠️ `{PROVIDER_ERROR}` — 종목코드 {code}의 일봉 조회에 실패했습니다 "
            f"({type(e).__name__}). 데이터 없음이 아니라 조회 실패이므로 재시도하세요."
        )

    # 수급은 최근 구간만 제공된다. 실패해도 주가 반응은 계산하되 결측으로 표시한다.
    flow_days = max(20, min(before + after + 10, 60))
    flow_error = None
    try:
        flows = await get_investor_flow(code, flow_days)
    except Exception as e:
        flows, flow_error = [], type(e).__name__

    reaction = build_event_reaction(
        code=code,
        event_date=event_date,
        ohlcv=ohlcv,
        flows=flows,
        before=before,
        after=after,
        flow_error=flow_error,
    )
    # 본문의 검증 상태를 같은 근거로 meta v3 봉투에 싣는다(SL-08).
    fields = reaction_meta_fields(reaction)
    return _append_result_meta(
        format_event_reaction(reaction),
        _kr_meta(
            kind="bars",
            code=code,
            data_as_of=fields["event_window"].get("last_usable_trading_date"),
            data_completeness=fields["data_completeness"],
            coverage=fields["coverage"],
            price_adjustment=reaction.get("price_adjustment"),
            extra={"event_window": fields["event_window"],
                   "flow_window": fields["flow_window"]},
            warnings=fields["warnings"] or None,
        ),
    )

# ETF/ETN 식별용 운용사 브랜드 — 거래량 랭킹 결과에서 종목명으로 빠르게 거름.
_ETF_NAME_PREFIXES = (
    "KODEX", "TIGER", "KBSTAR", "ARIRANG", "HANARO", "ACE", "RISE", "PLUS",
    "SOL", "KOSEF", "TIMEFOLIO", "TRUE", "WOORI", "1Q", "HK ", "마이다스",
)


def _looks_like_etf(name: str) -> bool:
    if not name:
        return False
    upper = name.upper()
    if "ETF" in upper or "ETN" in upper:
        return True
    return any(upper.startswith(p) for p in _ETF_NAME_PREFIXES)


@mcp.tool()
@safe_tool
@track_metrics("get_event_reactions")
async def get_event_reactions(
    code: str,
    max_events: int = 8,
    before: int = 5,
    after: int = 10,
) -> str:
    """공시반응이력 — 최근 공시들에 이 종목이 **어떻게 반응해 왔는지** 한 번에.

    "이 종목은 호재에 잘 오르나", "공시 나오면 어떻게 움직였나", "최근 반응이
    예전보다 약해졌나" 같은 질문에 사용합니다.

    `get_event_reaction`은 날짜 **하나**만 봅니다. 공시 8건을 보려면 8번 불러야 하고
    응답도 8배가 됩니다. 이 도구는 일봉·수급을 **한 번만** 받아 여러 공시일을
    한꺼번에 계산합니다.

    ⚠️ 반응이 좋았다고 앞으로도 그러리라는 뜻이 아닙니다. 과거 기록일 뿐입니다.
    ⚠️ 네이버 공시 목록에는 '가격제한폭 확대요건 도달' 같은 시장 안내도 섞입니다.
    제목을 보고 실제 기업 공시인지 구분하세요.

    Args:
        code: 종목코드 6자리
        max_events: 볼 공시 개수 (기본 8, 최대 15). 같은 날 여러 건이면 하나로 묶습니다.
        before: 기준일 전 비교 거래일 수 (기본 5)
        after: 기준일 후 비교 거래일 수 (기본 10, 최대 60)
    """
    import re as _re
    import statistics as _st

    if not _re.match(r"^[A-Za-z0-9]{6}$", code):
        return f"⚠️ 종목코드 형식이 올바르지 않습니다: {code}"
    max_events = max(1, min(int(max_events), 15))
    before = max(0, min(int(before), 60))
    after = max(1, min(int(after), 60))

    items = await naver_get_disclosure_list(code)
    if not items:
        return f"종목 {code}의 공시 목록을 찾지 못했습니다."

    # 같은 날 여러 건은 첫 건으로 대표시킨다 — 같은 날짜의 반응은 어차피 같다.
    seen: dict[str, str] = {}
    for it in items:
        day = str(it.get("date", "")).replace(".", "-").strip()
        if len(day) == 10 and day not in seen:
            seen[day] = str(it.get("title", ""))
        if len(seen) >= max_events:
            break
    if not seen:
        return f"종목 {code}의 공시에서 날짜를 읽지 못했습니다."

    try:
        ohlcv = await get_ohlcv(code, "day", 500)
    except Exception as e:
        return (f"⚠️ `{PROVIDER_ERROR}` — 일봉 조회 실패 ({type(e).__name__}). "
                f"데이터 없음이 아니라 조회 실패이므로 재시도하세요.")
    flow_error = None
    try:
        flows = await get_investor_flow(code, 60)
    except Exception as e:
        flows, flow_error = [], type(e).__name__

    rows = []
    for day, title in sorted(seen.items(), reverse=True):
        try:
            r = build_event_reaction(code=code, event_date=day, ohlcv=ohlcv, flows=flows,
                                     before=before, after=after, flow_error=flow_error)
        except ValueError:
            continue
        rows.append((day, title, r))
    if not rows:
        return f"종목 {code}의 공시일을 주가 데이터와 맞추지 못했습니다."

    def pt(r, key):
        p = (r.get("points") or {}).get(key) or {}
        return p.get("return_pct") if p.get("status") == "available" else None

    lines = [f"# 공시 반응 이력 — {code} (최근 {len(rows)}건)", ""]
    lines.append(f"기준일 | 공시 | D+1 | D+{after//2 if after>=4 else 1} | D+{after} | 당일거래량/직전평균")
    lines.append("---|---|---:|---:|---:|---:")
    mid_key = f"D+{after//2}" if after >= 4 else "D+1"
    d1s = []
    for day, title, r in rows:
        v = r.get("volume") or {}
        pre, basis = v.get("pre_avg"), v.get("basis")
        ratio = f"{basis/pre:.1f}배" if isinstance(pre, (int, float)) and pre and isinstance(basis, (int, float)) else "-"
        d1 = pt(r, "D+1"); dm = pt(r, mid_key); dl = pt(r, f"D+{after}")
        if d1 is not None:
            d1s.append(d1)
        f = lambda x: f"{x:+.2f}%" if isinstance(x, (int, float)) else "-"
        t = (title or "").replace("|", "/")
        t = (t[:34] + "…") if len(t) > 34 else t
        lines.append(f"{day} | {t} | {f(d1)} | {f(dm)} | {f(dl)} | {ratio}")

    if d1s:
        up = sum(1 for x in d1s if x > 0)
        lines.append("")
        lines.append("## 이 종목의 반응 습관")
        lines.append(f"- 계산된 공시 {len(d1s)}건 중 **다음 날 오른 경우 {up}건** ({up/len(d1s)*100:.0f}%)")
        lines.append(f"- D+1 중앙값 **{_st.median(d1s):+.2f}%** · 평균 {sum(d1s)/len(d1s):+.2f}%")
        lines.append(f"- 최대 {max(d1s):+.2f}% / 최소 {min(d1s):+.2f}%")
        if len(d1s) >= 4:
            half = len(d1s) // 2
            recent, past = d1s[:half], d1s[half:]   # rows 가 최신순이라 앞쪽이 최근
            rm, pm = _st.median(recent), _st.median(past)
            trend = "약해지는 중" if rm < pm - 0.5 else ("강해지는 중" if rm > pm + 0.5 else "비슷")
            lines.append(f"- 최근 {len(recent)}건 중앙값 {rm:+.2f}% vs 이전 {len(past)}건 {pm:+.2f}% → **{trend}**")
            lines.append("  · 좋은 소식이 계속 나오는데 반응이 약해진다면 기대가 이미 반영됐을 수 있습니다.")

    lines.append("")
    lines.append("※ 과거 반응이지 예측이 아닙니다. 같은 공시라도 시장 상황에 따라 다르게 움직입니다.")
    lines.append(f"※ {CORPORATE_ACTION_NOTE}")
    lines.append("※ 시장 전체가 크게 움직인 날은 개별 반응과 섞입니다 — get_index 로 그날 지수를 함께 보세요.")
    if flow_error:
        lines.append(f"※ 수급 조회 실패({flow_error}) — 주가 반응만 계산했습니다.")

    # 사건별 검증 상태를 집계한다. 하나라도 창이 덜 찼으면 이 표의 해당 칸은
    # 값이 아니라 결측이고, 그 사실이 메타에 있어야 프로그램이 읽는다(SL-08).
    statuses = [(r.get("validation") or {}).get("status") for _, _, r in rows]
    all_codes: set[str] = set()
    for _, _, r in rows:
        all_codes.update((r.get("validation") or {}).get("codes") or [])
    event_coverage = {
        "events": len(rows),
        "usable": sum(1 for s in statuses if s == "usable"),
        "partial": sum(1 for s in statuses if s == "partial"),
        "unavailable": sum(1 for s in statuses if s == "unavailable"),
    }
    incomplete = event_coverage["partial"] + event_coverage["unavailable"]
    if incomplete:
        ev_reason = ("insufficient_post_event_history"
                     if INSUFFICIENT_POST_EVENT_HISTORY in all_codes
                     else "source_limit")
        ev_cov = {"truncated": False, "coverage_complete": False, "reason": ev_reason}
    else:
        ev_cov = None
    return _append_result_meta(
        "\n".join(lines),
        _kr_meta(
            kind="bars", code=code,
            data_completeness=rmeta.PARTIAL if incomplete else rmeta.COMPLETE,
            coverage=ev_cov,
            price_adjustment=price_adjustment_meta(),
            extra={"event_coverage": event_coverage},
            warnings=([f"{incomplete}건 사건은 창이 덜 찼거나 분석 불가입니다"
                       " - 표의 해당 칸은 값이 아니라 결측입니다."]
                      if incomplete else None),
        ),
    )



@mcp.tool()
@safe_tool
@track_metrics("get_flow_batch")
async def get_flow_batch(codes: list[str], days: int = 5) -> str:
    """수급벌크 — 여러 종목의 투자자 수급(기관·외국인 순매매)을 한 번에 병렬 조회.

    개별 get_flow를 N번 부르는 것보다 훨씬 빠릅니다 (서버 내부 asyncio.gather).
    "이 종목들 외국인 매수 같이 들어왔나", "관심종목 수급 비교" 같은 질문에 사용.

    Args:
        codes: 종목코드 리스트 (최대 30개)
        days: 종목당 조회 일수 (기본 5, 최대 20). 토큰 절감 위해 단일 도구보다 짧음
    """
    if not codes:
        return "종목코드 리스트가 비어 있습니다."

    valid: list[str] = []
    invalid: list[str] = []
    for c in codes:
        if _CODE_RE.match(c or ""):
            valid.append(c)
        else:
            invalid.append(c)

    if not valid:
        return "유효한 종목코드가 없습니다 (6자리 영숫자)."

    # 상한에 걸려 잘린 사실은 보정 전 값을 들고 있어야만 말할 수 있다.
    # 지금까지 본문에는 "최근 20일"만 적혀서, 60일을 물어본 사람은 그 20일치를
    # 자기가 요청한 범위로 읽었다.
    requested_codes = len(codes)
    valid = valid[:30]
    requested_days = days
    effective_days = max(1, min(days, 20))
    days = effective_days
    days_truncated = requested_days != effective_days
    codes_truncated = requested_codes > len(valid) + len(invalid)

    async def fetch_one(code: str) -> tuple[str, list[dict] | None]:
        try:
            data = await get_investor_flow(code, days)
            return code, data
        except Exception:
            return code, None

    results = await asyncio.gather(*[fetch_one(c) for c in valid])

    # 종목명은 한 번에 모아서 표시 (토큰 절감, 병렬 호출)
    name_map: dict[str, str] = {}
    try:
        basic = await naver_get_multi_stocks(valid)
        name_map = {b["code"]: b.get("name", "") for b in basic}
    except Exception:
        pass

    lines: list[str] = []
    if days_truncated:
        lines.append(
            f"⚠️ {requested_days}일을 요청했지만 이 도구의 상한은 {effective_days}일입니다. "
            f"아래는 최근 {effective_days}일치입니다 (더 긴 구간은 get_flow 로 종목별 조회)."
        )
    if codes_truncated:
        lines.append(
            f"⚠️ {requested_codes}종목을 요청했지만 한 번에 30종목까지만 조회합니다. "
            f"앞 {len(valid)}종목만 아래에 있습니다."
        )
    if lines:
        lines.append("")
    lines += [
        f"수급 배치 ({len(valid)}종목, 최근 {days}일):",
        "각 행: 날짜 | 기관 순매매(주) | 외국인 순매매(주)  ※ 양수=순매수",
        "",
    ]

    failed: list[str] = list(invalid)
    matched_count = 0
    # 상한 안이어도 종목마다 실제 행 수는 다르다(신규 상장·거래정지). 합계만
    # 보여주면 3일치와 20일치가 같은 무게로 읽힌다.
    per_entity: dict[str, int] = {}
    # 배치의 기준일은 시장 달력이 아니라 실제 반환 데이터에서 나와야 한다.
    # 거래정지·수집 지연 종목의 옛 데이터가 "오늘 자료"로 읽히면 안 된다.
    per_entity_as_of: dict[str, str] = {}
    for code, data in results:
        name = name_map.get(code, "")
        header = f"[{code}] {name}".rstrip()
        per_entity[code] = len(data or [])
        row_days = [rmeta.normalize_day(r.get("date")) for r in (data or [])]
        row_days = [d for d in row_days if d]
        if row_days:
            per_entity_as_of[code] = max(row_days)
        if not data:
            failed.append(code)
            continue

        matched_count += 1
        lines.append(header)
        for row in data:
            lines.append(
                f"  {row['date']}  {row['institutional']:+,}  {row['foreign']:+,}"
            )
        total_inst = sum(r["institutional"] for r in data)
        total_frgn = sum(r["foreign"] for r in data)
        lines.append(f"  합계({len(data)}일)  {total_inst:+,}  {total_frgn:+,}")
        lines.append("")

    # 20일을 요청해 3일만 온 종목이 섞였으면 그 배치는 요청 범위를 못 채운 것이다.
    # 상한과 조회 실패만 보면 이 경우가 통째로 새어 나간다.
    short_entities = sorted(
        c for c, n in per_entity.items() if 0 < n < effective_days
    )

    if matched_count == 0:
        msg = "조회 가능한 종목이 없습니다."
        if failed:
            msg += f" (실패: {', '.join(failed)})"
        return msg

    if failed:
        lines.append(f"※ 조회 실패: {', '.join(failed)}")
    if short_entities:
        lines.append(
            f"※ 요청한 {effective_days}일보다 짧게 온 종목: "
            + ", ".join(f"{c}({per_entity[c]}일)" for c in short_entities)
            + ". 상장한 지 얼마 안 됐거나 거래정지 구간이 있을 수 있습니다."
        )

    truncated = days_truncated or codes_truncated
    has_gap = truncated or bool(failed) or bool(short_entities)
    coverage = {
        "requested": {"unit": "day", "value": requested_days},
        "effective": {"unit": "day", "value": effective_days},
        "requested_entities": requested_codes,
        "returned_entities": matched_count,
        "per_entity_returned_count": per_entity,
        "returned_count": sum(per_entity.values()),
        "total_count": None,
        "truncated": truncated,
        "short_entities": short_entities,
        "coverage_complete": not has_gap,
        "reason": (
            "server_cap" if truncated
            else "source_limit" if short_entities
            else "unknown" if failed
            else None
        ),
    }

    return _append_result_meta(
        "\n".join(lines),
        _kr_meta(
            kind="bars",
            data_as_of=max(per_entity_as_of.values()) if per_entity_as_of else None,
            data_completeness=rmeta.PARTIAL if has_gap else rmeta.COMPLETE,
            coverage=coverage,
            extra={"per_entity_data_as_of": per_entity_as_of},
        ),
    )


@mcp.tool()
@safe_tool
@track_metrics("screen_by_flow")
async def screen_by_flow(
    top_n: int = 100,
    market: str = "ALL",
    foreign_days: int = 2,
    inst_days: int = 2,
    exclude_etf: bool = True,
    sort_by: str = "trade_value",
) -> str:
    """수급스크리닝 — 거래대금/거래량 상위 N개 중 외국인·기관이 최근 며칠 연속 순매수한 종목만 추립니다.

    "거래대금 상위 중 외국인·기관 동반 매수", "이틀 연속 수급 들어온 종목" 같은 스크리닝 전용.
    랭킹+수급을 서버에서 join해 매치 종목만 반환 → 토큰·시간 절감.

    Args:
        top_n: 상위 후보 수 (기본 100, 최대 500). 클수록 정확하나 느림(500≈20~50초)
        market: "KOSPI"/"KOSDAQ"/"ALL" (기본 ALL)
        foreign_days: 최근 N일 모두 외국인 순매수여야 매치 (0=미적용)
        inst_days: 최근 N일 모두 기관 순매수여야 매치 (0=미적용)
        exclude_etf: ETF/ETN 제외 (기본 True)
        sort_by: "trade_value"(거래대금, 기본)/"volume"(거래량)
    """
    top_n = max(1, min(top_n, 500))
    foreign_days = max(0, min(foreign_days, 10))
    inst_days = max(0, min(inst_days, 10))
    needed_days = max(foreign_days, inst_days, 1)

    ranks = await naver_get_volume_ranking(market=market, count=top_n, sort_by=sort_by)
    if not ranks:
        meta = _kr_meta(kind="bars", data_completeness=rmeta.NONE, warnings=[f"{market} 거래량 순위를 가져오지 못했습니다."])
        return _append_result_meta(f"{market} 거래량 순위를 가져올 수 없습니다.", meta)

    # 코드 단위 dedup — 네이버 페이지 경계나 시장 합산에서 같은 종목이
    # 두 번 등장하는 케이스 방지 (먼저 등장한 항목 우선, rank 보존)
    seen_codes: set[str] = set()
    candidates = []
    for r in ranks:
        code = r.get("code")
        if code and code not in seen_codes:
            seen_codes.add(code)
            candidates.append(r)

    excluded_etf: list[str] = []
    if exclude_etf:
        kept = []
        for r in candidates:
            if _looks_like_etf(r.get("name", "")):
                excluded_etf.append(f"{r['code']} {r['name']}")
            else:
                kept.append(r)
        candidates = kept

    async def fetch_flow(item: dict) -> tuple[dict, list[dict] | None]:
        try:
            data = await get_investor_flow(item["code"], needed_days)
            return item, data
        except Exception:
            return item, None

    flow_results = await asyncio.gather(*[fetch_flow(c) for c in candidates])
    failed_fetch = sum(1 for _, data in flow_results if data is None)

    matched: list[tuple[dict, list[dict]]] = []
    for item, data in flow_results:
        if not data or len(data) < needed_days:
            continue
        head = data[:needed_days]  # 최신 N일 (네이버는 최신순)
        ok_frgn = foreign_days == 0 or all(
            r["foreign"] > 0 for r in head[:foreign_days]
        )
        ok_inst = inst_days == 0 or all(
            r["institutional"] > 0 for r in head[:inst_days]
        )
        if ok_frgn and ok_inst:
            matched.append((item, data))

    sort_label = "거래대금" if sort_by == "trade_value" else "거래량"
    cond_parts = []
    if foreign_days > 0:
        cond_parts.append(f"외국인 {foreign_days}일 연속 순매수")
    if inst_days > 0:
        cond_parts.append(f"기관 {inst_days}일 연속 순매수")
    cond_label = " + ".join(cond_parts) if cond_parts else "조건 없음"

    lines = [
        f"수급 스크리닝 — {sort_label} 상위 {top_n} ({market}) 중 {cond_label}",
        f"후보 {len(candidates)}개 → 매치 {len(matched)}개"
        + (f" (ETF/ETN 제외: {len(excluded_etf)})" if exclude_etf else ""),
        "",
    ]

    warnings = []
    if failed_fetch:
        warnings.append(f"수급 조회 실패 종목 {failed_fetch}개 — 스크리닝 후보에서 제외됨.")
    completeness = rmeta.PARTIAL if failed_fetch else rmeta.COMPLETE

    if not matched:
        lines.append("⚠️ 조건을 만족하는 종목이 없습니다.")
        lines.append("")
        lines.append("💡 조건 완화 예: foreign_days=1 또는 inst_days=0 (외국인만 검사)")
        if exclude_etf and excluded_etf:
            lines.append(f"제외된 ETF/ETN ({len(excluded_etf)}개): {', '.join(excluded_etf[:5])}{' …' if len(excluded_etf) > 5 else ''}")
        meta = _kr_meta(kind="bars", data_completeness=completeness, warnings=warnings)
        return _append_result_meta("\n".join(lines), meta)

    lines.append("매치 종목 (각 행: 날짜 | 기관 순매매 | 외국인 순매매, 양수=순매수):")
    lines.append("")
    for item, data in matched:
        head = data[:needed_days]
        lines.append(f"[{item['code']}] {item['name']}  현재가 {item['price']:,} ({item.get('change_rate', '')})")
        for row in head:
            lines.append(
                f"  {row['date']}  기관 {row['institutional']:+,}  외국인 {row['foreign']:+,}"
            )
        lines.append("")

    meta = _kr_meta(kind="bars", data_completeness=completeness, warnings=warnings)
    return _append_result_meta("\n".join(lines), meta)


@mcp.tool()
@safe_tool
@track_metrics("get_financial")
async def get_financial(code: str) -> str:
    """재무지표 — 한 종목의 **전체** 재무지표(연간·분기 추이 19개 항목).

    ⚠️ **여러 종목을 비교하려면 `get_financial_batch`를 쓰세요.** 이 도구는 종목당
    3,500자를 뱉어서, 5종목만 비교해도 재무가 전체 토큰의 78%를 먹습니다.
    "PER", "PBR", "재무제표", "시가총액", "저평가" 같은 질문에 사용합니다.

    Args:
        code: 종목코드 6자리 (예: "005930")
    """
    data = await get_financials(code)
    if not data:
        return f"종목코드 {code}의 재무지표를 가져올 수 없습니다."
    if PARSE_MISS_KEY in data:
        return (
            f"{data.get('name', code)}({code})의 재무 표를 읽지 못했습니다 "
            f"(재무 자료가 없는 것이 아니라 **파싱 실패**입니다).\n"
            f"네이버 증권 페이지 구조가 바뀌었을 수 있습니다. "
            f"StockLens 업데이트를 확인해 주세요."
        )

    periods = data.get("_periods") or {}
    annual = periods.get("annual", [])
    quarterly = periods.get("quarterly", [])

    def _fmt(p: list[str], v: list[str]) -> str:
        parts = []
        for pp, vv in zip(p, v):
            if not pp and not vv:
                continue
            if not vv:
                vv = "추정치 없음" if "(E)" in pp else "데이터 없음"
            parts.append(f"{pp or '?'}={vv}")
        return " | ".join(parts)

    lines = [f"종목: {data.get('name', code)} ({code})", ""]

    # 단위·재무기준을 안 적어 매출액 1,558이 억원인지 백만원인지 알 수 없었다.
    # 행 이름에 단위가 붙은 항목(EPS(원)·PER(배) 등)은 그대로 두고, 단위가 없는
    # 금액·비율 행이 무엇인지만 여기서 한 번 선언한다.
    lines.append(
        "※ 단위 — 매출액·영업이익·당기순이익: **억원** / "
        "영업이익률·순이익률·ROE·부채비율·당좌비율·유보율: **%** "
        "(네이버 금융 기업실적분석 표 기준). 그 외 행은 이름에 단위 표기."
    )
    basis = data.get("_fs_basis") or []
    if basis:
        uniq = sorted(set(b for b in basis if b))
        if len(uniq) == 1:
            lines.append(f"※ 재무기준: {uniq[0]} (전 기간 동일)")
        elif uniq:
            all_periods = list(annual) + list(quarterly)
            pairs = [f"{p}={b}" for p, b in zip(all_periods, basis) if b]
            lines.append(
                "※ ⚠️ 재무기준이 기간마다 다릅니다 — " + ", ".join(pairs)
                + ". 연결과 별도는 비교 범위가 달라 나란히 읽으면 안 됩니다."
            )
    lines.append("※ (E)는 확정 실적이 아니라 증권사 컨센서스 추정치입니다.")
    # PER·PBR 은 '가격 ÷ 실적'이라 어느 시점 주가로 나눴는지가 값을 좌우한다.
    # 이 표는 각 기간말 주가 기준이고, get_sector_valuation 은 현재가÷TTM 이라
    # 같은 종목인데 두 숫자가 다르게 보인다(한화오션 14.82 vs 12.79).
    if any(k.startswith(("PER", "PBR")) for k in data):
        lines.append(
            "※ 표의 **PER·PBR 은 각 기간말 주가** 기준입니다. "
            "**현재가** 기준이 필요하면 `get_sector_valuation`(현재가÷최근 4분기 EPS) "
            "또는 네이버 종목 화면을 보세요 — 같은 종목이라도 값이 다릅니다."
        )
    lines.append("")

    for key, value in data.items():
        if key in ("code", "name", "_periods", "_fs_basis"):
            continue
        if isinstance(value, list):
            if annual or quarterly:
                a_vals = value[: len(annual)]
                q_vals = value[len(annual) : len(annual) + len(quarterly)]
                lines.append(f"{key}:")
                if annual:
                    lines.append(f"  [연간] {_fmt(annual, a_vals)}")
                if quarterly:
                    lines.append(f"  [분기] {_fmt(quarterly, q_vals)}")
            else:
                lines.append(f"{key}: {' | '.join(value)}")
        else:
            lines.append(f"{key}: {value}")

    # 재무는 '날짜'가 아니라 '기간'이 기준이다. (E)가 안 붙은 마지막 기간이
    # 확정 실적의 끝 — 날짜를 지어내는 대신 그 라벨을 그대로 싣는다.
    confirmed = [p for p in (list(annual) + list(quarterly)) if p and "(E)" not in p]
    meta = _kr_meta(
        kind="filing", code=code, name=data.get("name"),
        data_period=f"{confirmed[-1]} 확정" if confirmed else None,
    )
    return _append_result_meta("\n".join(lines), meta)


# 비교표에 쓰는 핵심 지표만. get_financial은 19행을 다 뱉어 종목당 3,559자인데,
# 여러 종목을 나란히 볼 때 필요한 건 이 정도다(5종목 비교 시 재무가 토큰의 78%였다).
_FIN_COMPARE_ROWS = [
    ("PER(배)", "PER"),
    ("PBR(배)", "PBR"),
    ("ROE(지배주주)", "ROE%"),
    ("영업이익률", "영업이익률%"),
    ("부채비율", "부채비율%"),
    ("시가배당률(%)", "배당률%"),
]


def _latest_confirmed(periods: dict, values: list[str]) -> tuple[str, str]:
    """(E)가 안 붙은 마지막 기간의 값. 추정치를 확정처럼 쓰지 않는다."""
    labels = list(periods.get("annual") or []) + list(periods.get("quarterly") or [])
    pairs = [(p, v) for p, v in zip(labels, values) if p and "(E)" not in p and v]
    return pairs[-1] if pairs else ("", "")


@mcp.tool()
@safe_tool
@track_metrics("get_financial_batch")
async def get_financial_batch(codes: list[str]) -> str:
    """재무벌크 — 여러 종목의 핵심 재무지표(PER/PBR/ROE/영업이익률/부채비율/배당률)를 한 표로.

    ⭐ **종목 비교의 기본 도구.** "A와 B를 PER·PBR로 비교", "이 5종목 중 저평가",
    "업종 내 비교" 같은 요청에 get_financial을 종목마다 부르지 말고 이걸 쓰세요.
    (실측: 5종목 비교 시 get_financial 5회가 전체 토큰의 78%를 차지했습니다.)

    한 종목의 **전체** 지표(연간·분기 시계열, EPS/BPS/유보율 등)가 필요하면
    그때만 get_financial을 쓰세요.

    Args:
        codes: 종목코드 6자리 리스트 (최대 30개)
    """
    if not codes:
        return "종목코드 리스트가 비어 있습니다."
    codes = [c.strip() for c in codes if c and c.strip()][:30]

    async def one(code: str):
        try:
            return code, await get_financials(code)
        except Exception as e:
            return code, {"_error": f"{type(e).__name__}"}

    results = await asyncio.gather(*(one(c) for c in codes))

    header = "코드 | 종목명 | " + " | ".join(lbl for _, lbl in _FIN_COMPARE_ROWS) + " | 기준"
    lines = [f"재무 비교 ({len(results)}종목)", "", header,
             "---|---|" + "|".join(["---:"] * len(_FIN_COMPARE_ROWS)) + "|---"]
    failed = 0
    # 종목마다 기준 분기가 다르면 이 표는 비교표가 아니다. A는 2026.06 확정인데
    # B는 아직 2026.03 이면 PER 을 나란히 놓는 순간 서로 다른 시점을 비교하게 된다.
    period_by_code: dict[str, str] = {}
    for code, data in results:
        if not data or data.get("_error"):
            failed += 1
            lines.append(f"{code} | (조회 실패) | " + " | ".join(["-"] * len(_FIN_COMPARE_ROWS)) + " | -")
            continue
        periods = data.get("_periods") or {}
        cells, basis = [], ""
        for key, _ in _FIN_COMPARE_ROWS:
            val = data.get(key)
            if isinstance(val, list):
                p, v = _latest_confirmed(periods, val)
                cells.append(v or "-")
                basis = basis or p
            else:
                cells.append(str(val) if val else "-")
        if basis:
            period_by_code[code] = basis
        star = "⭐ " if code in wl.codes() else ""
        lines.append(f"{code} | {star}{data.get('name', code)} | " + " | ".join(cells) + f" | {basis or '-'}")

    lines.append("")
    lines.append("※ 단위 — PER·PBR: 배 / 나머지: % · 값은 **(E)가 붙지 않은 최신 확정 기간** 기준")
    lines.append("※ 전체 지표(연간·분기 추이, EPS/BPS 등)가 필요하면 get_financial을 종목별로 쓰세요.")
    if failed:
        lines.append(f"※ {failed}종목 조회 실패 — 상장폐지·거래정지 등일 수 있습니다.")

    period_coverage = _period_coverage_of(period_by_code)
    consistency = period_coverage["consistency"]
    if consistency == "mixed":
        uniq = sorted(set(period_by_code.values()))
        lines.append(
            "※ 종목마다 기준 기간이 다릅니다(" + " / ".join(uniq) + "). "
            "서로 다른 시점의 값이라 PER·ROE 를 그대로 나란히 놓고 비교하면 안 됩니다."
        )
    # 기준을 하나도 못 읽었으면 "전부 같다"가 아니라 "모른다"다. 판별 실패는
    # complete 가 될 수 없다(설계 5.2).
    incomplete = bool(failed) or consistency in ("mixed", "unknown")
    fin_coverage = None
    if consistency in ("mixed", "unknown"):
        fin_coverage = {
            "returned_count": len(period_by_code),
            "total_count": len(results),
            "truncated": False,
            "coverage_complete": False,
            "reason": "mixed_periods" if consistency == "mixed" else "unknown",
        }

    return _append_result_meta(
        "\n".join(lines),
        _kr_meta(kind="filing",
                 data_completeness=rmeta.PARTIAL if incomplete else rmeta.COMPLETE,
                 coverage=fin_coverage,
                 extra={"period_coverage": period_coverage}),
    )


@mcp.tool()
@safe_tool
@track_metrics("get_index")
async def get_index() -> str:
    """시장지수 — KOSPI, KOSDAQ 지수 현재값을 가져옵니다.
    "코스피", "코스닥", "시장 지수", "오늘 시장 어때" 같은 질문에 사용합니다.
    """
    data = await get_market_index()
    if not data:
        return "시장 지수를 가져올 수 없습니다."

    lines = ["시장 지수:"]
    warns: list[str] = []
    for item in data:
        miss = item.get(PARSE_MISS_KEY) or []
        value = item.get("value")
        # 못 읽은 지수를 '-'로만 두면 휴장인지 파싱 실패인지 구분이 안 된다.
        value_str = f"{value:,.2f}" if isinstance(value, float) else "데이터 없음"
        # change_raw(네이버 원문)를 그대로 쓰면 방향 라벨이 값과 어긋난다 —
        # 하락한 날에도 원문 끝에 "상승"이 남아 있고, 퍼센트도 두 번 찍힌다.
        # 부호를 이미 적용해 둔 change_value 를 쓴다.
        rate = item.get("change_rate")
        rate_str = f" ({rate:+.2f}%)" if rate is not None else ""
        point = item.get("change_value")
        point_str = f" {point:+,.2f}p" if isinstance(point, (int, float)) else ""
        lines.append(f"  {item['index']}: {value_str}{rate_str}{point_str}".rstrip())
        if miss:
            warns.append(f"{item['index']} 읽지 못한 항목: {', '.join(miss)}")

    completeness = rmeta.PARTIAL if warns else rmeta.COMPLETE
    if warns and all(item.get(PARSE_MISS_KEY) for item in data):
        completeness = rmeta.NONE
        warns.append("모든 지수를 읽지 못했습니다 — 네이버 페이지 구조 변경 가능성.")
    return _append_result_meta(
        "\n".join(lines),
        _kr_meta(kind="snapshot", data_completeness=completeness, warnings=warns or None),
    )


@mcp.tool()
@safe_tool
@track_metrics("list_themes")
async def list_themes(page: int = 1) -> str:
    """테마목록 — 네이버 증권의 테마 목록을 가져옵니다.
    "어떤 테마가 있어?", "테마 리스트", "오늘 강세 테마" 같은 질문에 사용합니다.
    총 7페이지가 있으며 한 페이지당 40개 테마가 있습니다. 전일대비 등락률 순으로 정렬되어 있어요.

    Args:
        page: 페이지 번호 (1~7, 기본 1)
    """
    page = max(1, min(page, 7))
    themes = await naver_list_themes(page=page)
    if not themes:
        return f"페이지 {page}의 테마 목록을 가져올 수 없습니다."

    lines = [f"테마 목록 (page {page}, {len(themes)}개):", ""]
    lines.append("테마명 | 전일대비 | 최근3일 | 상승/보합/하락 | 주도주")
    lines.append("---|---|---|---|---")
    for t in themes:
        leaders = ", ".join(
            f"{ld['name']}({ld['code']})" if ld.get("code") else ld["name"]
            for ld in t["leaders"]
        )
        counts = f"{t['up_count']}/{t['flat_count']}/{t['down_count']}"
        lines.append(
            f"{t['name']} | {t['change_rate']} | {t['recent_3d_rate']} | {counts} | {leaders}"
        )
    return "\n".join(lines)


@mcp.tool()
@safe_tool
@track_metrics("get_theme_stocks")
async def get_theme_stocks(
    theme_name: str,
    count: int = 30,
    include_reason: bool = True,
) -> str:
    """테마종목 — 특정 테마에 속한 종목 리스트를 가져옵니다.
    "반도체 테마 종목", "2차전지 관련주", "AI 테마주" 같은 질문에 사용합니다.
    테마명 부분 매칭을 지원합니다.

    ⚠️ **등락률 내림차순**으로 반환합니다. count로 자르면 그날 많이 오른 종목만 남고
    **소외된(많이 내린) 종목은 목록에서 빠집니다.** 저평가·소외 종목을 찾는 용도라면
    count를 테마 전체 종목 수 이상으로 주거나, 시가총액 기준
    `get_market_cap_ranking`을 모집단으로 쓰세요.

    Args:
        theme_name: 테마명 (예: "2차전지", "AI", "반도체")
        count: 반환할 최대 종목 수 (기본 30)
        include_reason: 편입사유 포함 여부. False로 하면 토큰 대폭 절감.
                        "왜 이 테마에 들어갔는지" 필요 없으면 False 권장.
    """
    count = min(count, 50)
    result = await naver_get_theme_stocks(
        theme_name,
        count=count,
        include_reason=include_reason,
    )
    if not result.get("theme_id"):
        return (
            f"'{theme_name}' 테마를 찾을 수 없습니다. "
            f"list_themes로 전체 테마 목록을 먼저 확인해보세요."
        )

    stocks = result["stocks"]
    lines = [f"테마: {result['theme_name']} ({len(stocks)}개 종목)", ""]

    if include_reason:
        lines.append("코드 | 종목명 | 현재가 | 등락률 | 거래량 | 편입사유")
        lines.append("---|---|---|---|---|---")
        for s in stocks:
            reason = s.get("reason", "")
            lines.append(
                f"{s['code']} | {s['name']} | {s['price']:,} | {s['change_rate']} | "
                f"{s['volume']:,} | {reason}"
            )
    else:
        lines.append("코드 | 종목명 | 현재가 | 등락률 | 거래량")
        lines.append("---|---|---|---|---")
        for s in stocks:
            lines.append(
                f"{s['code']} | {s['name']} | {s['price']:,} | {s['change_rate']} | {s['volume']:,}"
            )

    return _append_result_meta("\n".join(lines), _kr_meta(kind="snapshot"))


@mcp.tool()
@safe_tool
@track_metrics("list_sectors")
async def list_sectors() -> str:
    """업종목록 — 네이버 증권의 업종(섹터) 목록을 가져옵니다.
    "업종별 현황", "섹터 리스트", "업종 등락률" 같은 질문에 사용합니다.
    약 79개 업종이 전일대비 등락률 순으로 정렬됩니다.
    """
    sectors = await naver_list_sectors()
    if not sectors:
        return "업종 목록을 가져올 수 없습니다."

    lines = [f"업종 목록 ({len(sectors)}개):", ""]
    lines.append("업종명 | 전일대비 | 종목수 | 상승/보합/하락")
    lines.append("---|---|---|---")
    for s in sectors:
        counts = f"{s['up_count']}/{s['flat_count']}/{s['down_count']}"
        lines.append(
            f"{s['name']} | {s['change_rate']} | {s['total_count']} | {counts}"
        )
    return "\n".join(lines)


@mcp.tool()
@safe_tool
@track_metrics("get_sector_stocks")
async def get_sector_stocks(sector_name: str, count: int = 30) -> str:
    """업종종목 — 특정 업종에 속한 종목 리스트를 가져옵니다.
    "통신장비 업종 종목", "반도체 업종", "제약 섹터 종목" 같은 질문에 사용합니다.
    업종명 부분 매칭을 지원합니다.

    ⚠️ **등락률 내림차순**으로 반환합니다. count로 자르면 그날 많이 오른 종목만 남고
    **소외된(많이 내린) 종목은 목록에서 빠집니다.** 업종 전체를 보려면 `list_sectors`로
    그 업종의 종목 수를 먼저 확인하고 count를 그 이상으로 주세요.

    Args:
        sector_name: 업종명 (예: "통신장비", "반도체", "제약")
        count: 반환할 최대 종목 수 (기본 30)
    """
    count = min(count, 50)
    result = await naver_get_sector_stocks(sector_name, count=count)
    if not result.get("sector_id"):
        return (
            f"'{sector_name}' 업종을 찾을 수 없습니다. "
            f"list_sectors로 전체 업종 목록을 먼저 확인해보세요."
        )

    stocks = result["stocks"]
    lines = [f"업종: {result['sector_name']} ({len(stocks)}개 종목)", ""]
    lines.append("코드 | 종목명 | 현재가 | 등락률 | 거래량")
    lines.append("---|---|---|---|---")
    for s in stocks:
        lines.append(
            f"{s['code']} | {s['name']} | {s['price']:,} | {s['change_rate']} | {s['volume']:,}"
        )
    return _append_result_meta("\n".join(lines), _kr_meta(kind="snapshot"))


@mcp.tool()
@safe_tool
@track_metrics("get_sector_valuation")
async def get_sector_valuation(
    sector_name: str = "",
    top_n: int = 40,
    kind: str = "sector",
    code: str = "",
) -> str:
    """업종밸류에이션 — 업종·테마의 PER·PBR·ROE **집계**와 종목별 할증/할인 위치.

    "반도체 업종 평균 PER", "이 업종에서 싼 편인가", "동종 대비 할증인가" 같은
    질문에 사용합니다. 개별 종목 지표(get_financial)로는 알 수 없는 **비교 기준**을 만듭니다.

    통계는 **중앙값이 기준**입니다. 평균은 적자 기업과 극단값(PER 500배 등)에
    쉽게 흔들려 업종 대표값으로 쓰기 어렵습니다.

    **집계는 항상 업종 전체**로 하고, top_n 은 아래 위치 표에 몇 종목을
    보여줄지만 정합니다 - top_n 을 바꿔도 중앙값은 변하지 않습니다.
    (업종이 집계 상한을 넘으면 전체 중앙값 대신 **표본 중앙값(sample_median)**
    으로 이름을 낮춰 표기합니다.)

    ⚠️ 적자 기업의 PER은 음수로 나오며 **집계에서 제외**합니다(제외 건수를 함께 표기).

    ⭐ **종목코드만 넘겨도 됩니다** — `code="005930"` 이면 그 종목의 업종을 찾아
    집계한 뒤, 그 종목이 업종 안에서 어디에 있는지 표시합니다.
    업종명을 모를 때 이 방식을 쓰세요(ETF·ETN은 소속 업종이 없어 조회되지 않습니다).

    Args:
        sector_name: 업종명 또는 테마명 (예: "반도체와반도체장비", "건설").
            code 를 주면 생략할 수 있습니다.
        top_n: 위치 표에 **표시할** 종목 수 (기본 40, 최대 80). 집계 분모가 아닙니다.
        kind: "sector"(업종, 기본) / "theme"(테마)
        code: 6자리 종목코드. 주면 업종을 자동 판정합니다.
    """
    import statistics as _st

    top_n = max(5, min(top_n, 80))

    # 종목코드로 업종을 자동 판정한다. 네이버에는 종목→업종 역방향 조회가 없어
    # 지금까지는 사용자가 업종 이름을 미리 알아야만 이 도구를 쓸 수 있었다.
    focus_code = (code or "").strip()
    focus_name = ""
    focus_per_ttm = None
    focus_sector_per = None
    if focus_code:
        sec = await naver_get_stock_sector(focus_code)
        if not sec.get("sector_name"):
            return (f"종목 {focus_code}의 업종을 찾지 못했습니다 — "
                    f"{sec.get('reason', '업종 정보 없음')}. "
                    f"업종명을 직접 넘기거나, ETF라면 get_etf_info 를 쓰세요.")
        sector_name = sec["sector_name"]
        kind = "sector"
        focus_per_ttm = sec.get("per_ttm")
        focus_sector_per = sec.get("sector_per_naver")
    if not sector_name:
        return "업종·테마 이름이나 종목코드(code) 중 하나는 필요합니다."

    # 목록은 등락률 순으로 오므로 top_n 으로 자르면 그날 오른 쪽만 남는다.
    # 먼저 업종 전체를 받아(페이지네이션 없음) 규모를 보고, 감당 가능하면 전량
    # 집계해 시가총액 순으로 세운다.
    FETCH_ALL = 300
    if kind == "theme":
        rows = await naver_get_theme_stocks(sector_name, count=FETCH_ALL, include_reason=False)
    else:
        rows = await naver_get_sector_stocks(sector_name, count=FETCH_ALL)
    # 두 도구 모두 {"stocks": [...]} 형태의 dict 를 돌려준다(리스트가 아니다).
    if isinstance(rows, dict):
        rows = rows.get("stocks") or rows.get("items") or []
    if not rows:
        return f"'{sector_name}'에 해당하는 종목을 찾지 못했습니다. list_sectors / list_themes 로 이름을 확인하세요."

    all_codes = [r.get("code") for r in rows if isinstance(r, dict) and r.get("code")]
    universe = len(all_codes)
    # 집계는 업종 전체가 기본이다(SL-11). 등락률 상위 top_n 으로 분모를 자르면
    # 그날 오른 종목 쪽으로 치우친 값이 "업종 중앙값"으로 나간다(실측: 반도체
    # 172개 중 40개만 집계). top_n 은 표시 전용이다.
    codes = all_codes[:_SECTOR_AGG_CAP]
    full_scan = universe <= len(codes)
    if focus_code and focus_code not in codes:
        codes = [focus_code] + codes[: _SECTOR_AGG_CAP - 1]
    if not codes:
        return f"'{sector_name}'에서 종목코드를 얻지 못했습니다."

    snap = await naver_scan_snapshot(codes, days=60, include_financial=True)
    if not snap:
        return f"'{sector_name}' 재무 데이터를 가져오지 못했습니다."

    def series(key, positive_only=False):
        out = []
        for s in snap:
            v = s.get(key)
            if isinstance(v, (int, float)):
                if positive_only and v <= 0:
                    continue
                out.append((s.get("code"), s.get("name"), float(v)))
        return out

    per = series("per", positive_only=True)
    pbr = series("pbr", positive_only=True)
    roe = series("roe")
    neg_per = [s for s in snap if isinstance(s.get("per"), (int, float)) and s["per"] <= 0]
    # 분기 기준을 쓰면 종목마다 최신 분기가 다를 수 있다(실적 발표 시점 차이).
    # 서로 다른 시점의 PER 을 한 표에서 비교하게 되므로 분포를 그대로 보여준다.
    from collections import Counter as _Counter
    period_counts = _Counter(s.get("fin_period") for s in snap if s.get("fin_period"))
    periods = [p for p, _ in sorted(period_counts.items(), reverse=True)]

    def stat_line(label, arr, unit=""):
        if not arr:
            return f"{label} | - | - | - | 0"
        vals = sorted(v for _, _, v in arr)
        med = _st.median(vals)
        avg = sum(vals) / len(vals)
        return (f"{label} | {med:,.2f}{unit} | {avg:,.2f}{unit} | "
                f"{vals[0]:,.2f} ~ {vals[-1]:,.2f} | {len(vals)}")

    if focus_code:
        focus_name = next((s.get("name") for s in snap if s.get("code") == focus_code), focus_code)
    lines = [f"## {sector_name} 밸류에이션 집계", ""]
    median_label = "중앙값" if full_scan else "표본 중앙값(sample_median)"
    if focus_code:
        lines.append(f"- 기준 종목: **{focus_name}({focus_code})** → 업종 `{sector_name}`")
    if full_scan:
        lines.append(f"- 집계 대상: **업종 전체 {len(snap)}개** (표본 편향 없음,"
                     f" 표시는 시총 상위 {top_n}개까지)")
    else:
        lines.append(
            f"- 집계 대상: {len(snap)}개 / 업종 전체 **{universe}개** "
            f"→ ⚠️ 집계 상한({_SECTOR_AGG_CAP})을 넘어 **표본 중앙값(sample_median)**입니다. "
            "업종 전체 중앙값이 아닙니다."
        )
    if periods:
        parts = [f"{p}({period_counts[p]}개)" for p in periods]
        lines.append(f"- 재무 기준 기간: {', '.join(parts)}")
        if len(periods) > 1:
            newest = periods[0]
            share = period_counts[newest] / max(1, sum(period_counts.values())) * 100
            lines.append(
                f"  ⚠️ 기준 시점이 종목마다 다릅니다 — 최신 `{newest}` 기준이 {share:.0f}%. "
                f"실적 발표가 늦은 종목은 오래된 분기로 계산돼 비교가 완전히 공정하지는 않습니다."
            )
    if neg_per:
        names = ", ".join(f"{s.get('name')}" for s in neg_per[:6])
        lines.append(f"- **PER 집계 제외(적자) {len(neg_per)}개**: {names}")
    lines.append("- 계산 방식: **종목별 지표의 중앙값**")
    lines.append("  · PER = 현재가 ÷ **최근 4분기 EPS 합(TTM)** — 네이버 종목 화면과 같은 기준")
    lines.append("  · PBR·ROE = 최근 확정 분기 기준")
    if focus_per_ttm or focus_sector_per:
        lines.append("")
        lines.append("### 참고 — 네이버 기준값 (계산 방식이 다릅니다)")
        lines.append("구분 | 값 | 방식")
        lines.append("---|---:|---")
        if focus_per_ttm:
            lines.append(f"{focus_name or focus_code} PER | {focus_per_ttm:,.2f} | "
                         f"네이버 표기값 — 위 표의 PER 과 **같은 기준**이라 일치해야 정상")
        if focus_sector_per:
            lines.append(f"동일업종 PER | {focus_sector_per:,.2f} | "
                         f"업종 **집계**(합산 시총÷합산 순이익 계열)")
        lines.append("")
        lines.append("⚠️ 위 집계 PER 은 **적자 기업이 있으면 크게 부풀 수 있습니다**. "
                     "분모(순이익 합계)가 깎이기 때문이며, 그러면 업종 내 대부분 종목이 "
                     "'할인'으로 보입니다. 실측에서 기계 업종은 113.39배였습니다. "
                     "아래 중앙값과 함께 보고, 두 값이 크게 다르면 그 이유부터 확인하세요.")
    lines.append("")
    lines.append(f"지표 | {median_label} | 평균 | 최소~최대 | 표본")
    lines.append("---|---:|---:|---|---:")
    lines.append(stat_line("PER(배)", per))
    lines.append(stat_line("PBR(배)", pbr))
    lines.append(stat_line("ROE(%)", roe))

    per_map = {c: v for c, _, v in per}
    med_per = _st.median([v for _, _, v in per]) if per else None
    if med_per:
        lines.append("")
        cap_map = {s.get("code"): _market_cap_eok(s.get("시가총액")) for s in snap}
        name_map = {s.get("code"): s.get("name") for s in snap}

        # 표가 길어지면 읽히지 않는다(건설 업종은 78종목). 소형주 PER 은 노이즈가
        # 크므로 시가총액 상위부터 남기고, 기준 종목은 잘려도 반드시 넣는다.
        # 몇 개를 보여줄지는 top_n 이 정한다 - 집계 분모와는 무관하다(SL-11).
        SHOW = top_n
        cands = [c for c in per_map if c in name_map]
        omitted = 0
        if len(cands) > SHOW:
            by_cap = sorted(cands, key=lambda c: (cap_map.get(c) or 0), reverse=True)
            keep = by_cap[:SHOW]
            if focus_code and focus_code in cands and focus_code not in keep:
                keep = [focus_code] + keep[:SHOW - 1]
            omitted = len(cands) - len(keep)
            cands = keep

        lines.append("")
        head = f"### {median_label}(PER {med_per:,.2f}배) 대비 위치"
        if omitted:
            head += f" — 시총 상위 {len(cands)}개 표시 ({omitted}개 생략)"
        lines.append(head)
        lines.append("종목 | PER | 중앙값 대비 | PBR | ROE(%) | 시총(억)")
        lines.append("---|---:|---:|---:|---:|---:")
        pbr_map = {c: v for c, _, v in pbr}
        roe_map = {c: v for c, _, v in roe}
        for code_ in sorted(cands, key=lambda c: per_map[c]):
            name_ = name_map.get(code_) or code_
            pv = per_map[code_]
            gap = (pv - med_per) / med_per * 100
            tag = "할인" if gap < -5 else ("할증" if gap > 5 else "비슷")
            pb = pbr_map.get(code_)
            ro = roe_map.get(code_)
            cap = cap_map.get(code_)
            mark = "**" if code_ == focus_code else ""
            lines.append(
                f"{mark}{name_}({code_}){mark} | {pv:,.2f} | {gap:+.0f}% {tag} | "
                f"{pb:,.2f}" .replace("nan", "-") if pb is not None
                else f"{mark}{name_}({code_}){mark} | {pv:,.2f} | {gap:+.0f}% {tag} | -"
            )
            lines[-1] += f" | {ro:,.2f}" if ro is not None else " | -"
            lines[-1] += f" | {cap:,.0f}" if cap is not None else " | -"

    lines.append("")
    lines.append("※ **기준이 다르면 결론이 뒤집힙니다.** 같은 종목이 작년 연간 재무로는 할증,")
    lines.append("  TTM 기준으로는 할인으로 보일 수 있습니다(실측 27.94배 vs 12.84배).")
    lines.append("  이 표의 PER 은 네이버와 같은 TTM 기준이라 종목 화면에서 대조할 수 있습니다.")
    lines.append("※ 적자이거나 확정 분기가 4개 미만이면 TTM PER 을 만들지 않습니다 —")
    lines.append("  3개를 12개월인 척 합치지 않습니다.")
    lines.append("※ 낮은 PER이 곧 저평가는 아닙니다. 이익이 정점이거나 성장이 멈출 것이라는")
    lines.append("  기대가 반영된 결과일 수 있습니다. **왜 그 숫자인지**를 함께 확인하세요.")
    lines.append("※ 적자 기업(PER 음수)은 집계에서 빠져 있습니다 — 위 제외 목록을 보세요.")
    lines.append("※ top_n 은 표시 개수만 정합니다. 집계(중앙값)는 바뀌지 않습니다.")

    # 중앙값의 실제 분모와 제외 사유(적자·결측)를 메타에 남긴다(SL-11 요구 2).
    per_numeric = sum(1 for s in snap if isinstance(s.get("per"), (int, float)))
    valuation_sample = {
        "scope": "full" if full_scan else "sample",
        "universe": universe,
        "aggregated": len(snap),
        "valid": {"per": len(per), "pbr": len(pbr), "roe": len(roe)},
        "excluded": {
            "loss_per": len(neg_per),
            "missing_per": len(snap) - per_numeric,
        },
        "median_name": "median" if full_scan else "sample_median",
    }
    return _append_result_meta(
        "\n".join(lines),
        _kr_meta(kind="filing",
                 data_completeness=(rmeta.PARTIAL
                                    if neg_per or not full_scan else rmeta.COMPLETE),
                 coverage=(None if full_scan else
                           {"requested": {"unit": "stock", "value": universe},
                            "effective": {"unit": "stock", "value": len(snap)},
                            "truncated": True, "coverage_complete": False,
                            "reason": "server_cap"}),
                 extra={"valuation_sample": valuation_sample}),
    )



@mcp.tool()
@safe_tool
@track_metrics("get_volume_ranking")
async def get_volume_ranking(
    market: str = "ALL",
    count: int = 50,
    sort_by: str = "volume",
) -> str:
    """거래량/거래대금 순위 — 상위 종목을 가져옵니다.

    "거래량 많은 종목" → sort_by="volume" (기본, 주수 기준)
    "거래대금 많은 종목"/"거래 규모 큰 종목" → sort_by="trade_value" (원 기준)

    대형 고단가 종목(삼전·하이닉스 등)은 거래량(주수)이 작아도 거래대금은 클 수 있어
    스크리닝 시에는 sort_by="trade_value"가 더 적합한 경우가 많습니다.

    Args:
        market: "KOSPI" / "KOSDAQ" / "ALL" (기본 ALL)
        count: 가져올 종목 수 (기본 50, 최대 500)
        sort_by: "volume"(거래량 주수) / "trade_value"(거래대금 원)
    """
    count = min(count, 500)
    ranks = await naver_get_volume_ranking(market=market, count=count, sort_by=sort_by)
    if not ranks:
        return f"{market} 거래량 순위를 가져올 수 없습니다."

    sort_label = "거래대금" if sort_by == "trade_value" else "거래량"
    lines = [f"{sort_label} 상위 ({market}, {len(ranks)}개, 정렬={sort_by}):", ""]
    # 헤더가 '거래대금(원)'인데 셀은 억 단위로 찍혀 헤더와 값이 어긋나 있었다.
    # 단위는 헤더 한 곳에서만 선언하고 셀은 숫자만 둔다.
    lines.append("순위 | 코드 | 종목명 | 현재가(원) | 등락률 | 거래량(주) | 거래대금(억원, 추산)")
    lines.append("---|---|---|---:|---:|---:|---:")
    for r in ranks:
        tv = r.get("trade_value_est_krw")
        tv_cell = f"{tv / 100_000_000:,.1f}" if tv is not None else "-"
        lines.append(
            f"{r['rank']} | {r['code']} | {r['name']} | {r['price']:,} | "
            f"{r['change_rate']} | {r['volume']:,} | {tv_cell}"
        )
    return _append_result_meta("\n".join(lines), _kr_meta(kind="snapshot"))


@mcp.tool()
@safe_tool
@track_metrics("get_change_ranking")
async def get_change_ranking(
    direction: str = "up",
    market: str = "ALL",
    count: int = 50,
) -> str:
    """등락률순위 — 등락률 상위/하위 종목을 가져옵니다.
    "상한가 종목", "급등주", "급락주", "상승률 상위" 같은 질문에 사용합니다.

    Args:
        direction: "up"(상승률 상위) / "down"(하락률 상위)
        market: "KOSPI" / "KOSDAQ" / "ALL" (기본 ALL)
        count: 가져올 종목 수 (기본 50, 최대 500)
    """
    count = min(count, 500)
    ranks = await naver_get_change_ranking(direction=direction, market=market, count=count)
    if not ranks:
        return f"{direction} 등락률 순위를 가져올 수 없습니다."

    # 급등주 목록은 시장경보 종목이 섞이기 가장 쉬운 자리다. 종목마다 페이지를
    # 여는 대신 경보 목록(수십 종목)만 한 번 받아 매칭한다.
    alerts = await naver_get_alert_codes()
    mine = wl.codes()

    dir_label = "상승률 상위" if direction.lower() == "up" else "하락률 상위"
    lines = [f"{dir_label} ({market}, {len(ranks)}개):", ""]
    lines.append("순위 | 코드 | 종목명 | 현재가 | 등락률 | 거래량 | 시장경보")
    lines.append("---|---|---|---:|---:|---:|---")
    flagged = 0
    mine_hits = []
    for i, r in enumerate(ranks, 1):
        marks = alerts.get(r["code"]) or []
        if marks:
            flagged += 1
        # 급등 목록에 내 관심종목이 있으면 그게 가장 먼저 알고 싶은 사실이다.
        star = "⭐ " if r["code"] in mine else ""
        if star:
            mine_hits.append(f"{r['name']}({r['code']})")
        lines.append(
            f"{i} | {r['code']} | {star}{r['name']} | {r['price']:,} | "
            f"{r['change_rate']} | {r['volume']:,} | {'⚠️ ' + ' · '.join(marks) if marks else '-'}"
        )
    if mine_hits:
        lines.append("")
        lines.append(f"⭐ **이 목록에 내 관심종목이 있습니다:** {', '.join(mine_hits)}")
    if flagged:
        lines.append("")
        lines.append(
            f"※ {flagged}개 종목이 시장경보(투자주의/경고/위험) 지정 상태입니다. "
            "일반 종목과 매매 조건·위험이 다르니 급등 사유를 함께 확인하세요."
        )
    return _append_result_meta(
        "\n".join(lines),
        _kr_meta(
            kind="snapshot",
            warnings=[f"시장경보 지정 {flagged}건 포함 — 표의 '시장경보' 열 확인"] if flagged else None,
        ),
    )


@mcp.tool()
@safe_tool
@track_metrics("get_market_cap_ranking")
async def get_market_cap_ranking(market: str = "KOSPI", count: int = 50) -> str:
    """시가총액순위 — 시가총액 상위 종목을 가져옵니다.
    "대형주", "시가총액 TOP", "코스피 대장주" 같은 질문에 사용합니다.

    Args:
        market: "KOSPI" / "KOSDAQ" (기본 KOSPI, ALL 미지원)
        count: 가져올 종목 수 (기본 50, 최대 500)
    """
    count = min(count, 500)
    ranks = await naver_get_market_cap_ranking(market=market, count=count)
    if not ranks:
        return f"{market} 시가총액 순위를 가져올 수 없습니다."

    lines = [f"시가총액 상위 ({market}, {len(ranks)}개):", ""]
    lines.append("순위 | 코드 | 종목명 | 현재가 | 등락률 | 시가총액(억원)")
    lines.append("---|---|---|---|---|---")
    for r in ranks:
        lines.append(
            f"{r['rank']} | {r['code']} | {r['name']} | {r['price']:,} | "
            f"{r['change_rate']} | {r['market_cap_billion']:,}"
        )
    return _append_result_meta("\n".join(lines), _kr_meta(kind="snapshot"))


@mcp.tool()
@safe_tool
@track_metrics("get_multi_stocks")
async def get_multi_stocks(codes: list[str]) -> str:
    """벌크조회 — 여러 종목의 기본 정보(가격/등락률/거래량)를 한 번에 가져옵니다.
    "이 종목들 현재가 보여줘", "리스트 종목 시세 한번에" 같은 질문에 사용합니다.
    개별 get_price를 여러 번 호출하는 것보다 훨씬 토큰 효율적입니다.
    스크리닝 결과 N개 종목을 비교 분석할 때 필수 도구.

    Args:
        codes: 종목코드 리스트 (최대 30개, 예: ["005930", "000660", "005380"])
    """
    if not codes:
        return "종목코드 리스트가 비어 있습니다."

    stocks = await naver_get_multi_stocks(codes)
    if not stocks:
        return "종목 정보를 가져올 수 없습니다."

    lines = [f"종목 정보 ({len(stocks)}개):", ""]
    lines.append("코드 | 종목명 | 현재가 | 전일대비 | 등락률 | 거래량")
    lines.append("---|---|---|---|---|---")
    for s in stocks:
        change = s["change"]
        change_str = f"{change:+,}" if change != 0 else "0"
        lines.append(
            f"{s['code']} | {s['name']} | {s['price']:,} | "
            f"{change_str} | {s['change_rate']} | {s['volume']:,}"
        )
    return _append_result_meta("\n".join(lines), _kr_meta(kind="snapshot"))


@mcp.tool()
@safe_tool
@track_metrics("get_multi_chart_stats")
async def get_multi_chart_stats(codes: list[str], days: int = 260) -> str:
    """차트통계벌크 — 여러 종목의 기간 **집계 통계**(현재가/최고가/최저가/낙폭/기간수익률)를 한 번에 병렬 조회.

    ⚠️ 시계열 아님 — 집계값만 반환(캔들·OHLCV는 get_chart). 개별 get_chart N번 대신 이걸로.
    ⭐ 스크리닝 필수: "52주 고점 대비 -30% 종목" 등 drawdown_pct·period_return_pct 필터에 사용.

    Args:
        codes: 종목코드 리스트 (최대 100개)
        days: 과거 조회 일수 (기본 260 = 52주)
    """
    if not codes:
        return "종목코드 리스트가 비어 있습니다."
    if days < 10:
        days = 10
    if days > 500:
        days = 500

    stats = await naver_get_multi_chart_stats(codes, days=days)
    if not stats:
        return "차트 통계를 가져올 수 없습니다."
    _cal_stats = krx_calendar()

    # 요청한 기간만큼 데이터가 실제로 있는지는 종목마다 다르다. 신규 상장·거래정지
    # 종목은 260일을 요청해도 몇 봉밖에 없는데, 헤더에 "최근 260일 집계"라고만 쓰면
    # 7일치 고점을 52주 고점으로 읽게 된다. 부족한 종목은 행에 표시하고 경고로 올린다.
    short = [s for s in stats
             if isinstance(s.get("bars_count"), int) and s["bars_count"] < days * 0.6]
    lines = [f"차트 통계 ({len(stats)}개 종목, 요청 {days}일 집계):", ""]
    lines.append("코드 | 현재가 | 최고가(날짜) | 최저가(날짜) | 고점대비낙폭 | 기간수익률 | 봉수")
    lines.append("---|---|---|---|---|---|---")
    for s in stats:
        bars = s.get("bars_count")
        bars_s = f"{bars}" if isinstance(bars, int) else "-"
        if isinstance(bars, int) and bars < days * 0.6:
            bars_s = f"⚠️{bars}"
        lines.append(
            f"{s['code']} | {s['current_price']:,} | "
            f"{s['high']:,}({s['high_date']}) | "
            f"{s['low']:,}({s['low_date']}) | "
            f"{s['drawdown_pct']:+.1f}% | "
            f"{s['period_return_pct']:+.1f}% | "
            f"{bars_s}"
        )
    lines.append("")
    lines.append(
        "※ 기간 **집계값**만 반환(시계열 OHLCV 아님). "
        "봉별 시계열이 필요하면 get_chart를 별도 호출."
    )
    lines.append(
        "※ **봉수**는 이 통계가 실제로 몇 거래일로 계산됐는지입니다. "
        "요청 기간보다 훨씬 적으면(⚠️) 고점·저점과 낙폭이 그 짧은 구간 기준이라, "
        "'52주 고점' 같은 표현을 쓰면 안 됩니다."
    )
    lines.append(f"※ {CORPORATE_ACTION_NOTE}")
    total_excluded = sum(s.get("excluded_bars", 0) or 0 for s in stats)
    warns = []
    if total_excluded:
        warns.append(
            f"거래정지 placeholder 등 비정상 봉 {total_excluded}개를 집계에서 "
            "제외했습니다. 해당 날짜의 고가·저가는 실제 값이 아닙니다."
        )
    if short:
        names = ", ".join(f"{s['code']}({s['bars_count']}봉)" for s in short[:8])
        warns += [
            f"요청 {days}일보다 데이터가 크게 짧은 종목 {len(short)}건: {names}"
            " — 신규 상장·거래정지 가능성. 낙폭·기간수익률이 짧은 구간 기준입니다."
        ]
    per_entity_as_of = {}
    for s in stats:
        d = rmeta.normalize_day(s.get("current_date"))
        if s.get("code") and d:
            per_entity_as_of[s["code"]] = d
    return _append_result_meta(
        "\n".join(lines),
        _kr_meta(kind="bars",
                 data_as_of=max(per_entity_as_of.values()) if per_entity_as_of else None,
                 data_completeness=rmeta.PARTIAL if short else rmeta.COMPLETE,
                 warnings=warns or None,
                 bar_state=_merge_bar_states([
                     _bar_state(timeframe="day", rows=[{"date": s.get("current_date")}],
                                market_calendar=_cal_stats, calendar_fallback=True)
                     for s in stats if s.get("current_date")
                 ]),
                 price_adjustment=price_adjustment_meta(),
                 extra={"per_entity_data_as_of": per_entity_as_of}),
    )


@mcp.tool()
@safe_tool
@track_metrics("get_indicators")
async def get_indicators(
    code: str,
    days: int = 260,
    include: list[str] | None = None,
    timeframe: str = "day",
    params: dict | None = None,
) -> str:
    """기술지표 — 이평선·RSI·MACD·볼린저·스토캐스틱 등 종합 판정 (JSON).

    스크리닝·조건 필터·상태 판정 등 **숫자 비교가 필요할 때만** 호출.
    차트 시각화용 아님(시각화는 `get_chart`). OHLCV 대신 판정 결과만 반환해 토큰 절약.
    반환값의 라벨 필드(`phase_label`, `type_label`, `position` 등)는 그대로 인용할 것.

    키 이름이 곧 정의입니다 — 임의로 바꿔 읽지 마세요:
      `volume.latest`(+`latest_date`)  마지막 **봉**의 거래량. '오늘'이 아님
      `volume.avg_20b` / `ratio_vs_avg_20b`  20**봉**(거래일) 평균 대비
      `volume.trade_value_est_krw`  종가×거래량 **추산**. 실제 거래대금과 다름
      `volume.volume_rank_252b`  252봉 중 순위, **1이 최다**
      `position.bars_since_high/low`  달력일이 아니라 **봉 개수**.
        달력일이 필요하면 `high_date`/`low_date`로 직접 계산하세요.
    `_meta.data_basis`가 `in_progress_bar`면 마지막 봉이 미마감이라 이 판정들은
    장 마감 시 달라질 수 있습니다.

    Args:
        code: 종목코드 (예: "005930")
        days: 조회 일수 (기본 260, 30~500). 구조 분석 지표는 500+ 권장.
        include: 지표 키. 기본 ["ma", "ma_phase", "volume", "candle"].
            스냅샷: ma ma_phase ma_slope ma_cross rsi macd bollinger stochastic obv volume position candle
            구조: support_resistance volume_profile price_channel
        timeframe: "day"/"week"/"month" (분봉 미지원)
        params: 비표준 파라미터 오버라이드(사용자 명시 요청 시만). 예: {"rsi":{"period":21}}
    """
    if not code:
        return "종목코드가 필요합니다."
    if include is None:
        include = ["ma", "ma_phase", "volume", "candle"]
    unknown = [k for k in include if k not in AVAILABLE_INDICATORS]
    if unknown:
        return (
            f"지원하지 않는 지표: {unknown}\n"
            f"사용 가능: {AVAILABLE_INDICATORS}"
        )
    days = max(30, min(days, 500))

    raw_ohlcv = await get_ohlcv(code, timeframe=timeframe, count=days)
    if not raw_ohlcv:
        return f"차트 데이터를 가져올 수 없습니다: {code}"
    ohlcv, _excluded = split_valid_bars(raw_ohlcv)
    exclusion_info, exclusion_warns = _bar_exclusion_info(_excluded)
    if not ohlcv:
        return (f"종목 {code}: 조회된 봉 전체가 거래정지 placeholder 등 비정상이라 "
                "지표를 계산할 수 없습니다.")

    result = compute_indicators(ohlcv, include, params=params)
    ind_errors = _indicator_error_list(result)
    ind_cov = _indicator_coverage(
        include=include, available_bars=len(ohlcv), params=params
    )
    payload = {
        "code": code,
        "timeframe": timeframe,
        "days": days,
        "indicators": result,
        # JSON 도구라 텍스트 봉투 대신 payload 안에 넣는다. 장중이면 data_basis가
        # in_progress_bar가 되어 "이 판정은 마감 때 뒤집힐 수 있다"가 명시된다.
        "_meta": _kr_meta(
            kind="bars", code=code, data_as_of=ohlcv[-1].get("date"),
            bar_state=_bar_state(
                timeframe=timeframe, rows=ohlcv, market_calendar=krx_calendar()
            ),
            price_adjustment=price_adjustment_meta(),
            data_completeness=(rmeta.PARTIAL
                               if ind_cov["insufficient"] or ind_errors
                               else rmeta.COMPLETE),
            coverage=({"truncated": False, "coverage_complete": False,
                       "reason": "indicator_error"} if ind_errors else None),
            warnings=([
                "봉이 모자라 계산되지 않은 지표: " + ", ".join(ind_cov["insufficient"])
                + f" (보유 {ind_cov['available_bars']}봉). 값이 빠진 것은 "
                "'그런 신호가 없다'가 아니라 '아직 계산할 이력이 없다'는 뜻입니다."
            ] if ind_cov["insufficient"] else [])
            + ([
                "계산에 실패한 지표: "
                + ", ".join(f"{e['indicator']}({e['error']})" for e in ind_errors)
                + " - 값이 없는 것이지 신호가 없는 것이 아닙니다."
            ] if ind_errors else [])
            + exclusion_warns or None,
            extra={"indicator_coverage": ind_cov,
                   **({"indicator_errors": ind_errors} if ind_errors else {}),
                   **({"bar_exclusions": exclusion_info} if exclusion_info else {})},
        ),
    }
    _mark_intraday_volume(result, payload["_meta"])
    return json.dumps(payload, ensure_ascii=False, separators=(",", ":"))


def _market_cap_eok(value) -> float | None:
    """'1,543조 4,176억원' → 억 단위 숫자. 못 읽으면 None(0 아님).

    네이버는 시가총액을 사람이 읽는 문자열로 준다. 정렬·비교에 쓰려면 숫자가
    필요한데, 실패를 0으로 채우면 시총 최하위로 내려가 순서가 뒤집힌다.
    """
    if value is None:
        return None
    txt = str(value).replace(",", "").replace(" ", "")
    if not txt:
        return None
    total = 0.0
    m = re.search(r"(\d+(?:\.\d+)?)조", txt)
    if m:
        total += float(m.group(1)) * 10000
    m2 = re.search(r"(\d+(?:\.\d+)?)억", txt)
    if m2:
        total += float(m2.group(1))
    return total if total > 0 else None


# 지표 하나를 계산하려면 최소 몇 봉이 있어야 하는가. 값은 _indicators.py 의
# 가드(len(df) < N)와 같은 수여야 한다. 모자라면 그 지표는 조용히 사라지는데,
# 사라진 자리를 읽는 쪽은 "이평선이 안 만들어졌다"가 아니라 "그런 신호가 없다"로
# 읽는다. 주봉 104개로 ma120 을 부르면 정확히 그 일이 벌어진다.
_INDICATOR_MIN_BARS = {
    "ma": lambda o: {f"ma{n}": n for n in o.get("periods") or (5, 20, 60, 120, 240)},
    "ma_phase": lambda o: {f"ma{n}": n for n in (5, 20, 60, 120)},
    "ma_slope": lambda o: {"ma20_slope": 30, "ma60_slope": 80, "ma120_slope": 140},
    "ma_cross": lambda o: {"ma20_60_cross": 90, "ma60_120_cross": 180},
    "rsi": lambda o: {"rsi": int(o.get("period", 14)) + 1},
    "macd": lambda o: {"macd": int(o.get("slow", 26)) + int(o.get("signal", 9))},
    "bollinger": lambda o: {"bollinger": int(o.get("period", 20))},
    "stochastic": lambda o: {"stochastic": int(o.get("k_period", 12))
                             + int(o.get("slow_k_period", 5))
                             + int(o.get("d_period", 3))},
    "obv": lambda o: {"obv": int(o.get("window", 20)) + 1},
    "volume": lambda o: {"volume_avg_20b": 20, "volume_rank_252b": 252},
    "position": lambda o: {"position_52w": 252},
    "candle": lambda o: {"candle": 2},
    "support_resistance": lambda o: {"support_resistance": int(o.get("window", 10)) * 2 + 1},
    "volume_profile": lambda o: {"volume_profile": 20},
    "price_channel": lambda o: {"price_channel": int(o.get("period", 20))},
}


def _indicator_coverage(
    *, include: list[str] | None, available_bars: int, params: dict | None = None
) -> dict:
    """요청한 지표가 이 봉 수로 실제 계산되는지. 부족 항목을 이름으로 돌려준다."""
    required: dict[str, int] = {}
    for key in include or []:
        builder = _INDICATOR_MIN_BARS.get(key)
        if builder is None:
            continue
        try:
            required.update(builder((params or {}).get(key) or {}))
        except Exception:
            continue
    return {
        "available_bars": available_bars,
        "required_bars": dict(sorted(required.items())),
        "insufficient": sorted(k for k, n in required.items() if available_bars < n),
    }


def _bar_exclusion_info(excluded: list[dict]) -> tuple[dict | None, list[str]]:
    """제외한 placeholder 봉을 메타·경고로 만든다. 없으면 (None, []).

    거래정지 구간 처리 정책은 exclude 로 고정한다 - 고가·저가가 0인 봉의
    빈 자리를 지어내는(carry) 순간 그 값이 확정치로 읽힌다. 무엇을 왜 뺐는지는
    메타와 경고 양쪽에 남긴다.
    """
    if not excluded:
        return None, []
    bars = [{"date": rmeta.normalize_day(e.get("date")) or str(e.get("date")),
             "reason": e.get("reason")} for e in excluded]
    info = {"policy": "exclude", "count": len(bars), "bars": bars[:10]}
    dates = ", ".join(b["date"] for b in bars[:5])
    warns = [
        f"거래정지 placeholder 등 비정상 봉 {len(bars)}개를 계산에서 제외했습니다"
        f"({dates}{' 외' if len(bars) > 5 else ''}). "
        "해당 날짜의 고가·저가·거래량은 실제 값이 아니라 원천의 빈자리 표시입니다."
    ]
    return info, warns


def _indicator_error_list(result: dict, code: str | None = None) -> list[dict]:
    """지표 결과에서 {"error": ...} 로 삼켜진 실패를 끄집어낸다.

    compute_indicators 는 지표 하나가 죽어도 나머지를 살리려고 오류를 값 안에
    넣는데, 그대로 두면 응답은 complete 가 된다. "지표가 오류인데 완전"은
    이 계약이 없애려는 거짓말이다.
    """
    out = []
    for key, val in (result or {}).items():
        if isinstance(val, dict) and isinstance(val.get("error"), str):
            item = {"indicator": key, "error": val["error"]}
            if code is not None:
                item = {"code": code, **item}
            out.append(item)
    return out


def _now_kst():
    """지금 시각(KST). 봉 마감 판정의 기준점이라 한 군데로 모아 둔다."""
    return _dt.datetime.now(KST)


def _bar_state(
    *,
    timeframe: str,
    rows: list[dict] | None,
    now=None,
    market_calendar=None,
    calendar_fallback: bool = False,
) -> dict | None:
    """마지막 봉이 끝났는지, 끝난 봉 중 가장 최근 것은 언제인지.

    `data_basis=in_progress_bar`는 **장중일 때만** 붙는다. 그래서 수요일 저녁의
    주봉처럼 "장은 닫혔는데 봉은 안 끝난" 경우를 못 잡는다. 그 주봉으로 계산한
    이평·RSI는 금요일에 값이 바뀌는데 응답에는 확정치로 적혀 나간다.

    값 자체는 건드리지 않는다. 상태만 계산해 돌려주고, 경고를 붙일지는 도구가
    정한다. 달력을 못 구했거나 조회가 깨지면 예외를 올리지 않고 `None`으로
    "모른다"고 적는다 - 추측한 값이 확정치로 읽히는 게 더 나쁘다.
    """
    dates = [
        rmeta.normalize_day(r.get("date"))
        for r in (rows or [])
        if isinstance(r, dict)
    ]
    dates = [d for d in dates if d]
    if not dates:
        return None

    last = dates[-1]
    state = {
        "timeframe": timeframe,
        "last_bar_date": last,
        "last_bar_complete": None,
        "last_completed_bar_date": None,
        "calculation_includes_incomplete": None,
    }
    if market_calendar is None:
        return state

    try:
        moment = now or _now_kst()
        closed = market_calendar.is_period_closed(
            _dt.date.fromisoformat(last), timeframe, moment
        )
    except Exception:
        return state
    if closed is None:
        return state

    state["last_bar_complete"] = bool(closed)
    state["calculation_includes_incomplete"] = not closed
    if closed:
        state["last_completed_bar_date"] = last
        return state

    # 직전 원소를 그냥 집으면 안 된다. 같은 날짜가 겹쳐 들어오면(배치에서 종목별
    # 마지막 날짜를 모아 넘기던 경우) 확정 봉이 자기 자신이 된다. 실제로 마감된
    # 구간에 속한 가장 최근 봉을 찾는다.
    for prior in reversed(dates[:-1]):
        try:
            if market_calendar.is_period_closed(
                _dt.date.fromisoformat(prior), timeframe, moment
            ):
                state["last_completed_bar_date"] = prior
                break
        except Exception:
            break

    # 집계값만 오고 시계열이 없는 도구는 직전 봉을 행에서 찾을 수 없다. 그때만
    # 거래소 달력에서 직전 마감 구간을 읽는다(봉을 지어내는 게 아니다).
    if state["last_completed_bar_date"] is None and calendar_fallback:
        try:
            prev = market_calendar.previous_closed_period_end(
                _dt.date.fromisoformat(last), timeframe, moment
            )
            if prev is not None:
                state["last_completed_bar_date"] = prev.isoformat()
        except Exception:
            pass
    return state


def _merge_bar_states(states: list[dict | None]) -> dict | None:
    """종목별 봉 상태를 배치 하나로 합친다.

    기준은 가장 최근 봉을 가진 종목이다. 한 종목이라도 미완성 봉을 물고 있으면
    이 배치의 계산에는 미완성 봉이 섞인 것이다. 확정 봉 날짜는 종목들이 실제로
    가진 확정 봉 중 가장 최근 것을 쓴다.
    """
    live = [s for s in states if s]
    if not live:
        return None
    newest = max(live, key=lambda s: s.get("last_bar_date") or "")
    done = [s.get("last_completed_bar_date") for s in live if s.get("last_completed_bar_date")]
    completes = [s.get("last_bar_complete") for s in live]
    if any(c is False for c in completes):
        complete = False
    elif any(c is None for c in completes):
        complete = None
    else:
        complete = True
    return {
        "timeframe": newest.get("timeframe"),
        "last_bar_date": newest.get("last_bar_date"),
        "last_bar_complete": complete,
        "last_completed_bar_date": max(done) if done else None,
        "calculation_includes_incomplete": None if complete is None else not complete,
    }


def _mark_intraday_volume(indicators: dict, meta: dict) -> None:
    """장중이면 거래량 파생 필드에 '미완성' 표식을 단다.

    장이 열려 있는 동안 마지막 봉의 거래량은 아직 쌓이는 중이라
    ratio_vs_avg_20b·volume_rank_252b가 실제보다 훨씬 낮게 나온다
    (오전 조회 시 대형주가 252봉 중 **252위**로 찍히는 식).
    _meta의 data_basis만으로는 필드를 그대로 옮겨 적는 소비자가
    "거래량이 바닥"이라는 틀린 사실을 읽게 된다 — 값 옆에 표식을 붙인다.
    """
    if not isinstance(meta, dict) or meta.get("data_basis") != rmeta.BASIS_IN_PROGRESS_BAR:
        return
    vol = indicators.get("volume") if isinstance(indicators, dict) else None
    if isinstance(vol, dict) and vol.get("latest") is not None:
        vol["intraday_incomplete"] = True
        vol["intraday_note"] = (
            "장중 미완성 봉 — 오늘 거래량이 아직 쌓이는 중이라 "
            "ratio_vs_avg_20b·volume_rank_252b가 실제보다 낮게 나옵니다. "
            "거래량 기준으로 판단하려면 장 마감 후 다시 조회하세요."
        )


@mcp.tool()
@safe_tool
@track_metrics("get_indicators_bulk")
async def get_indicators_bulk(
    codes: list[str],
    days: int = 260,
    include: list[str] | None = None,
    timeframe: str = "day",
    params: dict | None = None,
) -> str:
    """기술지표벌크 — 여러 종목(최대 100개)의 지표를 병렬 판정. 스크리닝 핵심.

    ⚠️ 시계열·캔들 아님 — 집계 판정값만(시각화는 get_chart). get_indicators N번 대신 이걸로.

    Args:
        codes: 종목코드 리스트 (최대 100개)
        days: 조회 일수 (기본 260)
        include: 지표 키 (기본 ["ma_phase","volume"]). get_indicators 참조.
        timeframe: "day"/"week"/"month"
        params: 지표 파라미터 오버라이드(전 종목 공통). get_indicators 참조.
    """
    if not codes:
        return "종목코드 리스트가 비어 있습니다."
    if include is None:
        include = ["ma_phase", "volume"]
    unknown = [k for k in include if k not in AVAILABLE_INDICATORS]
    if unknown:
        return (
            f"지원하지 않는 지표: {unknown}\n"
            f"사용 가능: {AVAILABLE_INDICATORS}"
        )
    codes = codes[:100]
    days = max(30, min(days, 500))

    bars_seen: list[int] = []
    cal = krx_calendar()

    excluded_by_code: dict[str, list[dict]] = {}

    async def one(code: str) -> tuple[str, dict, dict | None]:
        try:
            raw = await get_ohlcv(code, timeframe=timeframe, count=days)
            if not raw:
                return code, {"error": "OHLCV 없음"}, None
            ohlcv, _exc = split_valid_bars(raw)
            if _exc:
                excluded_by_code[code] = _exc
            if not ohlcv:
                return code, {"error": f"봉 {len(_exc)}개 전체가 비정상 "
                                       "(거래정지 placeholder 등 입력 데이터 이상)"}, None
            bars_seen.append(len(ohlcv))
            # 봉 상태는 **그 종목의 시계열로** 계산해야 한다. 종목별 마지막
            # 날짜만 모아 한 시계열인 척 넘기면, 두 종목이 같은 미완성 봉을
            # 가질 때 확정 봉이 자기 자신으로 나온다.
            state = _bar_state(timeframe=timeframe, rows=ohlcv, market_calendar=cal)
            return code, compute_indicators(ohlcv, include, params=params), state
        except Exception as e:
            return code, {"error": f"{type(e).__name__}: {e}"}, None

    results = await asyncio.gather(*[one(c) for c in codes])
    failed = sum(1 for _, data, _ in results if "error" in data)
    # 지표 하나가 죽어도 {"error": ...} 로 삼켜져 응답이 complete 가 되던
    # 자리다(실측: 이월드 position ZeroDivisionError 인데 complete·경고 없음).
    ind_errors: list[dict] = []
    for code, data, _ in results:
        if "error" not in data:
            ind_errors += _indicator_error_list(data, code=code)
    bulk_exclusions = None
    if excluded_by_code:
        bulk_exclusions = {
            "policy": "exclude",
            "count": sum(len(v) for v in excluded_by_code.values()),
            "by_code": {c: len(v) for c, v in excluded_by_code.items()},
        }
    # 종목마다 이력 길이가 다르다. 가장 짧은 종목을 기준으로 잡아야 "전 종목
    # ma120 이 나왔다"는 과대 주장을 하지 않는다.
    ind_cov = _indicator_coverage(
        include=include, available_bars=min(bars_seen) if bars_seen else 0, params=params
    )
    # 봉 상태는 종목별로 계산한 뒤 합친다. 한 종목이라도 미완성 봉을 물고
    # 있으면 이 배치의 계산에는 미완성 봉이 섞인 것이다.
    batch_bar_state = _merge_bar_states([s for _, _, s in results])
    per_entity_as_of = {
        code: s["last_bar_date"]
        for code, _, s in results
        if s and s.get("last_bar_date")
    }
    payload = {
        "timeframe": timeframe,
        "days": days,
        "include": include,
        "count": len(results),
        "results": {code: data for code, data, _ in results},
        "_meta": _kr_meta(
            kind="bars",
            data_as_of=max(per_entity_as_of.values()) if per_entity_as_of else None,
            data_completeness=(rmeta.PARTIAL
                               if failed or ind_cov["insufficient"] or ind_errors
                               else rmeta.COMPLETE),
            coverage=({"truncated": False, "coverage_complete": False,
                       "reason": "indicator_error"} if ind_errors else None),
            warnings=([f"{failed}개 종목 조회 실패 — results의 error 필드 확인."] if failed else [])
                     + ([
                         "봉이 모자라 계산되지 않은 지표: " + ", ".join(ind_cov["insufficient"])
                         + f" (가장 짧은 종목 {ind_cov['available_bars']}봉 기준)."
                     ] if ind_cov["insufficient"] else [])
                     + ([
                         "계산에 실패한 지표 "
                         + ", ".join(f"{e['code']}:{e['indicator']}" for e in ind_errors[:8])
                         + " - 값이 없는 것이지 신호가 없는 것이 아닙니다."
                     ] if ind_errors else [])
                     + ([
                         f"거래정지 placeholder 등 비정상 봉 {bulk_exclusions['count']}개를 "
                         "계산에서 제외했습니다(by_code 참고)."
                     ] if bulk_exclusions else []) or None,
            bar_state=batch_bar_state,
            price_adjustment=price_adjustment_meta(),
            extra={"indicator_coverage": ind_cov,
                   "per_entity_data_as_of": per_entity_as_of,
                   **({"indicator_errors": ind_errors} if ind_errors else {}),
                   **({"bar_exclusions": bulk_exclusions} if bulk_exclusions else {})},
        ),
    }
    for _data in payload["results"].values():
        _mark_intraday_volume(_data, payload["_meta"])
    return json.dumps(payload, ensure_ascii=False, separators=(",", ":"))


@mcp.tool()
@safe_tool
@track_metrics("export_to_excel")
async def export_to_excel(
    data_type: str,
    code: str = "",
    days: int = 180,
    filename: str = "",
) -> str:
    """엑셀내보내기 — 단일 종목의 데이터를 Excel 파일로 저장합니다.

    Gemini/GPT 같은 다른 AI에 파일 업로드로 넘기거나,
    엑셀에서 직접 분석/차트 작성할 때 사용합니다.

    Args:
        data_type: "chart"(일봉 OHLCV) / "flow"(투자자별 수급) / "financial"(재무지표)
        code: 종목코드 6자리 (예: "005930")
        days: chart/flow의 경우 과거 일수 (기본 180)
        filename: 파일명 (비우면 자동 생성)

    Returns:
        저장된 파일 경로
    """
    if not code:
        return "종목코드가 필요합니다."

    if data_type == "chart":
        data = await get_ohlcv(code, "day", days)
        if not data:
            return f"차트 데이터를 가져올 수 없습니다: {code}"
        df = pd.DataFrame(data)
        prefix = f"chart_{code}"
        sheet = "OHLCV"

    elif data_type == "flow":
        try:
            data = await get_investor_flow(code, days)
        except NaverParseError as e:
            return (
                f"수급 표를 읽지 못했습니다 (데이터 없음이 아니라 파싱 실패): {code}\n"
                f"원인: {e}"
            )
        if not data:
            return f"수급 데이터를 가져올 수 없습니다: {code}"
        df = pd.DataFrame(data)
        prefix = f"flow_{code}"
        sheet = "Investor Flow"

    elif data_type == "financial":
        data = await get_financials(code)
        if not data:
            return f"재무지표를 가져올 수 없습니다: {code}"
        # dict → DataFrame 변환
        rows = [{"항목": k, "값": str(v)} for k, v in data.items() if k not in ("code",)]
        df = pd.DataFrame(rows)
        prefix = f"financial_{code}"
        sheet = "Financial"

    else:
        return f"지원하지 않는 data_type: {data_type}. 'chart', 'flow', 'financial' 중 선택."

    fname = filename or generate_filename(prefix)
    if not fname.endswith(".xlsx"):
        fname += ".xlsx"
    file_path = get_snapshot_dir() / fname
    saved = save_dataframe_to_excel(df, file_path, sheet_name=sheet)

    return (
        f"✓ Excel 파일 저장 완료\n"
        f"경로: {saved}\n"
        f"행 수: {len(df)}\n"
        f"컬럼: {', '.join(df.columns)}\n\n"
        f"💡 이 파일을 Gemini/ChatGPT에 업로드하면 다른 AI에서도 분석할 수 있어요."
    )


@mcp.tool()
@safe_tool
@track_metrics("scan_to_excel")
async def scan_to_excel(
    codes: list[str],
    fields: list[str] | None = None,
    days: int = 260,
    include_financial: bool = True,
    filename: str = "",
) -> str:
    """시장스캔 — **원하는 항목만 골라** 여러 종목을 Excel 한 장으로 저장.

    로컬 캐시 패턴: 한 번 스캔 → 이후 query_excel(파일경로, 조건)로 즉시 반복 필터링.

    ## fields — 넣을 항목을 직접 고른다

    | 값 | 들어가는 열 | 추가 조회 비용 |
    |---|---|---|
    | `price` | 현재가·거래량 | 기본 (항상 수집) |
    | `chart` | 기간 최고/최저·낙폭·기간수익률·평균거래량·집계봉수 | 기본 (항상 수집) |
    | `financial` | PER·PBR·EPS·BPS·ROE·배당수익률·시가총액·재무기준기간 | 종목당 1회 |
    | `flow` | 기관·외국인 순매매 누적(20일) | 30종목당 1회 |
    | `sector` | 업종 | 종목당 1회 |
    | `indicators` | 이평배열·RSI·거래량배수 | 100종목당 1회 |

    생략하면 `["price", "chart", "financial"]` 이 들어갑니다(기존과 동일).
    **필요 없는 항목을 빼면 그만큼 조회가 줄어 빨라집니다.**

    ## 무엇을 넣을지 사용자와 정하는 법

    사용자가 "엑셀로 만들어줘"라고만 하면 **바로 기본 세트로 만들지 말고,
    무엇에 쓸 파일인지 한 번 물어보세요.** 목적에 따라 넣을 항목이 달라집니다.

    > "어떤 걸 보시려고 하시나요? 아래처럼 목적을 말씀해 주시면 맞춰 담아드릴게요.
    >  — 싼 종목 고르기 / 수급 흐름 보기 / 차트 흐름 보기 / 업종 안에서 비교 / 전부"

    **목적별 추천 구성**

    | 사용자가 하려는 일 | fields |
    |---|---|
    | 싸게 거래되는 종목 고르기 | `["price", "financial", "sector"]` |
    | 외국인·기관이 사는 종목 보기 | `["price", "flow"]` |
    | 많이 빠진 종목 훑기 | `["price", "chart"]` |
    | 추세·과열 상태 보기 | `["price", "chart", "indicators"]` |
    | 같은 업종끼리 비교 | `["price", "financial", "sector"]` |
    | 일단 다 담기 (느림) | `["price","chart","financial","flow","sector","indicators"]` |

    이미 사용자가 목적을 밝혔다면 다시 묻지 말고 위 표대로 골라 담으세요.
    종목이 30개를 넘으면 항목이 많을수록 눈에 띄게 느려지므로,
    그때는 필요한 것만 담고 나중에 더 필요하면 다시 부르는 편이 낫다고 안내하세요.

    Args:
        codes: 종목코드 리스트 (최대 500개)
        fields: 넣을 항목. 위 표의 값들 중에서 고릅니다. 생략 시 기본 세트.
        days: 차트 통계 과거 일수 (기본 260 = 52주)
        include_financial: (구버전 호환) fields 를 안 줬을 때만 적용됩니다.
        filename: 파일명 (비우면 자동 생성)
    """
    if not codes:
        return "종목코드 리스트가 비어 있습니다."

    ALL_FIELDS = ("price", "chart", "financial", "flow", "sector", "indicators")
    if fields:
        chosen = [f for f in fields if f in ALL_FIELDS]
        unknown = [f for f in fields if f not in ALL_FIELDS]
        if not chosen:
            return (f"알 수 없는 항목입니다: {', '.join(fields)}"
                    f"\n쓸 수 있는 값: {', '.join(ALL_FIELDS)}")
    else:
        chosen = ["price", "chart"] + (["financial"] if include_financial else [])
        unknown = []
    want = set(chosen) | {"price", "chart"}   # 뼈대는 항상 남긴다

    rows = await naver_scan_snapshot(codes, days=days,
                                     include_financial="financial" in want)
    if not rows:
        return "데이터를 수집하지 못했습니다."
    by_code = {r.get("code"): r for r in rows}

    # ── 선택 항목만 추가로 수집한다 ──────────────────────────────
    if "flow" in want:
        async def one_flow(c: str):
            try:
                return c, await get_investor_flow(c, 20)
            except Exception:
                return c, None
        for code_, series in await asyncio.gather(*[one_flow(c) for c in by_code]):
            r = by_code.get(code_)
            if not r or not isinstance(series, list) or not series:
                continue
            # 결측을 0으로 채우면 '순매매 0'과 구분이 안 된다 — 값이 있는 날만 더한다.
            inst = [d.get("institutional") for d in series if d.get("institutional") is not None]
            forg = [d.get("foreign") for d in series if d.get("foreign") is not None]
            if inst:
                r["inst_net"] = int(sum(inst))
            if forg:
                r["foreign_net"] = int(sum(forg))

    if "sector" in want:
        async def one_sector(c):
            try:
                return c, (await naver_get_stock_sector(c)).get("sector_name")
            except Exception:
                return c, None
        for c, nm in await asyncio.gather(*[one_sector(c) for c in by_code]):
            if nm and c in by_code:
                by_code[c]["sector"] = nm

    if "indicators" in want:
        async def one_ind(c: str):
            try:
                ohlcv = await get_ohlcv(c, timeframe="day", count=max(days, 120))
                return c, compute_indicators(ohlcv, ["ma_phase", "rsi", "volume"])
            except Exception:
                return c, None
        for code_, ind in await asyncio.gather(*[one_ind(c) for c in by_code]):
            if True:
                r = by_code.get(code_)
                if not r or not isinstance(ind, dict):
                    continue
                ph = (ind.get("ma_phase") or {}).get("phase_label")
                if ph:
                    r["ma_phase"] = ph
                rsi = (ind.get("rsi") or {}).get("value")
                if rsi is not None:
                    r["rsi"] = rsi
                vr = (ind.get("volume") or {}).get("ratio_vs_avg_20b")
                if vr is not None:
                    r["vol_ratio"] = vr

    # ── 고른 항목의 열만 남긴다 ─────────────────────────────────
    FIELD_COLUMNS = {
        "price": ["current_price", "volume"],
        "chart": ["high", "high_date", "low", "low_date", "drawdown_pct",
                  "recovery_pct", "period_return_pct", "avg_volume", "bars_count"],
        "financial": ["per", "per_basis", "pbr", "eps", "bps", "roe",
                      "배당수익률", "시가총액", "fin_period"],
        "flow": ["inst_net", "foreign_net"],
        "sector": ["sector"],
        "indicators": ["ma_phase", "rsi", "vol_ratio"],
    }
    keep = ["code", "name"]
    for f in ALL_FIELDS:
        if f in want:
            keep += FIELD_COLUMNS[f]

    df = pd.DataFrame(list(by_code.values()))
    df = df[[c for c in keep if c in df.columns]]

    fname = filename or generate_filename(f"snapshot_{len(df)}stocks")
    if not fname.endswith(".xlsx"):
        fname += ".xlsx"
    file_path = get_snapshot_dir() / fname

    saved = save_dataframe_to_excel(
        df, file_path, sheet_name="Snapshot",
        metadata={"days": days, "fields": ", ".join(chosen),
                  "requested_codes": len(codes)},
    )

    msg = [f"✓ 스냅샷 저장 완료 ({len(df)}개 종목)",
           f"경로: {saved}",
           f"넣은 항목: {', '.join(chosen)}",
           f"컬럼: {', '.join(df.columns)}"]
    if unknown:
        msg.append(f"⚠️ 알 수 없어 무시한 항목: {', '.join(unknown)} "
                   f"(쓸 수 있는 값: {', '.join(ALL_FIELDS)})")
    msg.append("")
    msg.append("💡 다른 항목이 더 필요하면 fields 를 바꿔 다시 부르세요. "
               "예: fields=[\"price\", \"flow\", \"sector\"]")
    # 파일에 없는 컬럼으로 예시를 주면 그 조건이 조용히 무시된다 —
    # 실제로 들어간 열 중에서 골라 안내한다.
    sample = None
    for key, cond in (("per", "{'per': {'min': 0, 'max': 15}}"),
                      ("drawdown_pct", "{'drawdown_pct_max': -30}"),
                      ("foreign_net", "{'foreign_net_min': 0}"),
                      ("period_return_pct", "{'period_return_pct_min': 0}")):
        if key in df.columns:
            sample = cond
            break
    if sample:
        msg.append(f"💡 조건 검색: query_excel(file_path='{saved}', filters={sample})")
    return "\n".join(msg)



@mcp.tool()
@safe_tool
@track_metrics("save_analysis_to_excel")
async def save_analysis_to_excel(
    title: str,
    rows: list[dict],
    notes: list[str] | None = None,
    sources: list[str] | None = None,
    detail_codes: list[str] | None = None,
    detail_days: int = 60,
    extras: dict | None = None,
    filename: str = "",
) -> str:
    """분석결과저장 — **당신이 정리한 표**를 그대로 Excel로 저장합니다.

    `scan_to_excel` 은 정해진 지표를 종목별로 담는 도구입니다. 이 도구는 그와 달리,
    여러 도구를 돌려 **직접 판단하고 정리한 결과**(후보 목록, 단계 판정, 겹침 분석
    같은 것)를 사용자가 파일로 가져갈 수 있게 만듭니다.

    ## 반드시 지킬 것

    1. **rows 의 숫자는 도구가 돌려준 값만 씁니다.** 기억이나 추정으로 채우지 마세요.
       확인 못 한 칸은 `"-"` 또는 `"확인 안 됨"` 으로 두고, 0 으로 채우지 않습니다.
    2. **`sources` 에 어떤 도구를 썼는지 적습니다.** 파일만 남았을 때 이 숫자가
       어디서 왔는지 알 수 있어야 합니다. 예: `["screen_by_flow", "get_indicators_bulk"]`
    3. **판단 근거와 한계를 `notes` 에 적습니다.** 무엇을 걸렀고 무엇을 못 봤는지,
       장중 잠정치인지 같은 것. 표만 남으면 나중에 잘못 읽힙니다.
    4. 매수·매도 권유, 목표가·손절가는 넣지 않습니다.

    ## rows 구성 요령

    - 각 dict 의 키가 곧 열 이름입니다. **한글로 쓰세요**(예: `"종목명"`, `"수급(2일)"`).
    - 모든 행이 같은 키를 갖도록 맞춥니다. 빠진 키는 빈칸이 됩니다.
    - 숫자는 숫자 그대로 넣으세요(문자열 `"12.5%"` 대신 `12.5`). 자릿점·색은 저장할 때
      자동으로 붙습니다. 단위는 열 이름에 넣으세요(`"기간수익률(%)"`).
    - 종목이 20개 이하면 비교 그래프가 자동으로 함께 저장됩니다.

    Args:
        title: 파일 제목. 첫 시트 이름과 요약에 쓰입니다. 예: "기대 부상후보 스캔"
        rows: 표 데이터. `[{"종목명": "성광벤드", "기간수익률(%)": -8.9}, ...]`
        notes: 판단 근거·제외 기준·한계. 한 줄에 하나씩.
        sources: 근거가 된 도구 이름들.
        detail_codes: **종목별 흐름 시트**를 만들 6자리 코드들(최대 8개).
            각 시트에 종가 추이·기관/외국인 순매매 그래프와 원자료가 들어갑니다.
            목록만 있으면 "그래서 어떻게 움직이고 있나"를 다시 조회해야 하므로,
            사용자가 더 볼지 정할 수 있게 **후보 상위 몇 개는 넣어 주세요**.
        detail_days: 흐름 시트에 담을 거래일 수 (기본 60, 최대 120).
        extras: **다른 렌즈에서 가져온 것을 종목 시트에 얹습니다.**
            StockLens 는 TelegramLens·DartLens 를 직접 부르지 못하므로,
            그 도구들을 먼저 호출해 결과를 여기에 담아 주세요.
            `{"042660": [{"title": "텔레그램 언급", "cols": ["시각","채널","내용"],
                          "rows": [["08-26 10:30","매경 자이앤트","원전 수주 기대"]]}]}`
            형식이며, 카드는 담은 순서대로 붙습니다.
            넣을 만한 것:
              · TelegramLens — `telegram_stock_buzz`(언급 원문·telegram_link 를 url 로),
                `telegram_timeline`(확산 흐름)
              · DartLens — `get_major_holders`(5%룰: 연기금·행동주의 진입),
                `get_insider_trades`(임원·주요주주 매매 = 안에서 먼저 움직인 흔적),
                `list_disclosures`(정기·수시 공시), `get_order_backlog`(수주잔고 추이)
              · StockLens — `get_us_short`(미국 공매도), `get_event_reactions`(공시 반응)
            행은 `["값", ...]` 또는 `{"cells": [...], "url": "원문 링크"}` 로 넣습니다.
            ⚠️ 원문에 있는 내용만 옮기고, 없는 것을 지어내지 마세요.
        filename: 파일명 (비우면 제목으로 자동 생성)
    """
    if not rows or not isinstance(rows, list):
        return "rows 가 비어 있습니다. 표로 만들 행을 넘겨주세요."
    clean = [r for r in rows if isinstance(r, dict) and r]
    if not clean:
        return "rows 에 사전(dict) 형태의 행이 없습니다."

    df = pd.DataFrame(clean)
    base = filename or generate_filename(re.sub(r"[^\w가-힣]+", "_", title)[:40] or "분석결과")
    if not base.endswith(".xlsx"):
        base += ".xlsx"
    file_path = get_snapshot_dir() / base

    meta = {"제목": title, "행 수": len(df)}
    if sources:
        meta["근거 도구"] = ", ".join(str(s) for s in sources)
    meta["작성 주체"] = "AI가 도구 결과를 정리한 표 (원본 조회 결과 아님)"

    # 탭 이름에 제목을 그대로 넣으면 '완성판' 같은 말이 탭에 박혀 무슨 시트인지
    # 알 수 없다. 탭은 '요약'으로 고정하고 제목은 시트 안에 적는다.
    saved = save_dataframe_to_excel(df, file_path, sheet_name="요약", metadata=meta)

    # 근거·한계는 별도 시트에 남긴다 — 표만 떨어져 나가면 맥락이 사라진다.
    # notes 를 안 주거나 한 줄만 줘도 '무엇을 보고 만든 파일인지'는 남아야 한다.
    auto_notes: list[str] = []
    try:
        clock = build_market_clock()
        krx = clock.get("krx") or {}
        auto_notes.append(
            "조회 시각 " + _dt.datetime.now().strftime("%Y-%m-%d %H:%M")
            + " · 한국장 " + ("장중(마감 전)" if krx.get("is_open") else "마감 후")
            + (" — 장중 수치는 마감 시 달라질 수 있습니다." if krx.get("is_open") else "")
        )
    except Exception:
        pass
    auto_notes.append("개인 순매매는 제공되지 않습니다(기관·외국인만). "
                      "'개인이 받았다'는 판단은 이 자료로 할 수 없습니다.")
    auto_notes.append("PER 은 현재가 ÷ 최근 4분기 EPS(TTM) 기준입니다. "
                      "실적표의 PER(각 기간말 주가 기준)과 값이 다를 수 있습니다.")

    if True:
        from openpyxl import load_workbook
        wb = load_workbook(saved)
        ws = wb.create_sheet("근거와 한계")
        ws.column_dimensions["A"].width = 110
        r = 1
        ws.cell(row=r, column=1, value=f"[{title}]"); r += 2
        if sources:
            ws.cell(row=r, column=1, value="■ 이 표를 만든 도구"); r += 1
            for s in sources:
                ws.cell(row=r, column=1, value=f"  · {s}"); r += 1
            r += 1
        if notes:
            ws.cell(row=r, column=1, value="■ 판단 근거와 한계"); r += 1
            for n in notes:
                ws.cell(row=r, column=1, value=f"  · {n}"); r += 1
            r += 1
        if auto_notes:
            ws.cell(row=r, column=1, value="■ 이 자료를 읽을 때"); r += 1
            for n in auto_notes:
                ws.cell(row=r, column=1, value=f"  · {n}"); r += 1
        r += 1
        ws.cell(row=r, column=1,
                value="※ 이 표는 조회 결과를 정리한 것이며 매수·매도 권유가 아닙니다.")
        wb.save(saved)

    # 종목별 흐름 시트 — 목록만으로는 "그래서 어떻게 움직이나"를 못 본다.
    made_detail: list[str] = []
    if detail_codes:
        from openpyxl import load_workbook
        from stock_mcp_server._excel import add_detail_sheet
        detail_days = max(20, min(int(detail_days), 120))
        picks = [str(c).strip() for c in detail_codes if str(c).strip()][:8]
        name_of = {}
        for row in clean:
            c = str(row.get("코드") or row.get("code") or "").strip()
            if c:
                name_of[c] = str(row.get("종목명") or row.get("name") or c)

        async def one_detail(c: str):
            """한 종목에 필요한 조각을 모은다. 일부가 실패해도 나머지는 살린다."""
            out: dict = {}
            bars: list = []
            fl: list = []
            try:
                bars = await get_ohlcv(c, "day", max(detail_days, 130))
                ind = compute_indicators(bars, ["ma_phase", "position", "volume"])
                pos = ind.get("position") or {}
                vol = ind.get("volume") or {}
                out["price"] = {
                    "현재가": pos.get("price"),
                    "등락률": None,
                    "고점대비": pos.get("pct_from_high_52w"),
                    "저점대비": pos.get("pct_from_low_52w"),
                    "이평배열": (ind.get("ma_phase") or {}).get("phase_label"),
                    "거래량배수": vol.get("ratio_vs_avg_20b"),
                    "저점가": pos.get("low_52w"),
                    "고점가": pos.get("high_52w"),
                }
            except Exception:
                pass
            try:
                fin = await get_financials(c)
                ttm, label = _ttm_eps_for(fin)
                px = (out.get("price") or {}).get("현재가")
                per = round(px / ttm, 2) if (ttm and px) else None
                pbr, _ = _latest_fin(fin, "PBR(배)")
                roe, basis = _latest_fin(fin, "ROE(지배주주)")
                out["valuation"] = {"PER": per, "PBR": pbr, "ROE": roe,
                                    "기준": f"PER=TTM~{label}" if label else (basis or "-")}
                meta = fin.get("_periods") or {}
                ann, qtr = meta.get("annual") or [], meta.get("quarterly") or []
                rev, op = fin.get("매출액") or [], fin.get("영업이익") or []
                qs = []
                for i, lb in enumerate(qtr):
                    idx = len(ann) + i
                    if idx >= len(rev) or idx >= len(op):
                        continue
                    # 추정치는 버리지 않고 라벨에 남긴다 — 앞으로의 방향이 보인다.
                    est = "(E)" in str(lb)
                    try:
                        qs.append({"기간": str(lb),
                                   "추정": est,
                                   "매출액": float(str(rev[idx]).replace(",", "")),
                                   "영업이익": float(str(op[idx]).replace(",", ""))})
                    except ValueError:
                        continue
                out["quarters"] = qs[-6:]

                def _last(key):
                    v, _ = _latest_fin(fin, key)
                    return v
                out["stability"] = {"부채비율": _last("부채비율"),
                                    "당좌비율": _last("당좌비율"),
                                    "주당배당금": _last("주당배당금(원)"),
                                    "시가배당률": _last("시가배당률(%)")}
                # 연간 구간의 (E) 는 증권사 추정치다. 확정치와 섞지 않고 따로 담는다.
                est_i = None
                for i, lb in enumerate(ann):
                    if "(E)" in str(lb):
                        est_i = i
                        break
                if est_i is not None:
                    def _est(series):
                        try:
                            return float(str(series[est_i]).replace(",", ""))
                        except (ValueError, IndexError, TypeError):
                            return None
                    out["forecast"] = {"기간": str(ann[est_i]),
                                       "매출액": _est(rev), "영업이익": _est(op)}
            except Exception:
                pass
            try:
                sec = await naver_get_stock_sector(c)
                if sec.get("sector_name"):
                    out.setdefault("valuation", {})["_업종"] = sec["sector_name"]
                if sec.get("sector_per_naver"):
                    out.setdefault("valuation", {})["업종PER"] = sec["sector_per_naver"]
                    p_ = (out.get("valuation") or {}).get("PER")
                    if p_ and sec["sector_per_naver"]:
                        gap = (p_ - sec["sector_per_naver"]) / sec["sector_per_naver"] * 100
                        out["valuation"]["위치"] = (
                            f"{sec.get('sector_name','업종')} 대비 {gap:+.0f}% "
                            f"{'할증' if gap > 5 else ('할인' if gap < -5 else '비슷')}")
            except Exception:
                pass
            try:
                sec_name = (out.get("valuation") or {}).get("_업종")
                if sec_name:
                    # 목록이 등락률 순이라 12개만 받으면 기준 종목이 밀려난다.
                    # 업종 전체를 받아 본인 + 상위 몇 개를 고른다.
                    peer_raw = await naver_get_sector_stocks(sec_name, count=200)
                    stocks = peer_raw.get("stocks") if isinstance(peer_raw, dict) else None
                    rows_ = list(stocks or [])
                    mine = [x for x in rows_ if x.get("code") == c]
                    others = [x for x in rows_ if x.get("code") != c]
                    # 같은 업종을 보여주는데 정작 본인이 빠지면 비교가 안 된다.
                    rows_ = (mine + others)[:6] if mine else others[:5]
                    if not mine:
                        rows_ = others[:5]
                    peers = []
                    for pr in rows_:
                        rate = str(pr.get("change_rate", "")).replace("%", "").replace("+", "")
                        try:
                            rate_v = float(rate)
                        except ValueError:
                            rate_v = None
                        peers.append({"종목": pr.get("name"), "등락률": rate_v,
                                      "현재가": pr.get("price"),
                                      "본인": pr.get("code") == c})
                    out["peers"] = peers
            except Exception:
                pass
            try:
                fl = await get_investor_flow(c, 20)
                inst = [d.get("institutional") for d in fl if d.get("institutional") is not None]
                forg = [d.get("foreign") for d in fl if d.get("foreign") is not None]
                out["flow"] = {"기관": int(sum(inst)) if inst else None,
                               "외국인": int(sum(forg)) if forg else None}
                # 누적만으로는 '언제' 들어왔는지 안 보인다 — 일별도 담는다.
                daily = []
                for d in list(fl)[:15][::-1]:
                    daily.append({"날짜": str(d.get("date", "")).replace(".", "-"),
                                  "기관": d.get("institutional"),
                                  "외국인": d.get("foreign")})
                out["flow_daily"] = daily
            except Exception:
                pass
            try:
                con = await naver_get_consensus(c)
                tp = con.get("target_price")
                px = (out.get("price") or {}).get("현재가")
                op = con.get("opinion") or {}
                total = sum(v for v in op.values() if isinstance(v, (int, float)))
                buy = sum(v for k, v in op.items()
                          if isinstance(v, (int, float)) and ("매수" in str(k)))
                out["consensus"] = {
                    "목표주가": tp,
                    "괴리율": round((tp - px) / px * 100, 1) if (tp and px) else None,
                    "의견": (f"매수 {buy} / 전체 {int(total)}건" if total else None),
                }
                es = con.get("earnings_surprise") or {}
                op_row = es.get("영업이익") or {}
                surp = []
                for period, v in list(op_row.items())[-3:]:
                    if isinstance(v, dict) and v.get("surprise") is not None:
                        # 202512 → 2025.12 로 읽히게 (원본은 붙여 쓴 문자열)
                        pl = str(period)
                        if len(pl) == 6 and pl.isdigit():
                            pl = f"{pl[:4]}.{pl[4:]}"
                        surp.append(f"{pl} {v['surprise']:+.1f}%")
                if surp:
                    out["consensus"]["서프라이즈"] = " · ".join(surp)
                pts = []
                for h in (con.get("target_price_history") or []):
                    y, d = h.get("price"), h.get("date")
                    if y is None or d is None:
                        continue
                    try:
                        lb = _dt.datetime.fromtimestamp(float(d) / 1000,
                                                        _dt.timezone.utc).strftime("%Y-%m")
                    except Exception:
                        continue
                    pts.append({"월": lb, "목표주가": float(y)})
                out["target"] = pts
            except Exception:
                pass
            news = []
            try:
                for it in (await naver_get_disclosure_list(c))[:4]:
                    link = it.get("link") or ""
                    if link.startswith("/"):
                        link = "https://finance.naver.com" + link
                    news.append({"date": it.get("date"), "title": it.get("title"),
                                 "url": link or None})
            except Exception:
                pass
            try:
                reps = await naver_get_reports(c)
                for it in (reps or [])[:3]:
                    nid = it.get("nid")
                    url = (f"https://finance.naver.com/research/company_read.naver?nid={nid}"
                           if nid else None)
                    news.append({
                        "date": it.get("date"),
                        "title": "[리포트] " + str(it.get("broker", "")) + " " + str(it.get("title", "")),
                        "url": url,
                        "brief": (it.get("summary") or "")[:90],
                    })
            except Exception:
                pass
            out["news"] = news

            # 공시 목록만 있으면 "그래서 주가가 어떻게 반응했나"를 알 수 없다.
            # 이미 받아둔 일봉·수급으로 계산하므로 추가 조회는 없다.
            if bars:
                rx_rows = []
                for item in news:
                    title = str(item.get("title", ""))
                    if title.startswith("[리포트]"):
                        continue
                    day = str(item.get("date", "")).replace(".", "-")
                    if len(day) != 10:
                        continue
                    try:
                        rx = build_event_reaction(code=c, event_date=day, ohlcv=bars,
                                                  flows=fl, before=3, after=5)
                    except Exception:
                        continue
                    pts = rx.get("points") or {}

                    def _ret(key):
                        p = pts.get(key) or {}
                        return p.get("return_pct") if p.get("status") == "available" else None

                    d1, d5 = _ret("D+1"), _ret("D+5")
                    if d1 is None and d5 is None:
                        continue
                    # lstrip 은 문자 집합을 지운다 — '(주) ' 로 지우면 뒤따르는
                    # '주'까지 먹어 '주식선물'이 '식선물'이 된다. 접두사만 떼어낸다.
                    short = title
                    nm = str(name_of.get(c, ""))
                    if nm and short.startswith(nm):
                        short = short[len(nm):]
                        for pre in ("(주)", "㈜", ")"):
                            if short.startswith(pre):
                                short = short[len(pre):]
                                break
                        short = short.lstrip()
                    rx_rows.append({"날짜": day, "공시": short[:34],
                                    "다음날": d1, "5일후": d5})
                out["reactions"] = rx_rows[:5]
            if isinstance(extras, dict):
                ex = extras.get(c) or extras.get(str(c).zfill(6))
                if isinstance(ex, list):
                    out["extras"] = [x for x in ex if isinstance(x, dict)]
            return c, out

        # 업종·시장 등락률은 종목마다 다시 부를 필요가 없다 — 한 번만 받아 나눠 쓴다.
        sector_rate: dict = {}
        market_rate = None
        try:
            for s_ in (await naver_list_sectors() or []):
                nm = s_.get("name")
                rate = s_.get("change_rate")
                if nm is None or rate is None:
                    continue
                try:
                    sector_rate[str(nm)] = float(str(rate).replace("%", "").replace("+", ""))
                except ValueError:
                    continue
        except Exception:
            pass
        try:
            for idx_ in (await get_market_index() or []):
                if idx_.get("index") == "KOSPI" and idx_.get("change_rate") is not None:
                    market_rate = float(idx_["change_rate"])
        except Exception:
            pass

        results = await asyncio.gather(*[one_detail(c) for c in picks])
        wb = load_workbook(saved)
        for c, payload in results:
            if not payload:
                continue
            label = name_of.get(c, c)
            summary = []
            for row in clean:
                rc = str(row.get("코드") or row.get("code") or "").strip()
                if rc == c:
                    summary = list(row.items())
                    if payload.get("price") is not None:
                        chg_ = row.get("등락률")
                        payload["price"]["등락률"] = chg_
                        p_ = payload["price"]
                        hi, lo, cur = p_.get("고점가"), p_.get("저점가"), p_.get("현재가")
                        if all(isinstance(x, (int, float)) for x in (hi, lo, cur)) and hi > lo:
                            payload["price"]["52주위치"] = (cur - lo) / (hi - lo) * 100
                        sec_nm = (payload.get("valuation") or {}).get("_업종")
                        rel = {}
                        if isinstance(chg_, (int, float)):
                            if sec_nm and sec_nm in sector_rate:
                                rel["업종"] = chg_ - sector_rate[sec_nm]
                            if market_rate is not None:
                                rel["시장"] = chg_ - market_rate
                        if rel:
                            payload["price"]["상대성과"] = rel
                    break
            try:
                add_detail_sheet(wb, f"{label}", summary, payload)
                made_detail.append(label)
            except Exception:
                continue
        # 요약 표의 종목명에서 그 종목 시트로 바로 넘어가게 한다.
        # 시트가 여럿이면 탭을 찾아 헤매게 되고, 훑는 흐름이 끊긴다.
        if made_detail:
            try:
                from openpyxl.styles import Font as _F
                first = wb[wb.sheetnames[0]]
                head = [c.value for c in first[1]]
                if "종목명" in head:
                    ncol = head.index("종목명") + 1
                    for row in first.iter_rows(min_row=2, min_col=ncol, max_col=ncol):
                        for cell in row:
                            if cell.value in made_detail:
                                cell.hyperlink = f"#'{cell.value}'!A1"
                                cell.font = _F(color="0563C1", underline="single")
            except Exception:
                pass

        if made_detail:
            # 탭을 넘기며 훑는 순서로 맞춘다 — 요약·차트 다음에 종목들,
            # 관리용 시트(근거·메타)는 뒤로. 안 그러면 종목 시트가 맨 끝에 밀린다.
            # 종목 시트는 이 시점에야 확정된다 — 근거 시트에 목록을 덧붙인다.
            if "근거와 한계" in wb.sheetnames:
                gs = wb["근거와 한계"]
                rr = gs.max_row + 2
                gs.cell(row=rr, column=1, value="■ 종목별 상세 시트")
                gs.cell(row=rr + 1, column=1,
                        value="  · " + ", ".join(made_detail)
                              + "   (요약 시트에서 종목명을 누르면 이동합니다)")
            tail = [n for n in ("근거와 한계", "Metadata") if n in wb.sheetnames]
            front = [n for n in wb.sheetnames if n not in tail and n not in made_detail]
            try:
                wb._sheets = [wb[n] for n in front + made_detail + tail]
            except Exception:
                pass
            wb.save(saved)

    out = [f"✓ 저장 완료 — {title} ({len(df)}행)",
           f"경로: {saved}",
           f"열: {', '.join(str(c) for c in df.columns)}"]
    if made_detail:
        out.append(f"종목별 흐름 시트: {', '.join(made_detail)} (종가·수급 그래프 포함)")
    if sources:
        out.append(f"근거 도구: {', '.join(str(s) for s in sources)}")
    if not notes:
        out.append("⚠️ notes 가 비어 있습니다 — 무엇을 걸렀고 무엇을 못 봤는지 적어주세요. "
                   "표만 남으면 나중에 잘못 읽힙니다.")
    return "\n".join(out)



@mcp.tool()
@safe_tool
@track_metrics("query_excel")
async def query_excel(
    file_path: str,
    filters: dict | None = None,
    sort_by: str = "",
    descending: bool = True,
    limit: int = 30,
) -> str:
    """엑셀쿼리 — 저장된 Excel 스냅샷(scan_to_excel 산출)에서 조건에 맞는 종목을 로컬 필터링. HTTP 없이 빠름.

    필터 형식(둘 다 지원):
        간단: {"per_max": 10, "pbr_max": 1.5, "drawdown_pct_max": -30}
        상세: {"per": {"max": 10, "min": 0}, "drawdown_pct": {"max": -30}}

    Args:
        file_path: scan_to_excel로 만든 파일 경로
        filters: 필터 조건 (컬럼명_max / 컬럼명_min 형식)
        sort_by: 정렬 기준 컬럼 (예: "market_cap", "drawdown_pct")
        descending: 내림차순 (기본 True)
        limit: 반환 최대 개수 (기본 30)
    """
    try:
        df = load_excel(file_path, sheet_name="Snapshot")
    except FileNotFoundError:
        return f"파일을 찾을 수 없습니다: {file_path}"
    except Exception as e:
        return f"파일 로드 실패: {type(e).__name__}: {e}"

    unapplied = unknown_filter_columns(df, filters) if filters else []
    available_cols = list(df.columns)
    if filters:
        df = apply_filters(df, filters)

    if sort_by and sort_by in df.columns:
        df = df.sort_values(sort_by, ascending=not descending)

    df = df.head(limit)

    if df.empty:
        msg = f"조건에 맞는 종목이 없습니다. (원본: {file_path})"
        if unapplied:
            msg += (
                "\n\n⚠️ 이 파일에 없는 컬럼이라 **적용되지 않은 조건**: "
                + ", ".join(unapplied)
                + "\n사용 가능한 컬럼: "
                + ", ".join(available_cols)
            )
        return msg

    # 주요 컬럼만 표시
    display_cols = [
        c for c in ["code", "name", "current_price", "drawdown_pct", "per", "pbr", "volume"]
        if c in df.columns
    ]
    if display_cols:
        df_display = df[display_cols]
    else:
        df_display = df
    # 화면 표기도 파일과 같은 한글 이름으로 맞춘다.
    df_display = df_display.rename(
        columns={c: COLUMN_LABELS_KO.get(c, c) for c in df_display.columns}
    )

    lines = [f"쿼리 결과 ({len(df)}개 종목, 파일: {Path(file_path).name}):", ""]
    if unapplied:
        # 없는 컬럼 조건은 조용히 무시된다 — 알려주지 않으면 걸리지도 않은 조건을
        # 걸린 것으로 읽는다(PER 없는 파일에 per_max=10 → 전 종목이 그대로 나옴).
        lines.append(
            f"⚠️ **적용되지 않은 조건**: {', '.join(unapplied)} — 이 파일에 없는 컬럼입니다. "
            f"아래 결과는 이 조건이 **빠진 채** 걸러진 것입니다."
        )
        lines.append(f"사용 가능한 컬럼: {', '.join(available_cols)}")
        lines.append("")
    lines.append("| " + " | ".join(df_display.columns) + " |")
    lines.append("|" + "|".join(["---"] * len(df_display.columns)) + "|")
    for _, row in df_display.iterrows():
        values = []
        for col in df_display.columns:
            v = row[col]
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                # 열 이름은 이미 한글로 바뀐 뒤라 영문 키로만 판정하면 서식이 빠진다.
                if col in ("current_price", "volume", "avg_volume", "high", "low") or col in (
                    "현재가", "거래량", "평균거래량", "기간최고가", "기간최저가",
                    "시가", "종가", "EPS(원)", "BPS(원)", "집계봉수",
                ):
                    values.append(f"{int(v):,}")
                elif col.endswith("_pct") or col.endswith("(%)"):
                    values.append(f"{v:+.2f}%")
                else:
                    values.append(f"{v:.2f}" if isinstance(v, float) else str(v))
            else:
                values.append(str(v))
        lines.append("| " + " | ".join(values) + " |")

    return "\n".join(lines)


@mcp.tool()
@safe_tool
@track_metrics("get_metrics_summary")
async def get_metrics_summary(days: int = 1) -> str:
    """사용량통계 — 최근 N일간 MCP 도구 사용량을 집계해서 보여줍니다.

    디버깅/최적화용. 도구별로:
    - 호출 횟수
    - 평균/p50/p95 실행 시간
    - 평균 토큰 소모량
    - 캐시 히트율
    - 에러 발생 횟수
    를 보여줍니다.

    로그 파일 위치: ~/.stocklens/logs/metrics_YYYYMMDD.jsonl
    (2026-08 이전 기록은 ~/Downloads/kstock/logs/ 에 있고, 그것도 같이 읽습니다)

    Args:
        days: 조회할 일수 (기본 1, 오늘만. 최대 30)
    """
    days = max(1, min(days, 30))

    records = load_metrics(days=days)
    if not records:
        return (
            f"최근 {days}일간 기록이 없습니다.\n"
            f"로그 파일: {get_metrics_file()}"
        )

    summary = summarize_metrics(records)

    sorted_tools = sorted(
        summary.items(),
        key=lambda x: x[1]["call_count"],
        reverse=True,
    )

    lines = [
        f"📊 MCP 사용량 통계 (최근 {days}일, 총 {len(records)}회 호출)",
        "",
        "도구 | 호출 | 평균시간 | p95시간 | 평균토큰 | 캐시율 | 에러",
        "---|---|---|---|---|---|---",
    ]

    total_tokens = 0
    total_calls = 0

    for tool, s in sorted_tools:
        total_tokens += s["total_output_tokens"]
        total_calls += s["call_count"]
        lines.append(
            f"{tool} | "
            f"{s['call_count']} | "
            f"{s['avg_duration_ms']}ms | "
            f"{s['p95_duration_ms']}ms | "
            f"{s['avg_tokens']:,} | "
            f"{s['cache_hit_rate']}% | "
            f"{s['errors']}"
        )

    lines.append("")
    lines.append(f"**총 소모 토큰: {total_tokens:,}**")
    lines.append(f"**총 호출: {total_calls}회**")
    lines.append(f"**평균 호출당 토큰: {total_tokens // max(total_calls, 1):,}**")
    lines.append("")
    lines.append(f"로그 파일: {get_metrics_file()}")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# ETF 도구
# ---------------------------------------------------------------------------


@mcp.tool()
@safe_tool
@track_metrics("get_etf_list")
async def get_etf_list(
    category: str = "",
    keyword: str = "",
    sort_by: str = "marketSum",
    limit: int = 20,
) -> str:
    """ETF목록 — ETF 전체 목록 조회, 카테고리·이름 키워드 필터링.

    category: 카테고리 필터 (빈 문자열=전체). 가능한 값:
      "국내 시장지수", "국내 업종/테마", "국내 파생",
      "해외 주식", "원자재", "채권/금리", "단기자금"
    keyword: ETF 이름에 포함된 키워드로 필터 (예: "우주", "반도체", "배당").
      "OO 테마 ETF 뭐 있어" 같은 질문은 category를 추측하지 말고 keyword로
      바로 찾는다 — 한국 ETF는 이름 자체가 테마를 직설적으로 담는 관행이라
      (예: "TIGER 미국우주테크") 이름 검색만으로 대부분 커버된다. 단, 특정
      회사(예: 비상장사)를 "보유"한 ETF를 찾는 건 이 키워드로 안 된다 — 그
      회사명이 ETF 이름 자체에 없으면 못 찾으니, 후보 ETF를 좁힌 뒤
      get_etf_info로 구성종목을 직접 확인해야 한다.
    sort_by: 정렬 기준 — "marketSum"(시가총액), "quant"(거래량),
      "threeMonthEarnRate"(3개월수익률)
    limit: 반환 개수 (기본 20, 최대 50)
    """
    data = await naver_get_etf_list(
        category=category or None,
        keyword=keyword or None,
        sort_by=sort_by,
        limit=limit,
    )

    items = data["items"]
    if not items:
        return "조건에 맞는 ETF가 없습니다."

    header_extra = "".join([
        f", 카테고리: {category}" if category else "",
        f", 키워드: '{keyword}'" if keyword else "",
    ])
    lines = [
        f"ETF 목록 ({data['total']}개 중 상위 {len(items)}개{header_extra})",
        "",
    ]

    for it in items:
        chg = it.get("change_rate", 0) or 0
        chg_sign = "+" if chg > 0 else ""
        ret3m = it.get("return_3m")
        ret3m_str = f" | 3M: {'+' if ret3m > 0 else ''}{ret3m:.1f}%" if ret3m else ""
        mcap = it.get("market_cap")
        price = it.get("price")
        nav = it.get("nav")

        lines.append(
            f"- **{it['name']}** ({it['code']}) "
            f"| {f'{price:,.0f}원' if price is not None else '가격 없음'} "
            f"({chg_sign}{chg:.2f}%) "
            f"| NAV {f'{nav:,.0f}원' if nav is not None else '없음'} "
            f"| 시총 {f'{mcap:,.0f}억원' if mcap else '없음'}"
            f"{ret3m_str}"
        )

    lines.append("")
    lines.append("카테고리: " + ", ".join(data["categories"].values()))

    return _append_result_meta("\n".join(lines), _kr_meta(kind="snapshot"))


@mcp.tool()
@safe_tool
@track_metrics("get_etf_info")
async def get_etf_info(code: str) -> str:
    """ETF정보 — ETF 상세 정보 (기초지수, 보수율, 수익률, 구성종목 TOP10).

    code: ETF 종목코드 (예: "069500" KODEX 200, "360750" TIGER 미국S&P500)
    """
    import re
    if not re.match(r"^[A-Za-z0-9]{6}$", code):
        return f"⚠️ 종목코드 형식이 올바르지 않습니다: {code}"

    data = await naver_get_etf_detail(code)

    if not data.get("name"):
        return f"ETF코드 {code}의 정보를 가져올 수 없습니다. 코드를 확인해주세요."

    # 예전엔 결측을 0으로 채워 "베타 0.00", "펀드보수 연 0.000%", "외국인 비율
    # 0.00%"처럼 **의미 있는 값**으로 보이게 출력했다. 0과 '없음'은 다르다.
    def _num(key: str, fmt: str, suffix: str = "", *, signed: bool = False) -> str:
        v = data.get(key)
        if v is None:
            return "데이터 없음"
        s = format(v, fmt)
        if signed and v > 0:
            s = "+" + s
        return s + suffix

    lines = [
        f"# {data['name']} ({code})",
        "",
        "## 기본 정보",
        f"- 기초지수: {data.get('base_index') or '-'}",
        f"- 유형: {data.get('etf_type') or '-'}",
        f"- 자산운용사: {data.get('issuer') or '-'}",
        f"- 상장일: {data.get('listing_date') or '-'}",
        f"- 펀드보수: 연 {_num('total_fee', '.3f', '%')}",
        f"- 펀드유형: {data.get('fund_type') or '-'}",
    ]

    if data.get("dividend_base"):
        lines.append(f"- 분배금 기준일: {data['dividend_base']}")

    lines.append("")
    lines.append("## 시세 정보")
    lines.append(f"- 현재가: {_num('price', ',.0f', '원')}")

    chg = data.get("price_change")
    chg_rate = data.get("price_change_rate")
    if chg is None or chg_rate is None:
        lines.append("- 전일대비: 데이터 없음")
    else:
        chg_sign = "+" if chg > 0 else ""
        lines.append(f"- 전일대비: {chg_sign}{chg:,.0f}원 ({chg_sign}{chg_rate:.2f}%)")
    lines.append(f"- 시가총액: {_num('market_cap', ',.0f', '억원')}")
    lines.append(
        f"- 52주 최고/최저: {_num('year_high', ',.0f', '원')} / "
        f"{_num('year_low', ',.0f', '원')}"
    )
    lines.append(f"- 베타: {_num('beta', '.2f')}")
    lines.append(f"- 외국인 비율: {_num('foreign_rate', '.2f', '%')}")
    # ETF 는 "원할 때 팔 수 있는가"가 개별주보다 중요하다. 20일 평균 거래량과
    # 유동성공급자(LP)는 받아오면서도 내보내지 않고 있었다.
    lines.append(f"- 20일 평균 거래량: {_num('volume_20d_avg', ',.0f', '주')}")
    lp = (data.get("lp_list") or "").strip()
    if lp:
        lines.append(f"- 유동성공급자(LP): {lp}")

    lines.append("")
    lines.append("## 수익률")
    for period, key in [("1개월", "return_1m"), ("3개월", "return_3m"),
                        ("6개월", "return_6m"), ("1년", "return_1y")]:
        lines.append(f"- {period}: {_num(key, '.2f', '%', signed=True)}")

    holdings = data.get("holdings", [])
    if holdings:
        has_weight = data.get("holdings_has_weight", True)
        lines.append("")
        lines.append(f"## 구성종목 (총 {data.get('holdings_count', len(holdings))}개)")
        lines.append("")

        if has_weight:
            lines.append("종목명 | 비중(%)")
            lines.append("---|---")
            for h in holdings[:10]:
                lines.append(f"{h['name']} | {h['weight']:.2f}%")
            top10_sum = sum(h["weight"] for h in holdings[:10])
            lines.append(f"**TOP10 합계** | **{top10_sum:.2f}%**")
        else:
            lines.append("종목명 | 주식수")
            lines.append("---|---")
            for h in holdings[:10]:
                lines.append(f"{h['name']} | {h['shares']:,.0f}")

    return _append_result_meta("\n".join(lines), _kr_meta(kind="snapshot", code=code))


# ========================================
# 🇺🇸 US Stock Tools (NYSE / NASDAQ via yfinance)
# ========================================

def safe_us_tool(func):
    """US tool 전용 에러 래퍼 — yfinance 컨텍스트 메시지. 라이선스 게이트 포함."""

    @functools.wraps(func)
    async def wrapper(*args, **kwargs):
        if not is_licensed():
            return locked_message()
        try:
            return await func(*args, **kwargs)
        except httpx.TimeoutException:
            return "⚠️ Yahoo Finance 응답이 지연되고 있습니다. 잠시 후 다시 시도해주세요."
        except httpx.ConnectError:
            return "⚠️ Yahoo Finance에 연결할 수 없습니다. 인터넷 연결을 확인해주세요."
        except Exception as e:
            return (
                f"⚠️ 미국 주식 데이터 처리 중 오류: {type(e).__name__}\n"
                f"티커가 올바른지 확인해주세요 (예: AAPL, MSFT, BRK.B)."
                + _support_hint()
            )

    return wrapper


def _ccy_prefix(code: str | None) -> str:
    """통화 코드 → 금액 접두사.

    USD 만 `$` 를 쓰고 나머지는 코드를 그대로 붙인다. ADR 은 주가 통화(USD)와
    재무제표 통화가 갈리는데(TSM: 거래 USD / 재무 TWD), TWD 금액에 `$` 를 붙이면
    달러로 읽힌다 — TSM 분기 매출 추정이 `$1,454.96B`(약 31배)로 보였다.
    """
    if not code or str(code).upper() == "USD":
        return "$"
    return f"{str(code).upper()} "


def _fmt_num(v, unit: str = "", digits: int = 2, currency: str = "$") -> str:
    if v is None:
        return "-"
    if isinstance(v, (int, float)):
        if abs(v) >= 1_000_000_000:
            return f"{currency}{v/1_000_000_000:,.{digits}f}B"
        if abs(v) >= 1_000_000:
            return f"{currency}{v/1_000_000:,.{digits}f}M"
        return f"{v:,.{digits}f}{unit}"
    return str(v)


def _fmt_ratio(v, digits: int = 2) -> str:
    """소수 비율(ROE=1.52, margin=0.27)을 백분율로. *100 후 %."""
    if v is None:
        return "-"
    if isinstance(v, (int, float)):
        return f"{v * 100:.{digits}f}%"
    return str(v)


def _fmt_yield(v, digits: int = 2) -> str:
    """yfinance의 dividendYield/fiveYearAvgDividendYield는 이미 %. 그대로 출력."""
    if v is None:
        return "-"
    if isinstance(v, (int, float)):
        return f"{v:.{digits}f}%"
    return str(v)


@mcp.tool()
@safe_us_tool
@track_metrics("get_us_price")
async def get_us_price(ticker: str) -> str:
    """US stock price — NYSE/NASDAQ 미국 주식 현재가 스냅샷 (US market via yfinance).
    "AAPL price", "Tesla 현재가", "MSFT quote", "Nvidia 얼마" 같은 질문에 사용합니다.

    현재가 + 전일대비 + 시/고/저 + 거래량 + 52주 고저 + 베타 + 시가총액 + 마켓 상태
    (정규장/프리/포스트) 반환. Yahoo Finance 데이터는 최대 15분 지연 가능.

    Args:
        ticker: US 티커 (예: "AAPL", "TSLA", "BRK.B", "SPY")
    """
    data = await us.get_price(ticker)
    if data is None:
        return f"티커 '{ticker}'를 찾을 수 없습니다. 정확한 미국 주식 심볼인지 확인해주세요."

    currency = data.get("currency") or "USD"
    sym = "$" if currency == "USD" else ""
    lines = [
        f"**{data.get('name', data['ticker'])}** ({data['ticker']}) — {data.get('exchange', '')}",
        f"현재가: {sym}{data['price']:,.2f}" if data.get("price") is not None else "현재가: -",
    ]
    ch = data.get("change")
    chp = data.get("change_percent")
    if ch is not None and chp is not None:
        lines.append(f"전일대비: {ch:+,.2f} ({chp:+.2f}%)")
    lines.append(f"시가/고가/저가: {_fmt_num(data.get('open'))} / {_fmt_num(data.get('day_high'))} / {_fmt_num(data.get('day_low'))}")
    lines.append(f"거래량: {data['volume']:,}" if data.get("volume") else "거래량: -")
    lines.append(f"52주 고저: {_fmt_num(data.get('52w_high'))} / {_fmt_num(data.get('52w_low'))}")
    if data.get("beta") is not None:
        lines.append(f"베타: {data['beta']:.2f}")
    if data.get("market_cap"):
        lines.append(f"시가총액: {_fmt_num(data['market_cap'])}")
    if data.get("market_state"):
        lines.append(f"마켓 상태: {data['market_state']}")
    return _append_result_meta("\n".join(lines), _us_meta(kind="snapshot", ticker=ticker))


@mcp.tool()
@safe_us_tool
@track_metrics("get_us_info")
async def get_us_info(ticker: str) -> str:
    """US stock info — 미국 주식 기업 정보 (섹터, 산업, 시총, 사업 요약 · US company profile).
    "Apple 어떤 회사", "NVDA 사업 설명", "TSLA sector", "기업 정보" 같은 질문에 사용합니다.

    Args:
        ticker: US 티커 (예: "AAPL", "NVDA")
    """
    data = await us.get_info(ticker)
    if data is None:
        return f"티커 '{ticker}'를 찾을 수 없습니다."

    lines = [
        f"**{data.get('name') or data['ticker']}** ({data['ticker']})",
        f"섹터/산업: {data.get('sector') or '-'} / {data.get('industry') or '-'}",
        f"거래소: {data.get('exchange') or '-'} ({data.get('quote_type') or '-'})",
        f"국가: {data.get('country') or '-'}",
    ]
    # 시가총액은 주가(거래통화), EV·현금·부채는 재무제표(재무통화)에서 나온다.
    # ADR은 둘이 갈리므로 각각 제 통화로 찍고, 다르면 나란히 비교하지 말라고 적는다.
    trade_ccy = data.get("currency")
    fin_ccy = data.get("financial_currency")
    ccy_mixed = bool(trade_ccy and fin_ccy and trade_ccy != fin_ccy)
    if data.get("market_cap"):
        lines.append(f"시가총액: {_fmt_num(data['market_cap'], currency=_ccy_prefix(trade_ccy))}")
    if data.get("enterprise_value"):
        lines.append(
            f"기업가치(EV): {_fmt_num(data['enterprise_value'], currency=_ccy_prefix(fin_ccy))}"
        )
    if ccy_mixed:
        lines.append(
            f"⚠️ **통화 불일치** — 시가총액은 주가 통화({trade_ccy}), 기업가치(EV)는 "
            f"재무제표 통화({fin_ccy}) 기준입니다. 두 값을 나눠 배수로 쓰면 안 됩니다."
        )
    if data.get("shares_outstanding"):
        lines.append(f"발행주식수: {data['shares_outstanding']:,.0f}")
    if data.get("employees"):
        lines.append(f"임직원: {data['employees']:,}명")
    if data.get("website"):
        lines.append(f"웹사이트: {data['website']}")
    summary = data.get("business_summary")
    if summary:
        lines.append("")
        lines.append("## 사업 요약")
        lines.append(summary[:800] + ("..." if len(summary) > 800 else ""))
    return _append_result_meta("\n".join(lines), _us_meta(kind="snapshot", ticker=ticker))


@mcp.tool()
@safe_us_tool
@track_metrics("get_us_chart")
async def get_us_chart(
    ticker: str,
    period: str = "3mo",
    interval: str = "1d",
    prepost: bool = False,
    limit: int = 500,
) -> str:
    """US stock chart OHLCV — 미국 주식 시계열 캔들 데이터 (US historical price data).
    "AAPL 차트", "Tesla 1년 주가", "NVDA history" 같은 질문에 사용. US 시계열 진입점(한국은 get_chart).

    Args:
        ticker: US 티커 (예: "AAPL")
        period: "1d","5d","1mo","3mo","6mo","1y","2y","5y","10y","ytd","max" (기본 3mo)
        interval: "1m","5m","15m","30m","1h","1d","1wk","1mo" (기본 1d)
        prepost: 프리/포스트 마켓 포함 (intraday에서만 유효)
        limit: 최대 행수 (기본 500, 최대 5000). 큰 데이터는 export_us_to_excel 권장(토큰 0).
    """
    limit = max(10, min(limit, 5000))  # 10~5000 범위로 클램프
    rows = await us.get_history(ticker, period=period, interval=interval, prepost=prepost)
    if not rows:
        return f"티커 '{ticker}'의 차트 데이터를 가져올 수 없습니다."

    total = len(rows)
    truncated = total > limit
    if truncated:
        rows = rows[-limit:]  # 최근 구간 유지

    is_intraday = interval in ("1m", "5m", "15m", "30m", "1h")

    header = f"**{ticker.upper()}** {period} {interval} OHLCV ({len(rows)} bars"
    if truncated:
        header += f" · 원본 {total}행 중 최근 {limit}행"
    header += ")"

    lines = [header, ""]
    lines.extend(_format_ohlcv_rows(rows, is_intraday=is_intraday, decimals=2))
    if truncated:
        lines.append("")
        lines.append(
            f"💡 더 필요하면 `limit={min(total, 5000)}`로 재호출하세요. "
            f"백테스트·CSV 저장 용도면 **export_us_to_excel** 로 파일 저장 (토큰 0)."
        )
    # 일·주·월봉은 마지막 봉의 마감 여부를 미국 거래소 달력으로 판정한다(SL-02).
    # yfinance 일봉·주봉·월봉은 정규장 체결만 담으므로 prepost 와 무관하게
    # 정규 마감이 그 봉의 끝이다. intraday 는 항상 진행 중 봉이 섞여 대상이 아니다.
    tf_map = {"1d": "day", "1wk": "week", "1mo": "month"}
    bar_state = None
    if interval in tf_map:
        bar_state = _bar_state(
            timeframe=tf_map[interval],
            rows=rows,
            market_calendar=us_calendar(),
        )
    return _append_result_meta(
        "\n".join(lines),
        _us_meta(
            kind="bars",
            ticker=ticker,
            data_as_of=(rows[-1].get("date") or rows[-1].get("datetime")),
            bar_state=bar_state,
        ),
    )


@mcp.tool()
@safe_us_tool
@track_metrics("get_us_financials")
async def get_us_financials(ticker: str) -> str:
    """US stock financials — 미국 주식 재무지표 (PER, PBR, PEG, ROE, 배당률 · US valuation ratios).
    "AAPL PER", "Apple 재무", "NVDA valuation", "forward P/E" 같은 질문에 사용합니다.

    Trailing / Forward P/E, PEG, P/B, P/S, EPS, ROE, ROA, 부채비율, 마진, 성장률,
    배당수익률, 배당성향을 반환합니다.

    Args:
        ticker: US 티커 (예: "AAPL")
    """
    data = await us.get_financial_info(ticker)
    if data is None:
        return f"티커 '{ticker}'의 재무 데이터를 가져올 수 없습니다."

    trade_ccy = data.get("currency") or "?"
    fin_ccy = data.get("financial_currency") or "?"
    mixed = trade_ccy != "?" and fin_ccy != "?" and trade_ccy != fin_ccy

    lines = [f"**{data['ticker']}** 재무지표", ""]
    lines.append(f"거래통화: {trade_ccy} / 재무제표 통화: {fin_ccy}")
    if mixed:
        # ADR은 주가와 재무제표의 통화가 달라 yfinance가 두 통화를 섞어 비율을
        # 계산한다. 값을 임의로 고치지 않고 어떤 항목이 오염됐는지만 명시한다.
        lines.append(
            f"⚠️ **통화 불일치** — 주가는 {trade_ccy}, 재무제표는 {fin_ccy}입니다. "
            f"주가와 재무를 함께 쓰는 항목(**P/B, P/S, PEG, Trailing/Forward P/E**)은 "
            f"두 통화를 섞어 계산된 값이라 그대로 해석하면 안 됩니다. "
            f"Per Share 항목도 EPS는 {trade_ccy}, Book Value·Revenue/Share는 "
            f"{fin_ccy} 기준일 수 있습니다. 비율(%)·성장률은 통화 무관이라 유효합니다."
        )
    lines.append("")
    lines.append("## Valuation" + (" ⚠️통화혼합" if mixed else ""))
    lines.append(f"- Trailing P/E: {_fmt_num(data.get('trailing_pe'))}")
    lines.append(f"- Forward P/E: {_fmt_num(data.get('forward_pe'))}")
    lines.append(f"- PEG Ratio: {_fmt_num(data.get('peg_ratio'))}")
    lines.append(f"- P/B: {_fmt_num(data.get('price_to_book'))}")
    lines.append(f"- P/S (TTM): {_fmt_num(data.get('price_to_sales'))}")

    lines.append("")
    lines.append("## Per Share" + (" ⚠️통화혼합" if mixed else f" ({trade_ccy}/{fin_ccy})"))
    lines.append(f"- EPS (Trailing): {_fmt_num(data.get('eps_trailing'))}")
    lines.append(f"- EPS (Forward): {_fmt_num(data.get('eps_forward'))}")
    lines.append(f"- Revenue/Share: {_fmt_num(data.get('revenue_per_share'))}")
    lines.append(f"- Book Value: {_fmt_num(data.get('book_value'))}")

    lines.append("")
    lines.append("## Profitability")
    lines.append(f"- ROE: {_fmt_ratio(data.get('return_on_equity'))}")
    lines.append(f"- ROA: {_fmt_ratio(data.get('return_on_assets'))}")
    lines.append(f"- Profit Margin: {_fmt_ratio(data.get('profit_margin'))}")
    lines.append(f"- Operating Margin: {_fmt_ratio(data.get('operating_margin'))}")

    lines.append("")
    lines.append("## Balance & Growth")
    lines.append(f"- Debt/Equity: {_fmt_num(data.get('debt_to_equity'))}")
    lines.append(f"- Current Ratio: {_fmt_num(data.get('current_ratio'))}")
    lines.append(f"- Revenue Growth (YoY): {_fmt_ratio(data.get('revenue_growth'))}")
    lines.append(f"- Earnings Growth (YoY): {_fmt_ratio(data.get('earnings_growth'))}")

    lines.append("")
    lines.append("## Dividend")
    lines.append(f"- Dividend Yield: {_fmt_yield(data.get('dividend_yield'))}")
    lines.append(f"- Payout Ratio: {_fmt_ratio(data.get('payout_ratio'))}")

    # 이 비율들의 재무 항목이 어느 분기말 기준인지(SL-03). 공급자가 기준일을
    # 안 주면 "언제 것인지 모르는 숫자"이고, 그걸 조용히 complete 로 내보내지
    # 않는다. P/E 등 가격 결합 비율의 가격 쪽은 현재가 기준이다.
    basis_q = data.get("most_recent_quarter")
    lines.append("")
    if basis_q:
        lines.append(f"※ 재무 항목 기준: 최근 분기말 {basis_q}"
                     + (f" (회계연도말 {data.get('last_fiscal_year_end')})"
                        if data.get("last_fiscal_year_end") else "")
                     + ". 가격 결합 비율(P/E·P/B·P/S)의 가격 쪽은 현재가 기준입니다.")
        completeness, coverage, warns = rmeta.COMPLETE, None, None
    else:
        lines.append("※ 공급자가 이 비율들의 기준일을 주지 않았습니다. 언제 것인지 확인되지 않은 값입니다.")
        completeness = rmeta.PARTIAL
        coverage = {"truncated": False, "coverage_complete": False, "reason": "unknown"}
        warns = ["비율 데이터의 기준일이 확인되지 않습니다. 최신 실적 반영 여부를 알 수 없습니다."]
    return _append_result_meta(
        "\n".join(lines),
        _us_meta(
            kind="filing",
            ticker=ticker,
            data_as_of=basis_q,
            data_completeness=completeness,
            coverage=coverage,
            extra=({"period_coverage": {
                "cycle": "ttm_or_mrq",
                "latest_period": basis_q,
                "fiscal_year_end": data.get("last_fiscal_year_end"),
                "currency": data.get("financial_currency") or data.get("currency"),
                "basis": "period_end",
            }} if basis_q else None),
            warnings=warns,
        ),
    )


@mcp.tool()
@safe_us_tool
@track_metrics("get_us_earnings")
async def get_us_earnings(ticker: str) -> str:
    """US earnings calendar — 다음 실적 발표일 + 최근 EPS 서프라이즈 이력 (US earnings date / EPS surprise).
    "AAPL 실적 언제", "NVDA earnings date", "Tesla 다음 실적" 같은 질문에 사용합니다.

    미국 시장은 분기 실적(10-Q)이 주가 변동의 핵심 이벤트입니다.

    Args:
        ticker: US 티커 (예: "NVDA")
    """
    data = await us.get_earnings(ticker)
    if data is None:
        return f"티커 '{ticker}'를 찾을 수 없습니다."

    upcoming = data.get("upcoming", []) or []
    history = data.get("history", []) or []
    # 사건 레코드(ev_data)가 있으면 일정·이력이 비어도 내용이 있다(SL-05).
    _ev_probe = await us.get_earnings_events(ticker)
    if not upcoming and not history and not (_ev_probe and _ev_probe.get("events")):
        return f"**{data['ticker']}** 실적 일정 & 이력\n\n실적 데이터 없음 (ETF나 개별 주식이 아닌 자산은 해당 없음)."

    lines = [f"**{data['ticker']}** 실적 일정 & 이력"]
    if upcoming:
        lines.append("")
        lines.append("## 🗓️ 예정")
        for e in upcoming[:4]:
            date = str(e.get("earnings_date", ""))[:10]
            est = e.get("eps_estimate")
            lines.append(f"- {date} (EPS 추정: {_fmt_num(est)})")
    else:
        lines.append("")
        lines.append("예정된 실적 발표일 정보 없음.")

    if history:
        lines.append("")
        lines.append("## 📊 최근 실적 서프라이즈")
        lines.append("발표일 | EPS 추정 | 실제 EPS | 서프라이즈 %")
        lines.append("---|---|---|---")
        for e in history[:8]:
            date = str(e.get("earnings_date", ""))[:10]
            est = e.get("eps_estimate")
            rep = e.get("reported_eps")
            sur = e.get("surprise(%)") or e.get("surprise_%") or e.get("surprise")
            lines.append(f"{date} | {_fmt_num(est)} | {_fmt_num(rep)} | {_fmt_num(sur)}%")

    # ── 사건별 레코드(SL-05): EPS·매출·현금흐름·발표 시각을 한 사건으로 ──
    ev_data = _ev_probe
    events_meta: list[dict] = []
    quality_warns: list[str] = []
    if ev_data and ev_data.get("events"):
        rev_by = ev_data.get("revenue_by_period") or {}
        ocf_by = ev_data.get("ocf_by_period") or {}
        rev_est_by = ev_data.get("revenue_estimates") or {}
        ccy = ev_data.get("financial_currency") or "USD"
        now_day = _now_kst().date().isoformat()
        past = [e for e in ev_data["events"]
                if e.get("eps_actual") is not None and (e.get("date") or "") <= now_day]

        lines.append("")
        lines.append(f"## 🧾 사건별 기록 (통화: {ccy}, 최근 {min(len(past), 6)}건)")
        lines.append("발표(ET) | 세션 | 회계기간 | EPS 예상->실제 | 매출 예상->실제 | OCF | 판정")
        lines.append("---|---|---|---|---|---|---")
        for e in past[:6]:
            session = _earnings_session(e.get("timestamp") or "")
            fiscal = _match_fiscal_period(e.get("date"), list(rev_by))
            rev_act = rev_by.get(fiscal) if fiscal else None
            rev_est = rev_est_by.get(fiscal) if fiscal else None
            rev_spct = (round((rev_act - rev_est) / rev_est * 100, 2)
                        if isinstance(rev_act, (int, float))
                        and isinstance(rev_est, (int, float)) and rev_est else None)
            ocf = ocf_by.get(fiscal) if fiscal else None

            eps_s = e.get("eps_surprise_pct")
            parts = []
            if isinstance(eps_s, (int, float)):
                parts.append("EPS beat" if eps_s > 0 else
                             ("EPS miss" if eps_s < 0 else "EPS in-line"))
            if rev_spct is not None:
                parts.append("매출 beat" if rev_spct > 0 else
                             ("매출 miss" if rev_spct < 0 else "매출 in-line"))
            else:
                parts.append("매출 예상 확인 불가")
            verdict = " · ".join(parts) if parts else "확인 불가"

            session_ko = {"pre_market": "장전", "post_market": "장후",
                          "during_market": "장중", "unknown": "시각미상"}[session]
            ts_short = (e.get("timestamp") or "")[:16].replace("T", " ")
            fmt_b = lambda v: f"{v/1e9:,.1f}B" if isinstance(v, (int, float)) else "-"
            lines.append(
                f"{ts_short} | {session_ko} | {fiscal or '미매칭'} | "
                f"{e.get('eps_estimate') if e.get('eps_estimate') is not None else '-'}"
                f"->{e.get('eps_actual')}"
                + (f" ({eps_s:+.1f}%)" if isinstance(eps_s, (int, float)) else "")
                + f" | {fmt_b(rev_est)}->{fmt_b(rev_act)}"
                + (f" ({rev_spct:+.1f}%)" if rev_spct is not None else "")
                + f" | {fmt_b(ocf)} | {verdict}"
            )
            events_meta.append({
                "date": e.get("date"), "timestamp": e.get("timestamp"),
                "session": session, "fiscal_period": fiscal,
                "eps_estimate": e.get("eps_estimate"),
                "eps_actual": e.get("eps_actual"),
                "eps_surprise_pct": eps_s,
                "revenue_actual": rev_act, "revenue_estimate": rev_est,
                "revenue_surprise_pct": rev_spct,
                "operating_cash_flow": ocf,
                "verdict": verdict,
            })
        lines.append("")
        lines.append("※ 매출 **예상**은 공급자가 현재·다음 분기만 제공합니다 - 과거 사건의 "
                     "매출 예상이 '-' 이면 확인 불가이지 예상이 없었다는 뜻이 아닙니다.")
        lines.append("※ **실적의 질(가이던스 변화·일회성 조정)은 이 피드로 확인 불가**입니다. "
                     "발표 당일 8-K 원문으로 확인하세요: "
                     "`get_us_filings(ticker, forms=[\"8-K\"])` -> "
                     "`get_us_filing_detail(..., find=\"guidance\")`.")
        lines.append("※ 반응창: 장전 발표는 당일, 장후 발표는 다음 거래일이 첫 반응입니다 - "
                     "`get_us_event_reaction(ticker, event_date)` 로 정렬하세요.")
        quality_warns.append(
            "가이던스·일회성 조정 근거가 이 피드에 없어 실적의 질은 확인 불가입니다. "
            "8-K 원문 확인 전에는 'EPS beat=좋은 실적'으로 단정하지 마세요."
        )

    return _append_result_meta(
        "\n".join(lines),
        _us_meta(
            kind="filing", ticker=ticker,
            data_as_of=(events_meta[0]["date"] if events_meta else None),
            extra=({"earnings_events": {
                "events": events_meta,
                "quality_assessable": False,
                "quality_missing": ["guidance", "one_off_adjustments"],
            }} if events_meta else None),
            warnings=quality_warns or None,
        ),
    )


@mcp.tool()
@safe_us_tool
@track_metrics("get_us_event_reaction")
async def get_us_event_reaction(
    ticker: str,
    event_date: str,
    session: str = "auto",
    after: int = 5,
) -> str:
    """US event reaction — 발표 세션에 맞는 기준 거래일로 주가 반응을 정렬 (SL-05).

    미국 실적은 장전(BMO)·장후(AMC) 발표가 갈린다. 장전 발표는 **당일** 봉이,
    장후 발표는 **다음 거래일** 봉이 첫 반응이다. 이 정렬을 자동으로 한다.
    완성된 일봉만 쓴다(진행 중 봉 제외).

    Args:
        ticker: US 티커
        event_date: 발표일 YYYY-MM-DD
        session: "auto"(실적 발표 시각에서 자동 판정) / "pre" / "post" / "unknown"
        after: 사건 후 비교 거래일 수 (기본 5, 최대 20)
    """
    after = max(1, min(int(after), 20))
    day = rmeta.normalize_day(event_date)
    if not day:
        return f"event_date 를 날짜로 읽지 못했습니다: {event_date}"

    warns: list[str] = []
    resolved = session
    if session == "auto":
        resolved = "unknown"
        try:
            ev = await us.get_earnings_events(ticker)
            for e in (ev or {}).get("events", []):
                if e.get("date") == day:
                    resolved = _earnings_session(e.get("timestamp") or "")
                    break
        except Exception:
            pass
        if resolved == "unknown":
            warns.append(
                f"{day} 발표 시각을 실적 일정에서 찾지 못했습니다. "
                "세션을 알면 session=\"pre\"/\"post\" 로 지정하세요."
            )
    elif session == "pre":
        resolved = "pre_market"
    elif session == "post":
        resolved = "post_market"

    if resolved == "unknown":
        # 시각을 모르면 보수적으로 다음 거래일을 기준으로 잡는다 - 장후 발표를
        # 당일로 정렬하면 발표 전 가격을 반응으로 읽게 되는 쪽이 더 위험하다.
        warns.append("발표 시각 미상 - 보수적으로 다음 거래일을 기준 거래일로 씁니다.")

    raw = await us.get_history(ticker, period="1y", interval="1d")
    if not raw:
        return f"티커 '{ticker}'의 일봉을 가져올 수 없습니다."
    bars, _exc = split_valid_bars(raw)
    # 완성된 봉만 쓴다 - 진행 중인 오늘 봉으로 반응을 재면 마감 때 값이 바뀐다.
    cal = us_calendar()
    while bars:
        last_day = rmeta.normalize_day(bars[-1].get("date") or bars[-1].get("datetime"))
        try:
            closed = cal.is_period_closed(_dt.date.fromisoformat(last_day), "day",
                                          _now_kst())
        except Exception:
            closed = None
        if closed:
            break
        bars = bars[:-1]
        warns.append(f"진행 중인 봉({last_day})은 반응 계산에서 제외했습니다.")
    if len(bars) < 3:
        return f"티커 '{ticker}'의 완성된 일봉이 부족합니다."

    dates = [rmeta.normalize_day(b.get("date") or b.get("datetime")) for b in bars]

    if resolved == "pre_market":
        basis_idx = next((i for i, d in enumerate(dates) if d >= day), None)
        basis_selection = "same_day_pre_market"
    elif resolved == "during_market":
        basis_idx = next((i for i, d in enumerate(dates) if d >= day), None)
        basis_selection = "same_day_during_market"
        warns.append("장중 발표 - 당일 봉에는 발표 전 체결이 섞여 있습니다.")
    else:   # post_market / unknown -> 다음 거래일
        basis_idx = next((i for i, d in enumerate(dates) if d > day), None)
        basis_selection = ("next_trading_day_post_market"
                           if resolved == "post_market" else
                           "next_trading_day_unknown_session")

    if basis_idx is None:
        return _append_result_meta(
            f"# US 이벤트 반응 - {ticker} ({day})\n\n"
            "사건 이후의 완성된 거래일이 아직 없습니다. 다음 정규장 이후 다시 조회하세요.",
            _us_meta(kind="bars", ticker=ticker, data_as_of=dates[-1],
                     data_completeness=rmeta.NONE,
                     coverage={"truncated": False, "coverage_complete": False,
                               "reason": "insufficient_post_event_history"},
                     warnings=warns or None),
        )
    if basis_idx == 0:
        return f"사건({day}) 이전 봉이 없어 기준 대비 수익률을 만들 수 없습니다."

    basis_date = dates[basis_idx]
    ref_close = float(bars[basis_idx - 1]["close"])   # 직전 완성봉 종가

    offsets = [0, 1]
    if after >= 3:
        offsets.append(after // 2 if after >= 4 else after)
    if after not in offsets:
        offsets.append(after)
    points = []
    actual_after = min(after, len(bars) - 1 - basis_idx)
    for off in sorted(set(offsets)):
        idx = basis_idx + off
        if idx >= len(bars):
            points.append((f"D+{off}", None, None))
            continue
        close = float(bars[idx]["close"])
        points.append((f"D+{off}" if off else "D0", dates[idx],
                       (close - ref_close) / ref_close * 100))

    session_ko = {"pre_market": "장전", "post_market": "장후",
                  "during_market": "장중", "unknown": "시각미상"}.get(resolved, resolved)
    lines = [
        f"# US 이벤트 반응 - {ticker} (발표일 {day}, {session_ko})",
        f"기준 거래일: **{basis_date}** ({basis_selection}) · "
        f"기준가: 직전 완성봉({dates[basis_idx - 1]}) 종가 {ref_close:,.2f}",
        "",
        "| 지점 | 거래일 | 기준 대비 |",
        "|---|---|---:|",
    ]
    for label, d, pct in points:
        if pct is None:
            lines.append(f"| {label} | - | - |")
        else:
            lines.append(f"| {label} | {d} | {pct:+.2f}% |")
    short = actual_after < after
    if short:
        lines.append("")
        lines.append(f"⚠️ 사건 후 완성 거래일이 {actual_after}일뿐입니다"
                     f"(요청 D+{after}). 빈 지점은 아직 거래일이 없는 것입니다.")
    lines.append("")
    lines.append(f"_{CORPORATE_ACTION_NOTE}_")

    return _append_result_meta(
        "\n".join(lines),
        _us_meta(
            kind="bars", ticker=ticker,
            data_as_of=dates[-1],
            data_completeness=rmeta.PARTIAL if short else rmeta.COMPLETE,
            coverage=({"requested": {"unit": "trading_day", "after": after},
                       "effective": {"unit": "trading_day", "after": actual_after},
                       "truncated": True, "coverage_complete": False,
                       "reason": "insufficient_post_event_history"} if short else None),
            price_adjustment=price_adjustment_meta(),
            extra={"event_window": {
                "requested": {"after": after},
                "actual": {"after": actual_after},
                "basis_trading_date": basis_date,
                "basis_selection": basis_selection,
                "session": resolved,
                "last_usable_trading_date": dates[-1],
            }},
            warnings=warns or None,
        ),
    )


@mcp.tool()
@safe_us_tool
@track_metrics("get_us_analyst")
async def get_us_analyst(ticker: str) -> str:
    """US analyst ratings — 미국 주식 애널리스트 목표주가 + 투자의견 (US analyst price target / rating).
    "AAPL 목표가", "NVDA analyst rating", "Tesla buy/hold", "Wall Street 의견" 같은 질문에 사용합니다.

    목표주가(mean/high/low) + buy/hold/sell 분포 + 최근 업·다운그레이드를 반환합니다.

    Args:
        ticker: US 티커 (예: "NVDA")
    """
    data = await us.get_analyst_ratings(ticker)
    if data is None:
        return f"티커 '{ticker}'를 찾을 수 없습니다."

    # ETF·펀드는 애널리스트 커버 없음
    has_any = bool(
        data.get("price_targets") or data.get("recommendation_key")
        or data.get("recommendation_mean") or data.get("analyst_count")
        or data.get("recommendations_by_month") or data.get("recent_upgrades_downgrades")
    )
    if not has_any:
        return f"**{data['ticker']}** 애널리스트 의견\n\n애널리스트 커버리지 없음 (ETF나 신규·소형주의 경우 해당 없음)."

    lines = [f"**{data['ticker']}** 애널리스트 의견"]

    cur = data.get("current_price")
    targets = data.get("price_targets") or {}
    if targets:
        lines.append("")
        lines.append("## 🎯 목표주가")
        if cur is not None:
            lines.append(f"현재가: ${cur:,.2f}")
        for label, key in [("평균", "mean"), ("중앙값", "median"), ("최고", "high"), ("최저", "low")]:
            v = targets.get(key)
            if v is not None:
                upside = ((v - cur) / cur * 100) if cur else None
                up_str = f" ({upside:+.1f}%)" if upside is not None else ""
                lines.append(f"- {label}: ${v:,.2f}{up_str}")

    rec_key = data.get("recommendation_key")
    rec_mean = data.get("recommendation_mean")
    n_analysts = data.get("analyst_count")
    if rec_key or rec_mean:
        lines.append("")
        lines.append("## 📈 투자의견")
        if rec_key:
            lines.append(f"- 종합: **{rec_key.upper()}**")
        if rec_mean is not None:
            lines.append(f"- 평균 점수: {rec_mean:.2f} (1=Strong Buy, 5=Strong Sell)")
        if n_analysts:
            lines.append(f"- 커버 애널리스트: {n_analysts}명")

    rec_monthly = data.get("recommendations_by_month") or []
    if rec_monthly:
        lines.append("")
        lines.append("## 월별 의견 분포 (최근)")
        lines.append("기간 | Strong Buy | Buy | Hold | Sell | Strong Sell")
        lines.append("---|---|---|---|---|---")
        for r in rec_monthly[:4]:
            period = r.get("period", "-")
            lines.append(
                f"{period} | {r.get('strongBuy', 0)} | {r.get('buy', 0)} | "
                f"{r.get('hold', 0)} | {r.get('sell', 0)} | {r.get('strongSell', 0)}"
            )

    # 하위 표마다 기준 시점이 다르다(요구 2·3). 목표가·분포·추정치는 조회
    # 시점 스냅샷인데, 업·다운그레이드 표본은 공급자가 갱신을 멈춘 채 몇 년
    # 전에 끝나 있기도 하다(실측 META: 2024-09-30). 각 표의 기준을 따로 적고,
    # 오래된 이력이 섞이면 응답 전체를 complete 로 내지 않는다.
    section_coverage: dict = {}
    if targets:
        section_coverage["price_targets"] = {"basis": "snapshot"}
    if rec_key or rec_mean:
        section_coverage["recommendation"] = {"basis": "snapshot"}
    if rec_monthly:
        section_coverage["recommendation_trend"] = {"basis": "relative_months"}

    updown = data.get("recent_upgrades_downgrades") or []
    updown_latest = None
    updown_stale = False
    if updown:
        dates = sorted(str(u.get("GradeDate", ""))[:10] for u in updown
                       if str(u.get("GradeDate", ""))[:10])
        if dates:
            updown_latest = dates[-1]
            try:
                age = (_now_kst().date()
                       - _dt.date.fromisoformat(updown_latest)).days
            except ValueError:
                age = None
            updown_stale = age is not None and age > 90
            section_coverage["upgrades_downgrades"] = {
                "latest": updown_latest,
                "age_days": age,
                "stale": updown_stale,
            }
    if updown:
        lines.append("")
        lines.append("## 🔄 최근 업·다운그레이드"
                     + (f" (표본 마지막: {updown_latest})" if updown_latest else ""))
        if updown_stale:
            lines.append(
                f"⚠️ 이 표본은 {updown_latest} 에서 멈춰 있습니다. 위의 목표가·분포"
                "(조회 시점 스냅샷)와 기준 시점이 달라 나란히 읽으면 안 됩니다."
            )
        for u in updown[:6]:
            date = str(u.get("GradeDate", ""))[:10]
            firm = u.get("Firm", "-")
            from_g = u.get("FromGrade", "")
            to_g = u.get("ToGrade", "")
            action = u.get("Action", "")
            lines.append(f"- {date} · {firm} · {from_g or '-'} → {to_g or '-'} ({action})")

    # 애널리스트 추정치 (EPS / Revenue / 성장률)
    estimates = await us.get_analyst_estimates(ticker)
    if estimates:
        eps_est = estimates.get("earnings_estimate") or []
        rev_est = estimates.get("revenue_estimate") or []
        eps_rev = estimates.get("eps_revisions") or []
        # 추정치는 재무제표 통화로 나온다. ADR은 주가 통화와 갈리므로 그 통화로 찍는다.
        est_ccy_code = estimates.get("financial_currency") or estimates.get("currency")
        est_ccy = _ccy_prefix(est_ccy_code)
        est_mixed = bool(
            estimates.get("financial_currency")
            and estimates.get("currency")
            and estimates["financial_currency"] != estimates["currency"]
        )

        def _pick(e: dict, *keys):
            """0을 결측으로 흘리지 않는다. `a or b`는 0을 falsy로 보고 넘겨버려서,
            같은 0인데 평균은 '-', 낮음은 '0.00'으로 갈라져 나왔다(BRK.B 0q)."""
            for k in keys:
                if e.get(k) is not None:
                    return e[k]
            return None

        if eps_est or rev_est:
            lines.append("")
            lines.append("## 🔮 애널리스트 추정치")
            if est_mixed:
                lines.append(
                    f"⚠️ 재무 통화 **{estimates['financial_currency']}** · 주가 통화 "
                    f"{estimates['currency']} — 아래 금액은 재무 통화 기준입니다."
                )
            if eps_est:
                lines.append("### EPS Estimate")
                lines.append(
                    f"기간 | 평균 | 낮음 | 높음 | # 애널리스트 | YoY 성장 "
                    f"(통화 {est_ccy_code or '?'})"
                )
                lines.append("---|---|---|---|---|---")
                for e in eps_est:
                    per = _pick(e, "period", "index") or "-"
                    avg = _pick(e, "avg", "average")
                    lo = e.get("low")
                    hi = e.get("high")
                    n = _pick(e, "numberofanalysts", "numberofanalyst")
                    gr = e.get("growth")
                    gr_s = f"{gr*100:+.2f}%" if isinstance(gr, (int, float)) else "-"
                    lines.append(
                        f"{per} | {_fmt_num(avg, currency=est_ccy)} | "
                        f"{_fmt_num(lo, currency=est_ccy)} | {_fmt_num(hi, currency=est_ccy)} | "
                        f"{int(n) if isinstance(n, (int, float)) else '-'} | {gr_s}"
                    )
            if rev_est:
                lines.append("")
                lines.append("### Revenue Estimate")
                lines.append(f"기간 | 평균 | 낮음 | 높음 | YoY 성장 (통화 {est_ccy_code or '?'})")
                lines.append("---|---|---|---|---")
                for e in rev_est:
                    per = _pick(e, "period", "index") or "-"
                    avg = _pick(e, "avg", "average")
                    lo = e.get("low")
                    hi = e.get("high")
                    gr = e.get("growth")
                    gr_s = f"{gr*100:+.2f}%" if isinstance(gr, (int, float)) else "-"
                    lines.append(
                        f"{per} | {_fmt_num(avg, currency=est_ccy)} | "
                        f"{_fmt_num(lo, currency=est_ccy)} | {_fmt_num(hi, currency=est_ccy)} | {gr_s}"
                    )
        if eps_rev:
            lines.append("")
            lines.append("### 최근 EPS 추정 변경 (7일/30일 up/down)")
            lines.append("기간 | up7d | down7d | up30d | down30d")
            lines.append("---|---|---|---|---")
            for e in eps_rev:
                per = e.get("period") or e.get("index") or "-"
                lines.append(
                    f"{per} | {e.get('upLast7days') or e.get('uplast7days') or 0} | "
                    f"{e.get('downLast7days') or e.get('downlast7days') or 0} | "
                    f"{e.get('upLast30days') or e.get('uplast30days') or 0} | "
                    f"{e.get('downLast30days') or e.get('downlast30days') or 0}"
                )

    return _append_result_meta(
        "\n".join(lines),
        _us_meta(
            kind="filing", ticker=ticker,
            data_completeness=rmeta.PARTIAL if updown_stale else rmeta.COMPLETE,
            coverage=({"truncated": False, "coverage_complete": False,
                       "reason": "source_limit"} if updown_stale else None),
            extra={"section_coverage": section_coverage},
            warnings=([
                f"업·다운그레이드 이력이 {updown_latest} 에서 멈춰 있습니다"
                "(공급자 표본 한계). 현재 목표가·분포와 기준 시점이 다릅니다."
            ] if updown_stale else None),
        ),
    )


@mcp.tool()
@safe_us_tool
@track_metrics("get_us_dividends")
async def get_us_dividends(ticker: str, limit: int = 12) -> str:
    """US dividends — 미국 주식 배당 이력 + ex-date + yield (US dividend history / yield).
    "AAPL 배당", "KO dividend yield", "SCHD 배당 이력", "ex-date" 같은 질문에 사용합니다.

    Args:
        ticker: US 티커 (예: "KO", "JNJ", "SCHD")
        limit: 표시할 최근 배당 건수 (기본 12)
    """
    data = await us.get_dividends(ticker, limit=limit)
    if data is None:
        return f"티커 '{ticker}'를 찾을 수 없습니다."

    lines = [f"**{data['ticker']}** 배당 정보"]

    lines.append("")
    lines.append("## 📊 요약")
    lines.append(f"- Dividend Yield: {_fmt_yield(data.get('dividend_yield'))}")
    lines.append(f"- Annual Dividend Rate: {_fmt_num(data.get('dividend_rate'))}")
    lines.append(f"- 5-Year Avg Yield: {_fmt_yield(data.get('five_year_avg_yield'))}")
    lines.append(f"- Payout Ratio: {_fmt_ratio(data.get('payout_ratio'))}")

    ex_date = data.get("ex_dividend_date")
    if ex_date:
        # yfinance는 Unix timestamp로 주는 경우가 많음
        if isinstance(ex_date, (int, float)):
            ex_date = datetime.utcfromtimestamp(ex_date).strftime("%Y-%m-%d")
        lines.append(f"- 다음 Ex-Dividend Date: {ex_date}")

    last_v = data.get("last_dividend_value")
    last_d = data.get("last_dividend_date")
    if last_v and last_d:
        if isinstance(last_d, (int, float)):
            last_d = datetime.utcfromtimestamp(last_d).strftime("%Y-%m-%d")
        lines.append(f"- 최근 배당: ${last_v:.4f} ({last_d})")

    history = data.get("history") or []
    if history:
        lines.append("")
        lines.append(f"## 배당 이력 (최근 {len(history)}건)")
        lines.append("날짜 | 배당금")
        lines.append("---|---")
        for h in reversed(history):
            date = str(h.get("date", ""))[:10]
            amt = h.get("dividends", 0)
            lines.append(f"{date} | ${amt:.4f}")
    else:
        lines.append("")
        lines.append("배당 이력 없음 (무배당 주식일 수 있습니다).")

    return _append_result_meta("\n".join(lines), _us_meta(kind="filing", ticker=ticker))


# --- US Phase 2 tools ---

@mcp.tool()
@safe_us_tool
@track_metrics("get_us_options")
async def get_us_options(
    ticker: str,
    expiration: str | None = None,
    strikes_around_spot: int = 10,
) -> str:
    """US options chain — 미국 주식 옵션 체인 (calls/puts, IV, OI · US equity options).
    "AAPL options", "Tesla 콜옵션", "NVDA implied volatility", "옵션 체인" 같은 질문에 사용합니다.

    기본: 최근접 만기의 현재가 근처 strike. Greeks (delta/gamma/theta)는 미포함.

    Args:
        ticker: US 티커 (예: "AAPL")
        expiration: 만기일 "YYYY-MM-DD". 미지정 시 가장 가까운 만기.
        strikes_around_spot: 현재가 기준 좌우 strike 개수 (기본 10)
    """
    data = await us.get_options(ticker, expiration=expiration, strikes_around_spot=strikes_around_spot)
    if data is None:
        return f"티커 '{ticker}'를 찾을 수 없습니다."
    if not data.get("expirations"):
        return f"**{data['ticker']}** — 옵션 거래 없음."

    lines = [
        f"**{data['ticker']}** 옵션 체인",
        f"현재가(spot): ${data['spot']:,.2f}" if data.get("spot") else "",
        f"선택 만기: **{data['selected_expiration']}**",
        f"사용 가능 만기: {', '.join(data['expirations'][:8])}{' ...' if len(data['expirations'])>8 else ''}",
    ]

    def _table(title: str, rows: list[dict]) -> list[str]:
        if not rows:
            return [f"\n### {title}", "데이터 없음"]
        out = [f"\n### {title}", "Strike | Last | Bid | Ask | Vol | OI | IV | ITM"]
        out.append("---|---|---|---|---|---|---|---")
        for r in rows:
            iv = r.get("impliedVolatility")
            iv_s = f"{iv*100:.1f}%" if isinstance(iv, (int, float)) else "-"
            out.append(
                f"{r.get('strike', 0):.2f} | {r.get('lastPrice', 0):.2f} | "
                f"{r.get('bid') or 0:.2f} | {r.get('ask') or 0:.2f} | "
                f"{int(r.get('volume') or 0)} | {int(r.get('openInterest') or 0)} | "
                f"{iv_s} | {'Y' if r.get('inTheMoney') else 'N'}"
            )
        return out

    lines.extend(_table("📞 Calls", data.get("calls", [])))
    lines.extend(_table("📉 Puts", data.get("puts", [])))
    return "\n".join(l for l in lines if l)


@mcp.tool()
@safe_us_tool
@track_metrics("get_us_insider")
async def get_us_insider(ticker: str) -> str:
    """US insider trading — 내부자 거래 Form 4 + 최근 6개월 순매수 요약 (US insider Form 4).
    "AAPL insider trading", "NVDA 내부자 매수", "CEO stock sale" 같은 질문에 사용합니다.

    Args:
        ticker: US 티커
    """
    data = await us.get_insider(ticker)
    if data is None:
        return f"티커 '{ticker}'를 찾을 수 없습니다."

    lines = [f"**{data['ticker']}** 내부자 거래"]

    summary = data.get("purchases_last_6m") or {}
    tx = data.get("recent_transactions") or []

    def _tx_action(row: dict) -> str:
        a = (row.get("transaction") or "").strip()
        if not a:
            a = (row.get("text") or "").split(" at price")[0].strip() or "-"
        return a

    # ── 유형별 집계(SL-15 요구 1) ─────────────────────────────────────────
    # 무상 부여·증여·옵션 행사·세금 원천징수는 주식만 움직인 것이지 시장에서
    # 산 것이 아니다. 경제적 방향은 현금 거래로 판단해야 한다.
    categories: dict[str, dict] = {}
    for key in ("open_market_buy", "sale", "grant", "gift", "exercise",
                "tax_withholding", "unknown"):
        categories[key] = {"count": 0, "shares": 0, "value": 0.0}
    for r in tx:
        cat = _classify_insider_tx(r)
        categories[cat]["count"] += 1
        if isinstance(r.get("shares"), (int, float)):
            categories[cat]["shares"] += int(r["shares"])
        if isinstance(r.get("value"), (int, float)):
            categories[cat]["value"] += float(r["value"])

    # ── 기본 헤드라인 = 실제 현금 거래(요구 2) ───────────────────────────
    if tx:
        buys, sells = categories["open_market_buy"], categories["sale"]
        lines.append("")
        lines.append("## 💵 실제 현금 거래만 (부여·증여 제외) - 기본 판단 기준")
        lines.append(f"- 공개시장 매수(Purchase): {buys['count']}건, ${buys['value']:,.0f}")
        lines.append(f"- 매도(Sale): {sells['count']}건, ${sells['value']:,.0f}")
        if not buys["count"]:
            lines.append("- ⚠️ 대금이 실제로 지급된 **매수는 없습니다** - "
                         "아래 제공자 요약의 순매수는 부여·증여입니다.")

        lines.append("")
        lines.append("## 📦 전체 주식 이동 (유형별 - 현금 거래와 별개)")
        lines.append("유형 | 건수 | 주식수 | 금액($)")
        lines.append("---|---:|---:|---:")
        for key, label in (("open_market_buy", "공개시장 매수"), ("sale", "매도"),
                           ("grant", "무상 부여(grant)"), ("gift", "증여(gift)"),
                           ("exercise", "옵션 행사(exercise)"),
                           ("tax_withholding", "세금 원천징수(tax_withholding)"),
                           ("unknown", "unknown")):
            c = categories[key]
            if not c["count"]:
                continue
            lines.append(f"{label} | {c['count']} | {c['shares']:,} | {c['value']:,.0f}")
        if categories["unknown"]["count"]:
            lines.append(
                f"※ unknown {categories['unknown']['count']}건은 원문에 유형·대금이 "
                "비어 있는 이동입니다. 무상 부여 등일 수 있으나 이 데이터로는 확인 "
                "불가라 매수로도 부여로도 세지 않습니다."
            )

    # ── 제공자 요약 - 부호 검산(요구 3) ──────────────────────────────────
    summary_valid = True
    summary_warns: list[str] = []
    if summary:
        net_shares = None
        net_pct = None
        for label, vals in summary.items():
            norm = str(label).strip().lower()
            v = vals.get("shares")
            if not isinstance(v, (int, float)):
                continue
            if norm.startswith("%") and "net" in norm:
                net_pct = float(v)
            elif "net" in norm and "shares" in norm:
                net_shares = float(v)
        if net_shares is not None and net_pct is not None and net_shares * net_pct < 0:
            summary_valid = False
            summary_warns.append(
                f"제공자 요약 자체 모순: 순주식수 {net_shares:+,.0f}주인데 "
                f"순비율은 {net_pct:+.3f}% 로 부호가 다릅니다. 이 요약의 순매수 "
                "방향은 신뢰할 수 없습니다(invalid). 무상 부여가 매수로 섞인 "
                "집계일 가능성이 높으니 위의 현금 거래를 기준으로 읽으세요."
            )

        lines.append("")
        lines.append("## 📊 제공자 요약 (부여·증여 포함 - 참고용)"
                     + (" ⚠️ invalid" if not summary_valid else ""))
        if not summary_valid:
            lines.append("⚠️ " + summary_warns[0])
        for label, vals in summary.items():
            shares = vals.get("shares")
            trans = vals.get("trans")
            if shares is None and trans is None:
                continue
            # 이 dict에는 주식수 행과 **비율 행**이 섞여 있다("% Buy Shares" 등).
            # 둘 다 정수로 자르면 0.069가 "0"이 되어 비율이 통째로 사라진다.
            if isinstance(shares, (int, float)):
                sh_str = f"{shares:.3f}" if str(label).lstrip().startswith("%") else f"{shares:,.0f}"
            else:
                sh_str = "-"
            tr_str = f" ({int(trans)}건)" if isinstance(trans, (int, float)) else ""
            lines.append(f"- {label}: {sh_str}{tr_str}")

    if tx:
        lines.append("")
        lines.append(f"## 🗓️ 최근 거래 ({len(tx)}건)")
        lines.append("일자 | 인사 | 직위 | 유형 | 주식수 | 금액")
        lines.append("---|---|---|---|---|---")
        for t in tx[:15]:
            date = str(t.get("start_date", ""))[:10]
            insider = (t.get("insider") or "-")[:25]
            pos = (t.get("position") or "-")[:20]
            # transaction이 빈 문자열인 경우가 많은데, text에 "Stock Gift at price
            # 0.00 per share" 같은 진짜 유형이 들어 있다. 이걸 버리면 대금이 $0인
            # 이유(증여·무상지급)를 알 수 없어 "왜 0원이지?"가 된다.
            action = _tx_action(t)
            shares = t.get("shares") or 0
            value = t.get("value")
            val_s = f"${value:,.0f}" if isinstance(value, (int, float)) else "-"
            lines.append(f"{date} | {insider} | {pos} | {action[:24]} | {int(shares):,} | {val_s}")

    if not summary and not tx:
        lines.append("")
        lines.append("내부자 거래 정보 없음 (ETF나 펀드는 해당 없음).")

    # 현재 내부자 명단 (roster)
    roster = await us.get_insider_roster(ticker)
    if roster:
        lines.append("")
        lines.append(f"## 👥 현재 내부자 명단 ({len(roster)})")
        lines.append("이름 | 직위 | 보유주식 | 최근 거래")
        lines.append("---|---|---|---")
        for r in roster[:15]:
            name = (r.get("name") or "-")[:25]
            pos = (r.get("position") or "-")[:25]
            sh = r.get("most_recent_transaction") or r.get("position_direct_date") or ""
            held = r.get("position_direct") or r.get("shares_owned_directly") or r.get("shares")
            held_s = f"{int(held):,}" if isinstance(held, (int, float)) else "-"
            recent = str(r.get("latest_transaction_date", ""))[:10]
            lines.append(f"{name} | {pos} | {held_s} | {recent}")

    tx_dates = sorted(str(r.get("start_date", ""))[:10] for r in tx
                      if str(r.get("start_date", ""))[:10])
    latest_tx = tx_dates[-1] if tx_dates else None
    insider_coverage = {
        "provider_window": "6m",
        "transactions_listed": len(tx),
        "latest_transaction_date": latest_tx,
        "categories": categories,
        "provider_summary_valid": summary_valid,
    }
    return _append_result_meta(
        "\n".join(lines),
        _us_meta(
            kind="filing", ticker=ticker,
            data_as_of=latest_tx,
            data_completeness=rmeta.COMPLETE if summary_valid else rmeta.PARTIAL,
            coverage=(None if summary_valid else
                      {"truncated": False, "coverage_complete": False,
                       "reason": "unknown"}),
            extra={"insider_coverage": insider_coverage},
            warnings=summary_warns or None,
        ),
    )


@mcp.tool()
@safe_us_tool
@track_metrics("get_us_holders")
async def get_us_holders(ticker: str) -> str:
    """US institutional holders — 기관·뮤추얼펀드 보유 현황 (13F holdings).
    "AAPL institutional holders", "Vanguard 보유", "13F holdings", "who owns" 같은 질문에 사용합니다.

    Args:
        ticker: US 티커
    """
    data = await us.get_holders(ticker)
    if data is None:
        return f"티커 '{ticker}'를 찾을 수 없습니다."

    lines = [f"**{data['ticker']}** 보유 현황"]

    # ETF는 institutional_holders/mutualfund_holders 데이터 없음
    has_any = bool(
        data.get("institutional_holders") or data.get("mutualfund_holders")
        or data.get("held_pct_institutions") or data.get("held_pct_insiders")
    )
    if not has_any:
        lines.append("")
        lines.append("보유자 정보 없음 (ETF·신규 상장·소형주의 경우 데이터가 제공되지 않을 수 있습니다).")
        return _append_result_meta("\n".join(lines), _us_meta(kind="filing", ticker=ticker))

    inst_pct = data.get("held_pct_institutions")
    insd_pct = data.get("held_pct_insiders")
    if inst_pct is not None or insd_pct is not None:
        lines.append("")
        lines.append("## 📊 전체 비중")
        if inst_pct is not None:
            lines.append(f"- 기관 보유: {inst_pct*100:.2f}%")
        if insd_pct is not None:
            lines.append(f"- 내부자 보유: {insd_pct*100:.2f}%")

    inst = data.get("institutional_holders") or []
    if inst:
        lines.append("")
        lines.append("## 🏦 기관 투자자 TOP 10")
        lines.append("기관 | 보유% | 주식수 | 평가액 | 변동% | 보고일")
        lines.append("---|---|---|---|---|---")
        for h in inst:
            holder = ((h.get("holder") or "-").strip())[:30].rstrip()
            pct = h.get("pctheld")
            pct_s = f"{pct*100:.2f}%" if isinstance(pct, (int, float)) else "-"
            shares = h.get("shares") or 0
            value = h.get("value") or 0
            chg = h.get("pctchange")
            chg_s = f"{chg*100:+.2f}%" if isinstance(chg, (int, float)) else "-"
            date = str(h.get("date_reported", ""))[:10]
            lines.append(f"{holder} | {pct_s} | {int(shares):,} | ${value:,.0f} | {chg_s} | {date}")

    mf = data.get("mutualfund_holders") or []
    if mf:
        lines.append("")
        lines.append("## 💼 뮤추얼 펀드 TOP 10")
        lines.append("펀드 | 보유% | 주식수 | 평가액")
        lines.append("---|---|---|---")
        for h in mf:
            holder = ((h.get("holder") or "-").strip())[:35].rstrip()
            pct = h.get("pctheld")
            pct_s = f"{pct*100:.2f}%" if isinstance(pct, (int, float)) else "-"
            shares = h.get("shares") or 0
            value = h.get("value") or 0
            lines.append(f"{holder} | {pct_s} | {int(shares):,} | ${value:,.0f}")

    # Major holders 요약 (breakdown: insiders/institutions/float 등)
    major = await us.get_major_holders(ticker)
    if major and major.get("summary"):
        lines.append("")
        lines.append("## 📋 Breakdown")
        for label, val in major["summary"].items():
            if isinstance(val, (int, float)):
                # yfinance는 0~1 소수 or 라벨+값 혼용
                val_s = f"{val*100:.2f}%" if 0 <= abs(val) <= 1 else f"{val:,.0f}"
            else:
                val_s = str(val)
            lines.append(f"- {label}: {val_s}")

    return _append_result_meta("\n".join(lines), _us_meta(kind="filing", ticker=ticker))


@mcp.tool()
@safe_us_tool
@track_metrics("get_us_short")
async def get_us_short(ticker: str) -> str:
    """US short interest — 공매도 잔고 + % of float + days to cover (US short interest).
    "AAPL short interest", "GME 공매도", "short squeeze" 같은 질문에 사용합니다.

    ⚠️ FINRA bi-monthly 공시라 데이터가 2~4주 stale합니다. 반드시 'date_short_interest'를
    확인하세요.

    Args:
        ticker: US 티커
    """
    data = await us.get_short_interest(ticker)
    if data is None:
        return f"티커 '{ticker}'를 찾을 수 없습니다."

    def _ts(v):
        if isinstance(v, (int, float)):
            return datetime.utcfromtimestamp(v).strftime("%Y-%m-%d")
        return "-"

    # 본문은 보고일과 지연 경고를 이미 보여주는데 메타는 data_as_of=null 에
    # complete 였다(실측 CRSP). 이 숫자의 기준일은 조회일이 아니라 FINRA
    # 보고일이고, 격주 공시라 원래 몇 주 뒤처진 값이다 - 그 나이를 구조화한다.
    report_day = us._epoch_day(data.get("date_short_interest"))
    staleness = None
    short_warns: list[str] = []
    if report_day:
        age_days = max(0, (_now_kst().date()
                           - _dt.date.fromisoformat(report_day)).days)
        # 격주 공시 + 공표 지연을 감안해도 35일이 넘으면 한 사이클 이상 밀린 것
        stale = age_days > 35
        staleness = {"report_date": report_day, "age_days": age_days,
                     "cadence": "FINRA bi-monthly", "stale": stale}
        if stale:
            short_warns.append(
                f"공매도 보고일({report_day})이 {age_days}일 전입니다 - "
                "통상 주기(2~4주)보다 오래됐습니다(stale). 최신 수치가 아닐 수 있습니다."
            )
    else:
        short_warns.append(
            "공매도 수치의 기준일(FINRA 보고일)을 확인하지 못했습니다. "
            "언제 것인지 모르는 값입니다."
        )

    lines = [f"**{data['ticker']}** 공매도 지표"]
    lines.append("")
    lines.append(f"⚠️ 보고 기준일: **{_ts(data.get('date_short_interest'))}** (FINRA bi-monthly — 최대 2~4주 stale)")
    lines.append("")
    lines.append("## 📊 현재 지표")
    ss = data.get("shares_short")
    lines.append(f"- Shares Short: {ss:,.0f}" if isinstance(ss, (int, float)) else "- Shares Short: -")
    spf = data.get("short_percent_of_float")
    if isinstance(spf, (int, float)):
        lines.append(f"- % of Float (shorted): **{spf*100:.2f}%**")
    sr = data.get("short_ratio")
    if isinstance(sr, (int, float)):
        lines.append(f"- Days to Cover (short ratio): {sr:.2f}일")
    fl = data.get("float_shares")
    if isinstance(fl, (int, float)):
        lines.append(f"- Float: {fl:,.0f}주")

    prior = data.get("shares_short_prior_month")
    prior_date = data.get("prior_month_date")
    if isinstance(prior, (int, float)) and isinstance(ss, (int, float)):
        lines.append("")
        lines.append("## 📈 전월 대비")
        change = ss - prior
        pct = (change / prior * 100) if prior else 0
        lines.append(f"- 전월 ({_ts(prior_date)}): {prior:,.0f}")
        lines.append(f"- 변동: {change:+,.0f} ({pct:+.2f}%)")

    return _append_result_meta(
        "\n".join(lines),
        _us_meta(
            kind="filing", ticker=ticker,
            data_as_of=report_day,
            data_completeness=rmeta.COMPLETE if report_day else rmeta.PARTIAL,
            coverage=(None if report_day else
                      {"truncated": False, "coverage_complete": False,
                       "reason": "unknown"}),
            extra=({"staleness": staleness} if staleness else None),
            warnings=short_warns or None,
        ),
    )


@mcp.tool()
@safe_us_tool
@track_metrics("get_us_filings")
async def get_us_filings(ticker: str, limit: int = 15, forms: list[str] | None = None) -> str:
    """US SEC filings — SEC EDGAR 공시 목록 (accession number·원문 URL 포함).

    "AAPL 10-K", "RIVN latest filings", "8-K", "SEC filing" 같은 질문에 사용합니다.
    본문·exhibit 를 읽으려면 결과의 accession number 로 `get_us_filing_detail` 을 부르세요.

    SEC 는 발행사를 티커가 아니라 CIK 로 식별합니다. GOOGL/GOOG 같은 클래스주는
    같은 발행사로 정규화되어 같은 공시 집합이 나오고, 요청 티커는 메타에 보존됩니다.

    Args:
        ticker: US 티커
        limit: 표시할 공시 건수 (기본 15, 최대 100)
        forms: 공시 유형 필터 (예: ["10-Q", "8-K"]). 비우면 전체.
    """
    limit = max(1, min(int(limit), 100))
    requested_forms = [str(f).upper() for f in (forms or [])] or None

    def _fail(reason_text: str, warn: str) -> str:
        # 티커 매핑 실패·조회 실패·0건은 서로 다른 상태다(요구 6). 어느 쪽이든
        # 메타 없이 "공시 정보 없음"으로 끝나면 읽는 쪽이 셋을 구분할 수 없다.
        return _append_result_meta(
            reason_text,
            _us_meta(
                kind="filing", ticker=ticker,
                data_completeness=rmeta.NONE,
                coverage={"requested": {"limit": limit, "forms": requested_forms},
                          "returned_count": 0, "total_count": None,
                          "truncated": False, "coverage_complete": False,
                          "reason": "unknown"},
                warnings=[warn],
            ),
        )

    try:
        issuer = await sec.resolve_issuer(ticker)
    except sec.SecFetchError as e:
        return _fail(f"⚠️ SEC 티커 목록을 조회하지 못했습니다 ({e}). 재시도하세요.",
                     f"SEC 조회 실패: {e}")
    if issuer is None:
        return _fail(
            f"티커 '{ticker}' 를 SEC 발행사에 연결하지 못했습니다. "
            "티커 철자를 확인하거나 get_us_search 로 먼저 검색하세요.",
            f"티커 매핑 실패: {ticker}",
        )

    try:
        subs = await sec.get_submissions(issuer["cik"])
    except sec.SecFetchError as e:
        return _fail(f"⚠️ SEC 공시 목록 조회에 실패했습니다 ({e}). "
                     "0건이 아니라 조회 실패이므로 재시도하세요.",
                     f"SEC 조회 실패: {e}")

    rows = subs["rows"]
    if requested_forms:
        rows = [r for r in rows if str(r.get("form") or "").upper() in requested_forms]
    total_matching = len(rows)
    shown = rows[:limit]
    has_older = subs["has_older"]

    issuer_line = (
        f"발행사: **{issuer['name']}** (CIK {issuer['cik']}) · "
        f"요청 티커 {issuer['requested_ticker']}"
        + (f" -> 대표 {issuer['primary_ticker']}"
           if issuer['primary_ticker'] != issuer['requested_ticker'] else "")
        + (f" · 동일 발행사 티커: {', '.join(issuer['issuer_tickers'])}"
           if len(issuer['issuer_tickers']) > 1 else "")
    )

    if not shown:
        note = ("최근 공시 목록에 해당 유형이 없습니다."
                if requested_forms else "최근 공시 목록이 비어 있습니다.")
        if has_older:
            note += " 더 오래된 공시가 SEC 에 존재하므로 '공시가 없다'고 결론 내리면 안 됩니다."
        return _append_result_meta(
            f"**{issuer['requested_ticker']}** SEC 공시\n{issuer_line}\n\n{note}",
            _us_meta(
                kind="filing", ticker=ticker,
                data_completeness=rmeta.PARTIAL if has_older else rmeta.NONE,
                coverage={"requested": {"limit": limit, "forms": requested_forms},
                          "returned_count": 0,
                          "total_count": None if has_older else total_matching,
                          "truncated": False, "coverage_complete": not has_older,
                          "reason": "source_limit" if has_older else None},
                extra={"issuer": issuer},
                warnings=(["최근 목록 밖(구간 외) 공시가 존재합니다."] if has_older else None),
            ),
        )

    lines = [f"**{issuer['requested_ticker']}** SEC 공시 (표시 {len(shown)}건"
             + (f" / 최근 목록 {total_matching}건" if total_matching > len(shown) else "")
             + ")", issuer_line, ""]
    lines.append("접수일 | 유형 | 보고서 기준일 | accession | 원문")
    lines.append("---|---|---|---|---")
    for r in shown:
        url = sec.filing_url(issuer["cik"], r["accession"], r["primary_document"] or "")
        doc = r.get("primary_document") or "-"
        lines.append(
            f"{r['filing_date']} | **{r['form']}** | {r.get('report_date') or '-'} | "
            f"`{r['accession']}` | [{doc}]({url})"
        )
    lines.append("")
    lines.append("_본문·exhibit 검색: `get_us_filing_detail(ticker, accession_no, find=...)`_")
    if has_older:
        lines.append("_이 목록은 SEC 의 최근 공시 구간입니다. 더 오래된 공시는 여기 없습니다._")

    truncated = len(shown) < total_matching
    incomplete = truncated or has_older
    return _append_result_meta(
        "\n".join(lines),
        _us_meta(
            kind="filing", ticker=ticker,
            data_as_of=shown[0].get("filing_date"),
            data_completeness=rmeta.PARTIAL if incomplete else rmeta.COMPLETE,
            coverage={
                "requested": {"limit": limit, "forms": requested_forms},
                "returned_count": len(shown),
                "total_count": None if has_older else total_matching,
                "truncated": truncated,
                "coverage_complete": not incomplete,
                "reason": ("server_cap" if truncated
                           else "source_limit" if has_older else None),
            },
            extra={"issuer": issuer},
        ),
    )


@mcp.tool()
@safe_us_tool
@track_metrics("get_us_filing_detail")
async def get_us_filing_detail(
    ticker: str,
    accession_no: str,
    find: str | None = None,
    analyze: str | None = None,
    document: str | None = None,
) -> str:
    """US SEC filing 본문 — 원문 키워드 검색·희석/계약 조항 구조화 (SL-04).

    - find: 본문에서 키워드 주변 발췌(전체 매치 수 + 최대 5건 표시).
      매치 0건은 "본문에 없다"가 아니다 - 표기가 다를 수 있다.
    - analyze: "dilution"(증권 수·전환가·워런트·리픽싱·자금용도) 또는
      "contract"(계약금액·기간·해지·최소구매·상대방 비공개)를 범주별 원문
      발췌로 구조화. **못 찾은 범주는 '미확인'이지 '없다'가 아니다.**
    - 둘 다 비우면 문서 목록(본문+exhibit)만 보여준다. 전체 본문은 반환하지
      않는다(10-Q 하나가 수십만 자다).

    Args:
        ticker: US 티커 (발행사 CIK 확인용)
        accession_no: SEC 접수번호 (get_us_filings 결과의 accession)
        find: 본문 검색 키워드 (선택)
        analyze: "dilution" / "contract" (선택)
        document: 읽을 문서 파일명 (비우면 본문. exhibit 는 목록에서 이름 확인)
    """
    def _fail(reason_text: str, warn: str) -> str:
        return _append_result_meta(
            reason_text,
            _us_meta(kind="filing", ticker=ticker,
                     data_completeness=rmeta.NONE,
                     coverage={"truncated": False, "coverage_complete": False,
                               "reason": "unknown"},
                     warnings=[warn]),
        )

    try:
        issuer = await sec.resolve_issuer(ticker)
    except sec.SecFetchError as e:
        return _fail(f"⚠️ SEC 티커 목록 조회 실패 ({e}).", f"SEC 조회 실패: {e}")
    if issuer is None:
        return _fail(f"티커 '{ticker}' 를 SEC 발행사에 연결하지 못했습니다.",
                     f"티커 매핑 실패: {ticker}")

    acc = (accession_no or "").strip()
    # 이 공시의 본문 파일명은 submissions 목록에 있다.
    primary = None
    form = None
    filing_date = None
    try:
        subs = await sec.get_submissions(issuer["cik"])
        for r in subs["rows"]:
            if (r.get("accession") or "").replace("-", "") == acc.replace("-", ""):
                primary, form, filing_date = (r.get("primary_document"),
                                              r.get("form"), r.get("filing_date"))
                break
    except sec.SecFetchError:
        pass

    filing_info = {"accession": acc, "form": form, "filing_date": filing_date,
                   "issuer": issuer}

    # 문서 목록 모드
    if not find and not analyze:
        try:
            docs = await sec.fetch_filing_index(issuer["cik"], acc)
        except sec.SecFetchError as e:
            return _fail(f"⚠️ 문서 목록 조회 실패 ({e}). 접수번호를 확인하세요.",
                         f"SEC 조회 실패: {e}")
        lines = [f"# SEC 공시 문서 (accession={acc})",
                 f"발행사: {issuer['name']} (CIK {issuer['cik']})"
                 + (f" · {form} {filing_date}" if form else ""), "",
                 "파일 | 크기 | URL", "---|---|---"]
        for d in docs[:40]:
            url = sec.filing_url(issuer["cik"], acc, d["name"])
            lines.append(f"{d['name']} | {d.get('size') or '-'} | [열기]({url})")
        lines.append("")
        lines.append("_본문 검색: find=..., 조항 구조화: analyze=\"dilution\"/\"contract\". "
                     "exhibit 를 읽으려면 document=파일명._")
        return _append_result_meta(
            "\n".join(lines),
            _us_meta(kind="filing", ticker=ticker, data_as_of=filing_date,
                     extra={"filing": filing_info,
                            "documents": [d["name"] for d in docs[:40]]}),
        )

    target_doc = document or primary
    if not target_doc:
        return _fail(
            f"accession {acc} 의 본문 파일을 찾지 못했습니다. "
            "document= 로 파일명을 지정하세요 (목록: find/analyze 없이 호출).",
            "본문 파일 미확인",
        )

    try:
        text = await sec.fetch_filing_text(issuer["cik"], acc, target_doc)
    except sec.SecFetchError as e:
        # 원문을 못 가져왔으면 그 무엇도 "없다"고 말할 수 없다(요구 7).
        return _fail(
            f"⚠️ 원문을 가져오지 못했습니다 ({e}). 검색·분석 결과가 0건이라는 뜻이 "
            f"아니라 **확인 불가**입니다. SEC 원문: "
            f"{sec.filing_url(issuer['cik'], acc, target_doc)}",
            f"원문 확인 불가: {e}",
        )

    header = [f"# SEC 공시 본문 (accession={acc}, {target_doc})",
              f"발행사: {issuer['name']} (CIK {issuer['cik']})"
              + (f" · {form} {filing_date}" if form else ""),
              f"원문: {sec.filing_url(issuer['cik'], acc, target_doc)}",
              f"본문 길이: {len(text):,}자", ""]

    if find:
        kw = find.strip()
        positions = sec.find_all_matches(text, kw)
        showed = sec.match_snippets(text, kw, positions[:sec.FIND_MAX_MATCHES])
        lines = list(header)
        if positions:
            if len(positions) > len(showed):
                lines.append(f"**매치:** 전체 {len(positions)}건 중 {len(showed)}건 표시")
            else:
                lines.append(f"**매치:** {len(positions)}건")
            for i, s in enumerate(showed, 1):
                lines += ["", f"## 매치 {i} (위치 ~{s['pos']:,})", "", s["snippet"]]
        else:
            lines.append(f"'{kw}' 매치 0건 - 본문에 해당 내용이 **없다는 뜻이 아닙니다**. "
                         "표기가 다르거나 다른 문서(exhibit)에 있을 수 있습니다.")
        truncated = len(positions) > len(showed)
        complete = bool(positions) and not truncated
        extra = {"filing": filing_info,
                 "match_coverage": {"keyword": kw,
                                    "total_matches": len(positions),
                                    "displayed_matches": len(showed),
                                    "truncated": truncated,
                                    "coverage_complete": complete}}
        if not positions:
            extra["absence_confirmed"] = False
        return _append_result_meta(
            "\n".join(lines),
            _us_meta(kind="filing", ticker=ticker, data_as_of=filing_date,
                     data_completeness=rmeta.COMPLETE if complete else rmeta.PARTIAL,
                     coverage=(None if complete else
                               {"truncated": truncated, "coverage_complete": False,
                                "reason": "server_cap" if truncated else "unknown"}),
                     extra=extra),
        )

    kind_map = {"dilution": (sec.extract_dilution_facts, "희석(증권 수·전환·워런트)"),
                "contract": (sec.extract_contract_facts, "계약 조항")}
    if analyze not in kind_map:
        return _fail(f"analyze 는 'dilution' 또는 'contract' 여야 합니다 (받음: {analyze}).",
                     "잘못된 analyze 값")
    extractor, label = kind_map[analyze]
    facts = extractor(text)
    lines = list(header)
    lines.append(f"## {label} 구조화 - 원문 발췌 기반")
    for category, info in facts["found"].items():
        lines += ["", f"### {category} (매치 {info['total_matches']}건)"]
        for s in info["shown"]:
            lines.append(f"- …{s['snippet'][:400]}")
    if facts["unconfirmed"]:
        lines += ["", "### 미확인 범주",
                  "- " + ", ".join(facts["unconfirmed"]),
                  "- 이 문서에서 패턴으로 찾지 못했다는 뜻이며 **없다는 뜻이 아닙니다**. "
                  "exhibit 나 다른 공시(S-1·424B 등)에 있을 수 있습니다."]
    return _append_result_meta(
        "\n".join(lines),
        _us_meta(kind="filing", ticker=ticker, data_as_of=filing_date,
                 extra={"filing": filing_info,
                        "analysis": {"kind": analyze,
                                     "found_categories": list(facts["found"]),
                                     "unconfirmed_categories": facts["unconfirmed"]}},
                 warnings=([f"미확인 범주({', '.join(facts['unconfirmed'])})는 "
                            "원문에 없다는 뜻이 아니라 이 문서에서 못 찾았다는 뜻입니다."]
                           if facts["unconfirmed"] else None)),
    )


@mcp.tool()
@safe_us_tool
@track_metrics("get_us_news")
async def get_us_news(ticker: str, limit: int = 10, direct_only: bool = False) -> str:
    """US stock news — 미국 주식 관련 뉴스 헤드라인 (관련도 판정 포함).
    "AAPL 뉴스", "Tesla news", "NVDA headlines" 같은 질문에 사용합니다.

    Yahoo 피드에는 다른 종목이 중심인 기사가 섞여 옵니다(실측: AAPL 피드에
    InterDigital·NVDA 단독 기사). 기사마다 관련도(direct/comparison/
    supply_chain/list_or_etf/mention/unrelated)와 근거가 붙습니다.
    이 종목이 주제인 기사만 원하면 direct_only=True.

    Args:
        ticker: US 티커
        limit: 헤드라인 개수 (기본 10)
        direct_only: True 면 이 종목이 주제(direct)인 기사만 남깁니다
    """
    data = await us.get_news(ticker, limit=limit)
    if data is None or not data.get("news"):
        return f"티커 '{ticker}'의 뉴스를 가져올 수 없습니다."

    # 회사명이 있어야 "Apple" 로 쓴 기사도 잡는다. 실패해도 티커로는 판정한다.
    names: list[str] = [ticker.upper()]
    matched_by = "ticker"
    try:
        info = await us.get_info_raw(ticker)
        for key in ("shortName", "longName"):
            nm = (info or {}).get(key) or ""
            # "Apple Inc." -> "Apple" (법인 접미사 제거)
            base = re.sub(
                r"[,.]?\s*(inc|corp|corporation|co|ltd|plc|sa|nv|ag|group|"
                r"holdings?|company)\.?$",
                "", nm.strip(), flags=re.I).strip()
            if base and base.upper() not in {n.upper() for n in names}:
                names.append(base)
                matched_by = "name+ticker"
    except Exception:
        pass

    from collections import Counter as _Counter
    judged = []
    for n in data["news"]:
        rel = _classify_news_relevance(n, tuple(names))
        judged.append((n, rel))
    counts = dict(_Counter(rel["category"] for _, rel in judged))

    excluded_counts: dict = {}
    if direct_only:
        kept = [(n, rel) for n, rel in judged if rel["category"] == "direct"]
        excluded_counts = dict(_Counter(
            rel["category"] for _, rel in judged if rel["category"] != "direct"))
        judged = kept

    lines = [f"**{data['ticker'].upper()}** 최근 뉴스", ""]
    if excluded_counts:
        parts = ", ".join(f"{k} {v}건" for k, v in sorted(excluded_counts.items()))
        lines.append(f"⚠️ direct_only=True - 이 종목이 주제가 아닌 {sum(excluded_counts.values())}건 제외: {parts}")
        lines.append("")
    for n, rel in judged:
        title = n.get("title") or "-"
        pub = (n.get("published") or "")[:16].replace("T", " ")
        provider = n.get("provider") or "-"
        url = n.get("url") or ""
        summary = (n.get("summary") or "").replace("\n", " ").strip()
        # HTML 태그 간단 제거
        summary = re.sub(r"<[^>]+>", "", summary)[:200]
        lines.append(f"### {title}")
        lines.append(f"_{pub} · {provider}_ · 관련도 **{rel['category']}** "
                     f"({rel['relevance_score']}) - {rel['basis']}")
        if summary:
            lines.append(summary)
        if url:
            lines.append(f"[링크]({url})")
        lines.append("")
    if direct_only and not judged:
        lines.append("(이 종목이 주제인 기사가 없습니다 - 기사가 없다는 뜻이 아니라 "
                     "이 피드에 직접 기사가 없다는 뜻입니다)")
        lines.append("")
    lines.append("※ 관련도는 제목·요약 텍스트 기준의 순서 척도입니다. "
                 "unrelated 기사를 이 종목의 재료로 읽지 마세요.")

    return _append_result_meta(
        "\n".join(lines),
        _us_meta(
            kind="snapshot", ticker=ticker,
            extra={"news_relevance": {
                "matched_by": matched_by,
                "names_used": names,
                "categories": counts,
                **({"direct_only": True, "excluded": excluded_counts}
                   if direct_only else {}),
            }},
        ),
    )


# ---------------------------------------------------------------------------
# 컨센서스 / 목표가 / 리포트 / 공시
# ---------------------------------------------------------------------------


@mcp.tool()
@safe_tool
@track_metrics("get_consensus")
async def get_consensus(code: str) -> str:
    """컨센서스 — 증권사 투자의견·목표주가 + 어닝 서프라이즈(컨센서스 대비 잠정치).

    "목표가 얼마야", "컨센서스", "증권사 의견", "적정가", "어닝 서프라이즈/쇼크" 같은
    질문에 사용합니다.

    어닝 서프라이즈 표는 **영업이익·당기순이익**만 다룹니다 (에프앤가이드 원표에
    매출액이 없음). 매출 실적·전망은 `get_financial`을 쓰세요.

    Args:
        code: 종목코드 6자리 (예: "005930")
    """
    import re
    if not re.match(r"^[A-Za-z0-9]{6}$", code):
        return f"⚠️ 종목코드 형식이 올바르지 않습니다: {code}"

    data = await naver_get_consensus(code)

    if not data.get("target_price") and not data.get("opinion"):
        return f"종목코드 {code}의 컨센서스 데이터가 없습니다. 분석 대상 종목이 아니거나 코드를 확인해주세요."

    lines = [f"## 컨센서스 ({code})", ""]

    # 목표주가
    tp = data.get("target_price")
    if tp:
        lines.append(f"**목표주가: {tp:,.0f}원**")

    # 목표주가 추이 — 지금까지 수집만 하고 버렸던 값이다. 현재 목표가 하나로는
    # "기대가 올라가는 중인지 내려가는 중인지"를 알 수 없는데, 방향과 속도가
    # 수준보다 중요할 때가 많다. 전원 매수 의견인데 목표가만 내려가는 경우도 있다.
    hist = data.get("target_price_history") or []
    points = []
    for h in hist:
        y, d = h.get("price"), h.get("date")
        if y is None or d is None:
            continue
        try:
            label = _dt.datetime.fromtimestamp(float(d) / 1000, _dt.timezone.utc).strftime("%Y-%m")
        except (ValueError, OSError, OverflowError):
            continue
        points.append((label, float(y)))
    if len(points) >= 2:
        first, last = points[0], points[-1]
        chg = (last[1] - first[1]) / first[1] * 100 if first[1] else None
        prev = points[-2][1]
        recent = (last[1] - prev) / prev * 100 if prev else None
        lines.append("")
        lines.append("### 📈 목표주가 추이 (기대의 방향)")
        lines.append("기준월 | 목표주가(원)")
        lines.append("---|---:")
        for label, y in points:
            lines.append(f"{label} | {y:,.0f}")
        if chg is not None:
            arrow = "상향" if chg > 1 else ("하향" if chg < -1 else "보합")
            lines.append("")
            lines.append(f"- 전체 구간({first[0]}→{last[0]}): **{chg:+.1f}% {arrow}**")
        if recent is not None:
            arrow2 = "상향" if recent > 1 else ("하향" if recent < -1 else "보합")
            lines.append(f"- 직전 대비: **{recent:+.1f}% {arrow2}**")
        lines.append("- ※ 목표주가는 증권사 추정치이며, 이 표는 그 추정이 "
                     "**어느 방향으로 바뀌어 왔는지**를 보여줍니다.")

    # 투자의견
    opinion = data.get("opinion", {})
    if opinion:
        total = sum(opinion.values())
        parts = [f"{k} {v}건" for k, v in opinion.items() if v > 0]
        # 전 항목이 0이면 "투자의견:  (총 0건)"이라는 빈 라벨만 남았다.
        if parts:
            lines.append(f"투자의견: {', '.join(parts)} (총 {total}건)")
        else:
            lines.append("투자의견: 최근 집계 없음")

        ago = data.get("opinion_1m_ago", {})
        if ago:
            ago_parts = [f"{k} {v}건" for k, v in ago.items() if v > 0]
            lines.append(f"1개월 전: {', '.join(ago_parts)}")

    # 어닝 서프라이즈 (컨센서스 vs 잠정치)
    surprise = data.get("earnings_surprise", {})
    periods = data.get("earnings_surprise_periods", [])
    if surprise and periods:
        period_labels = [f"{p[:4]}.{p[4:]}" for p in periods]
        lines.append("")
        lines.append("### 어닝 서프라이즈 (컨센서스 대비 잠정치)")
        lines.append("")
        lines.append("지표 | 구분 | " + " | ".join(period_labels))
        lines.append("---|---|" + "|".join(["---:"] * len(periods)))

        for metric in ("영업이익", "당기순이익"):
            rows = surprise.get(metric)
            if not rows:
                continue
            for field, field_label in (
                ("consensus", "컨센서스(억원)"),
                ("actual", "잠정치(억원)"),
                ("surprise", "Surprise(%)"),
            ):
                cells = []
                for p in periods:
                    v = (rows.get(p) or {}).get(field)
                    if v is None:
                        cells.append("-")
                    elif field == "surprise":
                        cells.append(f"{v:+.1f}%")
                    else:
                        cells.append(f"{v:,.0f}")
                lines.append(f"{metric} | {field_label} | " + " | ".join(cells))

        dates = data.get("earnings_surprise_dates") or []
        if dates:
            pairs = [f"{lab}: {d}" for lab, d in zip(period_labels, dates)]
            lines.append("")
            lines.append("잠정치 발표(예정)일 · 회계기준 — " + " / ".join(pairs))
        lines.append("")
        lines.append(
            "※ 이 표는 **매출액이 아니라 영업이익·당기순이익**입니다. "
            "`-`는 아직 발표 전이거나 미집계."
        )

    lines.append("")
    lines.append("※ 금액 단위: 억원, 에프앤가이드 기준")
    lines.append("※ 목표주가·투자의견·컨센서스는 증권사 전망치이며 미래 수익을 보장하지 않습니다.")

    # 가장 최근 잠정치 발표일을 기준일로. 이 날짜가 그대로
    # get_event_reaction(event_date=...)로 넘어간다.
    announced = [d for d in (data.get("earnings_surprise_dates") or []) if rmeta.normalize_day(d)]
    meta = _kr_meta(
        kind="filing", code=code,
        data_as_of=announced[-1] if announced else None,
        data_period=f"{periods[-1][:4]}.{periods[-1][4:]}" if periods else None,
    )
    return _append_result_meta("\n".join(lines), meta)


@mcp.tool()
@safe_tool
@track_metrics("get_reports")
async def get_reports(code: str, count: int = 5) -> str:
    """증권사리포트 — 종목의 최근 증권사 분석 리포트 (목표가, 투자의견, 본문 요약, PDF).

    "리포트", "증권사 분석", "애널리스트 의견", "리서치" 같은 질문에 사용합니다.

    Args:
        code: 종목코드 6자리 (예: "005930")
        count: 가져올 리포트 수 (기본 5, 최대 10)
    """
    import re
    if not re.match(r"^[A-Za-z0-9]{6}$", code):
        return f"⚠️ 종목코드 형식이 올바르지 않습니다: {code}"

    count = min(count, 10)
    report_list = await naver_get_reports(code, count)

    if not report_list:
        return f"종목코드 {code}의 증권사 리포트가 없습니다."

    # 각 리포트 상세를 병렬 조회
    import asyncio as _asyncio
    details = await _asyncio.gather(
        *[naver_get_report_detail(r["nid"]) for r in report_list]
    )

    lines = [f"## 증권사 리포트 ({code}, 최근 {len(report_list)}건)", ""]

    for report, detail in zip(report_list, details):
        tp = detail.get("target_price")
        op = detail.get("opinion", "")
        tp_str = f" | 목표가 {tp:,}원" if tp else ""
        op_str = f" | {op}" if op else ""

        # 조회수는 파싱해두고 버리던 값이다. 같은 날 나온 리포트라도 몇 배씩
        # 차이가 나며, "어느 설명이 실제로 읽혔는가"를 재는 몇 안 되는 숫자다.
        views = report.get("views")
        views_str = f" | 조회 {views:,}" if isinstance(views, int) and views > 0 else ""

        lines.append(f"### {report['title']}")
        lines.append(f"{report['broker']} | {report['date']}{tp_str}{op_str}{views_str}")

        summary = detail.get("summary", "")
        if summary:
            lines.append(f"> {summary[:300]}")
        elif PARSE_MISS_KEY in detail:
            lines.append("> _(본문을 읽지 못했습니다 — 요약이 없는 게 아니라 파싱 실패)_")

        pdf = detail.get("pdf_url", "")
        if pdf:
            lines.append(
                f"[PDF 원문]({pdf}) · 본문을 읽으려면 "
                f"`get_report_content(nid=\"{report['nid']}\")` "
                f"— 기본은 앞부분 발췌, `mode=\"full\"` 전체 / `mode=\"link\"` 링크만"
            )

        lines.append("")

    lines.append("※ 목표가·투자의견은 각 증권사 분석 의견이며 미래 수익을 보장하지 않습니다.")

    return _append_result_meta("\n".join(lines), _kr_meta(kind="filing", code=code))


@mcp.tool()
@safe_tool
@track_metrics("get_report_content")
async def get_report_content(
    nid: str, mode: str = "summary", max_chars: int | None = None
) -> str:
    """리포트읽기 — 증권사 리포트 **한 건**의 PDF 본문을 읽어옵니다.

    ⚠️ **한 번에 한 건만.** 리포트 하나가 1만 자를 넘어서, 여러 건을 이어 부르면
    대화 토큰을 통째로 먹습니다. 목록·목표가·짧은 요약은 `get_reports`로 충분하고,
    "이 리포트 자세히", "목표가 근거" 처럼 **한 건을 깊게** 볼 때만 쓰세요.

    **어느 모드를 고를지는 사용자 말에 맞춥니다.** 사용자가 따로 말하지 않으면
    `summary`로 두세요.

    | mode | 무엇을 주나 | 이렇게 말할 때 |
    |---|---|---|
    | `summary` (기본) | 앞 4,000자 발췌 | "리포트 봐줘", "목표가 근거가 뭐야" |
    | `full` | 본문 전체 (상한 50,000자) | "전문 다 보여줘", "빠짐없이", "부록까지" |
    | `link` | 본문 없이 링크만 | "링크만 줘", "내가 직접 볼게" |

    `summary`로 충분한 이유: 실측상 리포트는 앞쪽에 투자의견·목표가 산출 근거·
    실적 추정이 모이고, 뒤쪽은 재무제표 부록과 컴플라이언스 고지문입니다.
    판단 근거는 대개 앞 3,000자 안에 다 있습니다.

    증권사마다 PDF 만드는 방식이 달라, 글자가 아니라 이미지로 렌더링해 내는 곳은
    본문을 읽을 수 없습니다. 그 경우 억지로 몇 글자 내놓지 않고 못 읽었다고
    알려주며 원문 링크로 안내합니다.

    Args:
        nid: 리포트 번호. `get_reports` 결과에 함께 표시됩니다.
        mode: "summary"(기본) / "full" / "link". 한글("요약"·"전문"·"링크")도 됩니다.
        max_chars: 글자 수를 직접 지정하고 싶을 때만. 비우면 mode 기본값을 씁니다.
    """
    import re as _re

    if not _re.match(r"^\d{1,12}$", str(nid)):
        return f"⚠️ 리포트 번호 형식이 올바르지 않습니다: {nid}"

    # 사용자가 한글로 말할 수도 있고 Claude가 영어로 넘길 수도 있다. 둘 다 받는다.
    _MODES = {
        "summary": "summary", "요약": "summary", "핵심": "summary", "발췌": "summary",
        "full": "full", "전문": "full", "전체": "full", "원문": "full",
        "link": "link", "링크": "link", "주소": "link",
    }
    mode_key = _MODES.get(str(mode).strip().lower())
    if mode_key is None:
        return (
            f"⚠️ 알 수 없는 mode: {mode}\n"
            f"summary(앞부분 발췌) / full(전체) / link(링크만) 중에서 골라 주세요."
        )
    cap = max_chars if max_chars else {"summary": 4000, "full": 50000, "link": 0}[mode_key]

    detail = await naver_get_report_detail(str(nid))
    pdf_url = detail.get("pdf_url") or ""
    # 사람이 열어 볼 자리는 원문 PDF보다 네이버 리포트 페이지가 낫다(요약·목표가가
    # 같이 있고, PDF가 없는 리포트도 여기는 열린다).
    page_url = f"{naver_report_read_url}?nid={nid}"

    if mode_key == "link":
        # 본문을 안 볼 거면 PDF를 받을 이유도 없다 — 내려받지 않고 링크만 준다.
        lines = [
            f"# 리포트 링크 (nid={nid})",
            "",
            f"- 리포트 페이지: {page_url}",
        ]
        lines.append(f"- PDF 원문: {pdf_url}" if pdf_url else "- PDF 원문: 제공되지 않는 리포트입니다")
        lines += [
            "",
            "_본문은 가져오지 않았습니다(mode=\"link\"). 내용까지 읽으려면 "
            "mode=\"summary\"(앞부분) 또는 mode=\"full\"(전체)로 다시 부르세요._",
        ]
        return _append_result_meta(
            "\n".join(lines),
            _kr_meta(kind="filing", data_completeness=rmeta.NONE,
                     warnings=["mode=link — 본문을 읽지 않고 링크만 반환했습니다."]),
        )

    if not pdf_url:
        return _append_result_meta(
            f"리포트 {nid}에는 PDF 원문 링크가 없습니다 "
            f"(일부 증권사는 PDF를 공개하지 않습니다).\n\n"
            f"- 리포트 페이지에서 보기: {page_url}\n"
            f"- 요약·목표가·투자의견은 `get_reports`로 확인하실 수 있습니다.",
            _kr_meta(kind="filing", data_completeness=rmeta.NONE,
                     warnings=["PDF 링크 없음 — 이 증권사는 원문을 공개하지 않습니다."]),
        )

    result = await pdfx.fetch_and_extract(pdf_url)

    if result.status != "ok":
        # 못 읽은 이유를 그대로 말한다. '내용 없음'으로 뭉개면 사용자는 리포트가
        # 부실한 줄 안다.
        reason = {
            "image_pdf": "이미지로 렌더링된 PDF라 본문 글자를 읽을 수 없습니다",
            "too_large": "PDF가 너무 커서 받지 않았습니다",
            "failed": "PDF를 받거나 여는 데 실패했습니다",
        }.get(result.status, "본문을 읽지 못했습니다")
        return _append_result_meta(
            f"**{reason}.**\n\n"
            f"- 상세: {result.detail}\n"
            f"- 리포트 페이지에서 보기: {page_url}\n"
            f"- PDF 원문 직접 열기: {pdf_url}\n\n"
            f"본문이 없는 게 아니라 **우리가 못 읽은 것**입니다. "
            f"위 링크로 직접 보시면 되고, 요약·목표가는 `get_reports`로 "
            f"확인하실 수 있습니다.",
            _kr_meta(kind="filing", data_completeness=rmeta.NONE,
                     warnings=[f"PDF 본문 추출 실패({result.status}): {result.detail}"]),
        )

    text = result.text
    truncated = len(text) > cap
    if truncated:
        text = text[:cap]

    heading = "리포트 전문" if mode_key == "full" else "리포트 발췌"
    lines = [
        f"# {heading} (nid={nid})",
        "",
        f"- **전문 보기: {page_url}** (PDF: {pdf_url})",
        f"- PDF {result.pages}쪽 · 본문 {result.chars:,}자 중 "
        + (f"앞 {cap:,}자" if truncated else "전체"),
        "",
        "---",
        "",
        text,
    ]
    if truncated and mode_key == "summary":
        lines.append("")
        lines.append(
            f"_… 이후 {result.chars - cap:,}자 생략(앞부분 발췌). 리포트는 뒤쪽이 "
            f"재무제표 부록과 고지문이라 판단 근거는 대개 위 안에 있습니다. "
            f"전부 필요하면 `mode=\"full\"`, 직접 보시려면 위 원문 링크를 쓰세요._"
        )
    elif truncated:
        lines.append("")
        lines.append(
            f"_… 이후 {result.chars - cap:,}자 생략. 상한({cap:,}자)을 넘겼습니다 — "
            f"이만큼 긴 리포트는 위 원문 링크로 보시는 편이 낫습니다._"
        )
    lines.append("")
    lines.append(
        "※ 증권사 리서치 자료입니다. 목표가·투자의견은 해당 증권사의 분석 의견이며 "
        "미래 수익을 보장하지 않습니다."
    )
    lines.append(
        "※ 위 발췌는 해당 증권사에 저작권이 있습니다. 열람·분석 용도로만 쓰시고 "
        "재배포하지 마세요. 인용하실 때는 원문 링크를 함께 표기해 주세요."
    )

    return _append_result_meta(
        "\n".join(lines),
        _kr_meta(
            kind="filing",
            data_completeness=rmeta.PARTIAL if truncated else rmeta.COMPLETE,
            warnings=[
                f"mode={mode_key} — 본문 {result.chars:,}자 중 앞 {cap:,}자만 실었습니다."
            ] if truncated else None,
        ),
    )


@mcp.tool()
@safe_tool
@track_metrics("get_disclosure")
async def get_disclosure(code: str) -> str:
    """공시목록 — 종목의 최근 공시 (DART 전자공시) 목록.

    "공시", "IR", "실적 발표", "공정공시" 같은 질문에 사용합니다.

    Args:
        code: 종목코드 6자리 (예: "005930")
    """
    import re
    if not re.match(r"^[A-Za-z0-9]{6}$", code):
        return f"⚠️ 종목코드 형식이 올바르지 않습니다: {code}"

    items = await naver_get_disclosure_list(code)

    if not items:
        return f"종목코드 {code}의 공시 내역이 없습니다."

    lines = [f"## 공시 목록 ({code}, 최근 {len(items)}건)", ""]
    lines.append("날짜 | 제목 | 출처")
    lines.append("---|---|---")

    for item in items:
        lines.append(f"{item['date']} | {item['title']} | {item['source']}")

    lines.append("")
    lines.append("※ 공시 본문은 DART(dart.fss.or.kr)에서 확인 가능합니다.")
    # 네이버는 이 목록의 전체 건수를 알려주지 않는다. 20건이 왔다고 20건이 전부가
    # 아니고, "최근 공시에 아무것도 없다"를 여기서 결론지을 수 없다.
    lines.append(
        "※ 위 목록은 네이버가 보여주는 **최근 표본**이며 전체 공시가 아니라 일부입니다. "
        "기간·유형으로 빠짐없이 보려면 DartLens `list_disclosures` 를 쓰세요."
    )
    lines.append("※ 제목은 원문에서 잘려 있을 수 있습니다. 판단 전에 본문을 확인하세요.")

    return _append_result_meta(
        "\n".join(lines),
        _kr_meta(
            kind="filing",
            code=code,
            data_completeness=rmeta.PARTIAL,
            coverage={
                "returned_count": len(items),
                "total_count": None,      # 원천이 전체 건수를 주지 않는다
                "truncated": None,        # 그래서 잘렸는지조차 확정할 수 없다
                "coverage_complete": False,
                "reason": "source_limit",
            },
            extra={"title_may_be_truncated": True},
        ),
    )


# --- US Phase 3: 탐색/시장/재무제표/ETF/멀티 ---

@mcp.tool()
@safe_us_tool
@track_metrics("get_us_search")
async def get_us_search(query: str) -> str:
    """US stock search — 미국 주식 종목명/티커 검색 (US stock search, ticker lookup).
    "Apple 티커", "Tesla symbol", "반도체 ETF", "Nvidia 찾아줘" 같은 질문에 사용합니다.

    ⚠️ 사용자가 "애플"·"테슬라" 같은 회사명만 주면 **이 도구를 먼저 호출**하세요.
    결과가 여러 개면 사용자에게 확인 요청. 추측 금지.

    Args:
        query: 검색어 (회사명·티커, 한/영 무관)
    """
    results = await us.search(query)
    if not results:
        return f"'{query}'에 대한 미국 주식 검색 결과가 없습니다."

    lines = [f"**'{query}'** 검색 결과 ({len(results)}건):", ""]
    for r in results:
        sector = r.get("sector") or ""
        sector_str = f" · {sector}" if sector else ""
        lines.append(f"- **{r['name']}** ({r['symbol']}) [{r.get('exchange', '-')}, {r.get('type', '-')}]{sector_str}")
    return "\n".join(lines)


@mcp.tool()
@safe_us_tool
@track_metrics("get_us_market")
async def get_us_market() -> str:
    """US market indices — 미국 시장 스냅샷 (주요 지수·VIX·Gold).

    "미국 시장 어때", "VIX 얼마", "금값" 같은 질문에 사용합니다.

    ⚠️ **반환 항목은 소스가 주는 대로이며 고정이 아닙니다.** 실측(v0.8.2,
    2026-08-27): S&P 500 / Dow Jones / NASDAQ / Russell 2000 / VIX / Gold.
    예전에는 선물 4종만 오던 시기도 있었습니다. **반환된 표에 있는 항목만**
    답변에 쓰고, 없는 지수를 학습지식으로 채우지 마세요. 이 값은 현재
    스냅샷입니다. 과거 기간 수익률이 필요하면 `get_us_chart`로 조회하세요.
    """
    data = await us.get_market_summary()
    indices = data.get("indices", [])
    if not indices:
        return "지수 데이터를 가져올 수 없습니다."

    lines = ["**미국 주요 지수**", ""]
    lines.append("지수 | 심볼 | 현재 | 변동 | 변동률 | 상태")
    lines.append("---|---|---|---|---|---")
    for i in indices:
        price = i.get("price")
        change = i.get("change")
        chp = i.get("change_percent")
        price_s = f"{price:,.2f}" if isinstance(price, (int, float)) else "-"
        ch_s = f"{change:+,.2f}" if isinstance(change, (int, float)) else "-"
        chp_s = f"{chp:+.2f}%" if isinstance(chp, (int, float)) else "-"
        lines.append(f"{i['label']} | {i.get('symbol', '-')} | {price_s} | {ch_s} | {chp_s} | {i.get('market_state', '-')}")
    return _append_result_meta("\n".join(lines), _us_meta(kind="snapshot"))


@mcp.tool()
@safe_us_tool
@track_metrics("get_us_screener")
async def get_us_screener(preset: str = "day_gainers", count: int = 20,
                          common_stock_only: bool = False) -> str:
    """US stock screener — 미국 주식 프리셋 스크리너 (US predefined screener).
    "오늘 급등주", "top gainers", "가장 많이 거래된 종목", "저평가 성장주" 같은 질문에 사용합니다.

    사용 가능 preset: day_gainers, day_losers, most_actives, most_shorted_stocks,
    aggressive_small_caps, growth_technology_stocks, undervalued_growth_stocks,
    undervalued_large_caps, small_cap_gainers, conservative_foreign_funds

    각 행에 증권 유형(common_stock/warrant/right/unit/adr/etf/unknown)이 붙습니다.
    small_cap_gainers 같은 프리셋에는 GRABW(워런트)·KLXER(라이츠) 같은 파생
    식별자가 섞여 나옵니다 - 보통주 후보만 원하면 common_stock_only=True.
    유형을 확인할 수 없는 항목은 unknown 으로 남고 보통주로 치지 않습니다.

    Args:
        preset: 스크리너 ID (기본 day_gainers)
        count: 반환 종목 수 (기본 20)
        common_stock_only: True 면 보통주만 남깁니다 (unknown 도 제외)
    """
    data = await us.screen(preset, count=count)
    if data.get("error"):
        return f"⚠️ {data['error']}"

    quotes = data.get("quotes", [])
    if not quotes:
        return f"'{preset}' 결과 없음."

    # 행마다 증권 유형을 판정한다(SL-12). 필터를 켜면 보통주만 남기되,
    # 무엇을 몇 개 걸렀는지는 본문·메타에 남긴다.
    from collections import Counter as _Counter
    for q in quotes:
        q["_sec_type"] = _classify_us_security(q)
    type_counts = dict(_Counter(q["_sec_type"] for q in quotes))
    excluded_counts: dict = {}
    if common_stock_only:
        kept = [q for q in quotes if q["_sec_type"] == "common_stock"]
        excluded_counts = dict(_Counter(
            q["_sec_type"] for q in quotes if q["_sec_type"] != "common_stock"))
        quotes = kept

    lines = [
        f"**{data.get('title', preset)}** ({len(quotes)} / 전체 {data.get('total', '-')})",
    ]
    if data.get("description"):
        lines.append(f"_{data['description']}_")
    if excluded_counts:
        parts = ", ".join(f"{k} {v}개" for k, v in sorted(excluded_counts.items()))
        lines.append(f"⚠️ common_stock_only=True - 보통주가 아닌 {sum(excluded_counts.values())}개 제외: {parts}")
    lines.append("")
    lines.append("티커 | 이름 | 유형 | 거래소 | 현재가 | 변동% | 거래량 | 시총 | P/E")
    lines.append("---|---|---|---|---|---|---|---|---")
    for q in quotes:
        chp = q.get("change_percent")
        chp_s = f"{chp:+.2f}%" if isinstance(chp, (int, float)) else "-"
        vol = q.get("volume")
        vol_s = f"{vol/1_000_000:.2f}M" if isinstance(vol, (int, float)) else "-"
        mcap = q.get("market_cap")
        mcap_s = _fmt_num(mcap) if mcap else "-"
        pe = q.get("pe")
        pe_s = f"{pe:.2f}" if isinstance(pe, (int, float)) else "-"
        price = q.get("price")
        price_s = f"${price:,.2f}" if isinstance(price, (int, float)) else "-"
        name = (q.get("name") or "-")[:30]
        lines.append(
            f"{q['symbol']} | {name} | {q['_sec_type']} | "
            f"{q.get('exchange') or '-'} | {price_s} | {chp_s} | {vol_s} | {mcap_s} | {pe_s}")
    if not quotes:
        lines.append("(필터 후 남은 종목 없음)")
    lines.append("")
    lines.append("※ 유형 unknown 은 확인 불가라는 뜻이며 보통주로 치지 않습니다. "
                 "warrant/right/unit 은 이름 또는 나스닥 5글자 접미사(W/R/U) 규약으로 판정합니다.")
    return _append_result_meta(
        "\n".join(lines),
        _us_meta(
            kind="snapshot",
            extra={"security_types": type_counts,
                   **({"security_filter": {"common_stock_only": True,
                                           "excluded": excluded_counts}}
                      if common_stock_only else {})},
        ),
    )


@mcp.tool()
@safe_us_tool
@track_metrics("get_us_financial_statement")
async def get_us_financial_statement(
    ticker: str,
    statement_type: str = "income",
    period: str = "annual",
) -> str:
    """US financial statements — 미국 주식 재무제표 3종 (income/balance/cash_flow).
    "AAPL 손익계산서", "NVDA 현금흐름표", "Apple balance sheet quarterly" 같은 질문에 사용합니다.

    핵심 row만 추출 (Total Revenue, Net Income, Total Assets, Free Cash Flow 등).

    Args:
        ticker: US 티커
        statement_type: "income" / "balance" / "cash_flow" (기본 income)
        period: "annual" / "quarterly" (기본 annual)
    """
    data = await us.get_financial_statement(ticker, statement_type=statement_type, period=period)
    if data is None:
        return f"티커 '{ticker}'를 찾을 수 없습니다."
    if data.get("error"):
        return f"⚠️ {data['error']}"

    rows = data.get("rows", [])
    if not rows:
        return f"**{data['ticker']}** {statement_type} ({period}) — 데이터 없음."

    name_map = {"income": "손익계산서", "balance": "재무상태표", "cash_flow": "현금흐름표"}
    # 헤더에 통화를 적어 놓고 금액엔 '$'를 붙이면 한 화면 안에서 서로 모순된다.
    stmt_ccy = _ccy_prefix(data.get("currency"))
    lines = [
        f"**{data['ticker']}** {name_map.get(statement_type, statement_type)} ({period})",
        f"통화: {data.get('currency', '-')}",
        "",
    ]

    # 기간 헤더 — 첫 row의 periods 사용
    periods = rows[0]["periods"]
    header = "항목 | " + " | ".join(periods)
    sep = "---|" + "---|" * len(periods)
    lines.append(header)
    lines.append(sep[:-1])

    for r in rows:
        vals = []
        for v in r["values"]:
            if v is None:
                vals.append("-")
            elif isinstance(v, (int, float)):
                vals.append(_fmt_num(v, digits=1, currency=stmt_ccy))
            else:
                vals.append(str(v))
        lines.append(f"{r['item']} | " + " | ".join(vals))

    # 최신 열이 언제 것인지가 메타에 있어야 한다(SL-03). 본문 표에는 기간이
    # 있지만 그 줄을 안 읽으면 그만이고, 배치 비교의 mixed 판정도 이 값에서 나온다.
    latest = max(periods) if periods else None
    if latest:
        completeness, coverage = rmeta.COMPLETE, None
        warns = None
    else:
        completeness = rmeta.PARTIAL
        coverage = {"truncated": False, "coverage_complete": False, "reason": "unknown"}
        warns = ["재무제표의 기간 열을 읽지 못했습니다. 이 값들이 언제 것인지 알 수 없습니다."]
    return _append_result_meta(
        "\n".join(lines),
        _us_meta(
            kind="filing",
            ticker=ticker,
            data_as_of=latest,
            data_completeness=completeness,
            coverage=coverage,
            extra={"period_coverage": {
                "cycle": period,
                "latest_period": latest,
                "periods_returned": len(periods),
                "currency": data.get("currency"),
                # yfinance 는 접수일을 주지 않는다. 이 날짜는 기간말이다.
                "basis": "period_end",
            }},
            warnings=warns,
        ),
    )


@mcp.tool()
@safe_us_tool
@track_metrics("get_us_sector")
async def get_us_sector(sector_key: str, top_n: int = 20) -> str:
    """US sector overview — 미국 섹터별 top 기업 + 시장 비중 (US sector top companies).
    "기술주 섹터", "healthcare top companies", "technology 대장주", "섹터 비중" 같은 질문에 사용합니다.

    Args:
        sector_key: technology, healthcare, financial-services, consumer-cyclical,
                    consumer-defensive, communication-services, industrials,
                    energy, basic-materials, utilities, real-estate
        top_n: top 기업 수 (기본 20)
    """
    data = await us.get_sector(sector_key, top_n=top_n)
    if data is None or data.get("error"):
        return f"⚠️ {data.get('error') if data else '섹터 조회 실패'}"

    lines = [
        f"**섹터: {sector_key}**",
        "",
        "## 📊 Overview",
    ]
    if data.get("companies_count"):
        lines.append(f"- 종목 수: {int(data['companies_count']):,}")
    if data.get("market_cap"):
        lines.append(f"- 시가총액: {_fmt_num(data['market_cap'])}")
    if data.get("market_weight"):
        lines.append(f"- 시장 비중: {data['market_weight']*100:.2f}%")
    if data.get("industries_count"):
        lines.append(f"- 산업 수: {int(data['industries_count'])}")
    if data.get("employee_count"):
        lines.append(f"- 총 고용: {int(data['employee_count']):,}명")

    desc = data.get("description")
    if desc:
        lines.append("")
        lines.append(f"_{desc[:400]}_")

    companies = data.get("top_companies", [])
    if companies:
        lines.append("")
        lines.append(f"## 🏢 Top {len(companies)} 기업")
        lines.append("순위 | 종목 | 투자의견 | 시장 비중")
        lines.append("---|---|---|---")
        for idx, c in enumerate(companies, 1):
            sym = c.get("symbol") or c.get("index") or "-"
            name = (c.get("name") or "-")[:35]
            rating = c.get("rating") or "-"
            mw = c.get("market_weight") or c.get("market weight")
            mw_s = f"{mw*100:.2f}%" if isinstance(mw, (int, float)) else "-"
            lines.append(f"{idx} | **{sym}** {name} | {rating} | {mw_s}")

    return _append_result_meta("\n".join(lines), _us_meta(kind="snapshot"))


@mcp.tool()
@safe_us_tool
@track_metrics("get_us_etf_info")
async def get_us_etf_info(ticker: str) -> str:
    """US ETF info — ETF 전용 상세 (top holdings, 섹터 비중, 자산 배분 · US ETF details).
    "SPY 구성종목", "QQQ holdings", "VOO 섹터", "ETF 보수" 같은 질문에 사용합니다.

    Args:
        ticker: ETF 티커 (예: "SPY", "QQQ", "VOO", "SCHD")
    """
    data = await us.get_etf_info(ticker)
    if data is None:
        return f"티커 '{ticker}'를 찾을 수 없습니다."
    if data.get("error"):
        return f"⚠️ {data['error']}"

    lines = [
        f"**{data.get('name', data['ticker'])}** ({data['ticker']})",
        f"카테고리: {data.get('category', '-')} · {data.get('family', '-')}",
    ]
    if data.get("expense_ratio") is not None:
        # yfinance netExpenseRatio는 이미 % 단위 (0.0945 = 0.0945%)
        lines.append(f"보수율: {data['expense_ratio']:.3f}%")
    if data.get("total_assets"):
        lines.append(f"운용자산(AUM): {_fmt_num(data['total_assets'])}")
    if data.get("ytd_return") is not None:
        # ytdReturn도 이미 % 단위
        lines.append(f"YTD 수익률: {data['ytd_return']:+.2f}%")

    ac = data.get("asset_classes") or {}
    if ac:
        lines.append("")
        lines.append("## 📊 자산 배분")
        for k, v in ac.items():
            if isinstance(v, (int, float)) and v > 0:
                lines.append(f"- {k}: {v*100:.2f}%")

    sw = data.get("sector_weightings") or {}
    if sw:
        lines.append("")
        lines.append("## 🏭 섹터 비중")
        for k, v in sorted(sw.items(), key=lambda x: -(x[1] if isinstance(x[1], (int, float)) else 0)):
            if isinstance(v, (int, float)) and v > 0:
                lines.append(f"- {k.replace('_', ' ').title()}: {v*100:.2f}%")

    th = data.get("top_holdings") or []
    if th:
        lines.append("")
        lines.append(f"## 🏆 Top Holdings ({len(th)}개)")
        lines.append("")
        lines.append(
            f"> ℹ️ **Yahoo Finance는 상위 {len(th)}개 보유 종목만 공개합니다.** "
            "ETF 전체 구성종목(예: SPY 500종목)은 발행사 공식 페이지에서 확인 가능합니다 "
            "(예: SPY → ssga.com, QQQ → invesco.com, SCHD → schwabassetmanagement.com)."
        )
        lines.append("")
        lines.append("종목 | 비중")
        lines.append("---|---")
        for h in th:
            sym = h.get("symbol") or h.get("index") or "-"
            name = h.get("name") or h.get("holding_name") or h.get("holding") or ""
            weight = h.get("holding_percent") or h.get("weight") or h.get("pct_held")
            w_s = f"{weight*100:.2f}%" if isinstance(weight, (int, float)) else "-"
            lines.append(f"**{sym}** {name} | {w_s}")

    desc = data.get("description")
    if desc:
        lines.append("")
        lines.append(f"_{desc[:400]}_")

    return _append_result_meta("\n".join(lines), _us_meta(kind="snapshot", ticker=ticker))


@mcp.tool()
@safe_us_tool
@track_metrics("get_us_multi_price")
async def get_us_multi_price(tickers: list[str]) -> str:
    """US multi-ticker prices — 여러 미국 주식 일괄 가격 조회 (US multi-ticker snapshot).
    "AAPL MSFT NVDA 동시", "big tech 가격", "내 포트폴리오 현재가" 같은 질문에 사용합니다.

    Args:
        tickers: 티커 리스트 (예: ["AAPL", "MSFT", "NVDA"]). 최대 20개 권장.
    """
    if not tickers:
        return "티커 리스트가 비어있습니다."
    if len(tickers) > 30:
        tickers = tickers[:30]

    rows = await us.get_multi_prices(tickers)
    if not rows:
        return "데이터를 가져올 수 없습니다."

    lines = [f"**멀티 티커 스냅샷** ({len(rows)}개)", ""]
    lines.append("티커 | 이름 | 현재가 | 변동 | 변동% | 거래량 | 시총")
    lines.append("---|---|---|---|---|---|---")
    for r in rows:
        if r.get("error"):
            lines.append(f"{r['ticker']} | ⚠️ {r['error']} | - | - | - | - | -")
            continue
        price = r.get("price")
        price_s = f"${price:,.2f}" if isinstance(price, (int, float)) else "-"
        ch = r.get("change")
        ch_s = f"{ch:+,.2f}" if isinstance(ch, (int, float)) else "-"
        chp = r.get("change_percent")
        chp_s = f"{chp:+.2f}%" if isinstance(chp, (int, float)) else "-"
        vol = r.get("volume")
        vol_s = f"{vol/1_000_000:.1f}M" if isinstance(vol, (int, float)) else "-"
        mcap = r.get("market_cap")
        mcap_s = _fmt_num(mcap) if mcap else "-"
        name = (r.get("name") or "-")[:25]
        lines.append(f"**{r['ticker']}** | {name} | {price_s} | {ch_s} | {chp_s} | {vol_s} | {mcap_s}")

    # 이 함수의 인자는 tickers(복수)다 — 없는 단수 ticker 를 넘기다 NameError 로
    # 죽어서, 데이터를 다 받아놓고 마지막 줄에서 도구 전체가 실패했다.
    # 여러 종목 스냅샷이라 entity 를 하나로 특정할 수 없으므로 ticker 는 넘기지 않는다.
    return _append_result_meta("\n".join(lines), _us_meta(kind="snapshot"))


@mcp.tool()
@safe_us_tool
@track_metrics("export_us_to_excel")
async def export_us_to_excel(
    ticker: str,
    period: str = "10y",
    interval: str = "1d",
    filename: str = "",
) -> str:
    """US Excel export — 미국 주식 장기 데이터를 Excel 파일로 저장 (토큰 소비 없음).
    "AAPL 10년치 CSV 저장", "TSLA 5년 일봉 엑셀" 같은 질문에 사용. get_us_chart는 500행 상한이라
    장기 데이터는 잘림 — 백테스트·CSV·다른 AI 업로드용이면 이 도구로(행 수 무제한).

    Args:
        ticker: US 티커 (예: "AAPL", "SPY", "BRK.B")
        period: "1d","5d","1mo","3mo","6mo","1y","2y","5y","10y","ytd","max" (기본 10y)
        interval: "1d","1wk","1mo" (기본 1d)
        filename: 파일명 (비우면 자동)
    """
    rows = await us.get_history(ticker, period=period, interval=interval, prepost=False)
    if not rows:
        return f"티커 '{ticker}'의 데이터를 가져올 수 없습니다."

    df = pd.DataFrame(rows)
    prefix = f"us_chart_{ticker.upper().replace('.', '-')}_{period}_{interval}"
    fname = filename or generate_filename(prefix)
    if not fname.endswith(".xlsx"):
        fname += ".xlsx"
    file_path = get_snapshot_dir() / fname
    saved = save_dataframe_to_excel(
        df,
        file_path,
        sheet_name="OHLCV",
        source="Yahoo Finance (yfinance)",
        metadata={
            "ticker": ticker.upper(),
            "period": period,
            "interval": interval,
        },
    )

    return (
        f"✓ US Excel 저장 완료\n"
        f"티커: {ticker.upper()} · 기간: {period} · 간격: {interval}\n"
        f"경로: {saved}\n"
        f"행 수: {len(df)}\n"
        f"컬럼: {', '.join(df.columns)}\n\n"
        f"💡 이 파일을 엑셀·Gemini·ChatGPT 등에서 바로 분석 가능. Claude 토큰 소비 없음."
    )


def main():
    mcp.run()


if __name__ == "__main__":
    main()
