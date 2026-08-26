"""Excel 파일 저장/조회 유틸리티.

사용자의 Downloads/kstock/ 폴더에 xlsx 파일을 저장하여
로컬 캐시 역할을 한다. DB 설치 없이도 반복 쿼리를 즉시 처리할 수 있다.
"""

import os
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd


def get_snapshot_dir() -> Path:
    """OS별 기본 저장 폴더. ~/Downloads/kstock/"""
    if sys.platform == "win32":
        base = Path(os.environ.get("USERPROFILE", "")) / "Downloads"
    elif sys.platform == "darwin":
        base = Path.home() / "Downloads"
    else:
        base = Path.home() / "Downloads"

    folder = base / "kstock"
    folder.mkdir(parents=True, exist_ok=True)
    return folder


def generate_filename(prefix: str, ext: str = "xlsx") -> str:
    """타임스탬프 붙은 파일명 생성."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{ts}.{ext}"


# 열 이름 → 한글 표시명. 파일을 여는 사람은 drawdown_pct 가 뭔지 모른다.
# 내부 필터(query_excel)는 영문 키를 그대로 받으므로 별칭으로 이어 준다.
COLUMN_LABELS_KO: dict[str, str] = {
    "code": "종목코드", "name": "종목명",
    "current_price": "현재가", "price": "현재가", "volume": "거래량",
    "high": "기간최고가", "high_date": "최고가일",
    "low": "기간최저가", "low_date": "최저가일",
    "drawdown_pct": "고점대비낙폭(%)", "recovery_pct": "저점대비회복(%)",
    "period_return_pct": "기간수익률(%)", "avg_volume": "평균거래량",
    "bars_count": "집계봉수",
    "per": "PER(배)", "pbr": "PBR(배)", "eps": "EPS(원)", "bps": "BPS(원)",
    "roe": "ROE(%)", "per_basis": "PER기준", "fin_period": "재무기준기간",
    "sector": "업종", "market_cap": "시가총액",
    "ma_phase": "이평배열", "rsi": "RSI", "vol_ratio": "거래량배수",
    "date": "날짜", "open": "시가", "close": "종가",
    "inst_net": "기관순매매", "foreign_net": "외국인순매매",
}

# 숫자 서식 — 정수형 금액/수량과 소수형 비율을 나눈다.
_INT_LIKE = ("현재가", "거래량", "기간최고가", "기간최저가", "평균거래량",
             "시가총액", "EPS(원)", "BPS(원)", "시가", "종가",
             "기관순매매", "외국인순매매", "집계봉수")
_DEC_LIKE = ("PER(배)", "PBR(배)", "배당수익률")
# 등락·수익률 계열은 부호가 곧 의미다. 국내 관례대로 오름 빨강 / 내림 파랑으로
# 칠한다(숫자 서식에 색을 넣으면 조건부 서식 없이 그대로 적용된다).
_SIGNED_PCT = ("고점대비낙폭(%)", "저점대비회복(%)", "기간수익률(%)", "ROE(%)",
               "기관순매매", "외국인순매매")
_SIGNED_FMT = '[Red]+#,##0.00;[Blue]-#,##0.00;0.00'
_SIGNED_INT_FMT = '[Red]+#,##0;[Blue]-#,##0;0'


# 한글 표시명 → 내부 키. 파일은 한글로 저장하지만 필터·정렬은 영문 키로 받는다
# (기존 사용자가 쓰던 per_max 같은 조건이 그대로 동작해야 한다).
LABEL_TO_KEY: dict[str, str] = {}
for _k, _v in COLUMN_LABELS_KO.items():
    LABEL_TO_KEY.setdefault(_v, _k)


# 열 이름을 미리 다 알 수 없다 — AI 가 정리한 표는 "등락률", "기관2일" 처럼
# 그때그때 다른 이름을 쓴다. 고정 목록으로만 판정하면 서식이 통째로 빠지므로,
# 이름의 힌트와 값의 성격을 함께 보고 정한다.
_PCT_HINTS = ("%", "률", "낙폭", "수익", "등락", "증감", "비율", "rsi", "RSI")
_MONEY_HINTS = ("가", "금액", "시총", "시가총액", "순매매", "거래량", "주식", "잔량", "봉")


def _guess_number_format(col_name, values) -> str | None:
    """열 하나에 맞는 숫자 서식. 판단이 안 서면 None(서식 없음)."""
    name = str(col_name)
    if name in _SIGNED_PCT:
        return _SIGNED_FMT
    if name in ("기관순매매", "외국인순매매"):
        return _SIGNED_INT_FMT
    if name in _INT_LIKE:
        return "#,##0"
    if name in _DEC_LIKE:
        return "#,##0.00"

    nums = [v for v in values
            if isinstance(v, (int, float)) and not isinstance(v, bool)]
    if not nums:
        return None
    has_neg = any(v < 0 for v in nums)
    has_frac = any(float(v) != int(v) for v in nums)
    low = name.lower()
    pct_like = any(h.lower() in low for h in _PCT_HINTS)
    money_like = any(h in name for h in _MONEY_HINTS)

    if pct_like:
        return _SIGNED_FMT if has_neg else "#,##0.00"
    if has_neg and money_like:
        return _SIGNED_INT_FMT
    if has_frac:
        return "#,##0.00"
    if money_like or max(abs(v) for v in nums) >= 1000:
        return "#,##0"
    return None


def _chartable_columns(df) -> list[str]:
    """종목 간 비교가 되는 숫자 열. 가격·거래량처럼 자릿수가 제각각인 값은 뺀다."""
    out = []
    for c in df.columns:
        name = str(c)
        vals = [v for v in df[c]
                if isinstance(v, (int, float)) and not isinstance(v, bool)]
        if len(vals) < 2:
            continue
        low = name.lower()
        if not any(h.lower() in low for h in _PCT_HINTS + ("per", "pbr", "roe", "배수")):
            continue
        if max(abs(v) for v in vals) > 100000:   # 금액 단위는 비교 차트에 안 맞는다
            continue
        out.append(name)
    return out


def _polish_sheet(ws, df) -> None:
    """열자마자 쓸 수 있는 상태로 만든다 — 필터·헤더고정·열너비·숫자서식.

    지금까지는 df.to_excel() 만 불러서, 파일을 열면 열 이름이 잘리고 숫자에
    자릿점이 없으며 정렬하려면 사용자가 직접 필터를 걸어야 했다.
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    if ws.max_row < 1:
        return
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 첫 행 고정 + 자동 필터
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for idx, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(idx)
        # 열 너비: 헤더와 값 중 긴 쪽에 맞추되 한글 폭을 감안해 넉넉히
        try:
            longest = max((len(str(v)) for v in df[col_name].head(200)), default=0)
        except Exception:
            longest = 0
        width = min(max(len(str(col_name)) * 1.6, longest * 1.3, 8), 30)
        ws.column_dimensions[letter].width = width

        try:
            fmt = _guess_number_format(col_name, list(df[col_name]))
        except Exception:
            fmt = None
        if fmt:
            for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
                for cell in row:
                    cell.number_format = fmt


# 그래프로 그렸을 때 의미가 있는 열 — 종목 간 비교가 되는 값만 고른다.
# 가격·거래량처럼 종목마다 단위가 크게 다른 값은 한 차트에 섞으면 못 읽는다.
_CHART_CANDIDATES = (
    ("기간수익률(%)", "기간 수익률 비교 (%)"),
    ("고점대비낙폭(%)", "고점 대비 낙폭 (%)"),
    ("PER(배)", "PER 비교 (배)"),
    ("ROE(%)", "ROE 비교 (%)"),
    ("RSI", "RSI 비교"),
)
_FLOW_PAIR = ("기관순매매", "외국인순매매")


def _add_charts(wb, ws_data, df, sheet_name: str) -> None:
    """비교용 막대그래프를 '차트' 시트에 넣는다.

    숫자만 있는 표는 종목이 몇 개만 넘어가도 눈으로 비교가 안 된다. 파일을 여는
    사람이 바로 알아볼 수 있도록, 종목 간 비교가 되는 열만 골라 그린다.
    종목이 너무 많으면(20개 초과) 축이 뭉개져 오히려 안 보이므로 그리지 않는다.
    """
    from openpyxl.chart import BarChart, Reference

    n = len(df)
    if n < 2 or n > 20:
        return
    cols = list(df.columns)
    if "종목명" not in cols:
        return
    name_idx = cols.index("종목명") + 1

    picked: list[tuple[list[int], str]] = []

    # 기간수익률·낙폭은 늘 따라오는 기본 항목이다. 사용자가 일부러 고른 항목
    # (수급·재무·지표)이 있으면 그쪽을 먼저 그려야 원하는 그림이 나온다.
    if all(c in cols for c in _FLOW_PAIR):
        picked.append(([cols.index(c) + 1 for c in _FLOW_PAIR],
                       "기관·외국인 순매매 (20일 누적, 주)"))
    for label, title in (("PER(배)", "PER 비교 (배)"),
                         ("ROE(%)", "ROE 비교 (%)"),
                         ("RSI", "RSI 비교")):
        if len(picked) >= 2:
            break
        if label in cols:
            picked.append(([cols.index(label) + 1], title))
    for label, title in _CHART_CANDIDATES:
        if len(picked) >= 2:
            break
        if label in cols and not any(label in ttl for _, ttl in picked):
            picked.append(([cols.index(label) + 1], title))
    # 고정 목록에 없는 이름(AI 가 정리한 표)도 비교 가능한 열이면 그린다.
    if len(picked) < 2:
        used = {c for cs, _ in picked for c in cs}
        for name in _chartable_columns(df):
            if len(picked) >= 2:
                break
            idx = cols.index(name) + 1
            if idx in used:
                continue
            picked.append(([idx], f"{name} 비교"))
    if not picked:
        return

    ws = wb.create_sheet("차트")
    ws["A1"] = "아래 그래프는 왼쪽 데이터 시트를 그대로 읽습니다 — 값이 바뀌면 함께 바뀝니다."
    anchor_row = 3
    for value_cols, title in picked:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = title
        chart.y_axis.title = None
        chart.x_axis.title = None
        chart.width = 24
        chart.height = 11
        for ci in value_cols:
            ref = Reference(ws_data, min_col=ci, min_row=1, max_row=n + 1)
            chart.add_data(ref, titles_from_data=True)
        cats = Reference(ws_data, min_col=name_idx, min_row=2, max_row=n + 1)
        chart.set_categories(cats)
        ws.add_chart(chart, f"A{anchor_row}")
        anchor_row += 23


def add_detail_sheet(wb, sheet_title: str, summary: list[tuple], data: dict) -> None:
    """종목 한 개의 '한눈에' 시트.

    주가 캔들은 넣지 않는다 — HTS·증권사 화면이 훨씬 낫고 엑셀로 옮기면 선 하나가
    된다. 대신 여러 화면을 돌아다녀야 보이던 것(밸류에이션·증권가 전망·수급·공시)을
    한 장에 모으고, 숫자만 늘어놓지 않도록 카드로 끊어 읽게 만든다.
    """
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.formatting.rule import DataBarRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    ws = wb.create_sheet(sheet_title[:31])
    ws.sheet_view.showGridLines = False

    NAVY, SKY, GREY = "1F4E79", "EAF1F8", "888888"
    RED, BLUE = "C00000", "1F6FB5"
    thin = Side(style="thin", color="BFD4E8")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, w in (("A", 16), ("B", 15), ("C", 14), ("D", 15), ("E", 14), ("F", 12)):
        ws.column_dimensions[col].width = w
    for col in "HIJKLMNO":
        ws.column_dimensions[col].width = 12
    for col in ("P", "Q", "R"):
        ws.column_dimensions[col].width = 13
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[4].height = 30

    price = data.get("price") or {}
    val = data.get("valuation") or {}
    con = data.get("consensus") or {}
    flow = data.get("flow") or {}

    def money(v):
        return format(v, ",.0f") if isinstance(v, (int, float)) else "-"

    def pct(v, digits=2):
        return format(v, "+." + str(digits) + "f") + "%" if isinstance(v, (int, float)) else "-"

    ws["A1"] = sheet_title
    ws["A1"].font = Font(bold=True, size=16, color=NAVY)
    note = ""
    for k, v in summary:
        if str(k) in ("읽는법", "판정", "코멘트") and v:
            note = str(v)
            break
    if note:
        ws["A2"] = note
        ws["A2"].font = Font(size=10, color=GREY)
        ws.merge_cells("A2:F2")

    chg = price.get("등락률")
    if isinstance(chg, (int, float)):
        chg_color = RED if chg > 0 else (BLUE if chg < 0 else NAVY)
    else:
        chg_color = NAVY
    per_txt = format(val["PER"], ",.2f") if val.get("PER") else "-"
    tiles = [
        ("현재가", money(price.get("현재가")), NAVY),
        ("등락률", pct(chg), chg_color),
        ("PER (TTM)", per_txt, NAVY),
        ("52주 고점 대비", pct(price.get("고점대비"), 1), BLUE),
    ]
    for i, (label, value, color) in enumerate(tiles):
        col = 1 + i
        lc = ws.cell(row=3, column=col, value=label)
        lc.font = Font(size=9, color=GREY)
        lc.alignment = Alignment(horizontal="center")
        lc.fill = PatternFill("solid", fgColor=SKY)
        lc.border = box
        vc = ws.cell(row=4, column=col, value=value)
        vc.font = Font(bold=True, size=15, color=color)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.fill = PatternFill("solid", fgColor=SKY)
        vc.border = box

    def section(row, title):
        c = ws.cell(row=row, column=1, value="  " + title)
        c.font = Font(bold=True, size=11, color="FFFFFF")
        for i in range(1, 7):
            ws.cell(row=row, column=i).fill = PatternFill("solid", fgColor=NAVY)
        return row + 1

    def line(row, pairs):
        for i, (k, v) in enumerate(pairs):
            kc = ws.cell(row=row, column=1 + i * 2, value=str(k))
            kc.font = Font(bold=True, size=10, color="333333")
            kc.border = box
            vc = ws.cell(row=row, column=2 + i * 2, value=v)
            vc.border = box
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                vc.number_format = "#,##0.00" if float(v) != int(v) else "#,##0"
        return row + 1

    def wide(row, key, value, color=None):
        kc = ws.cell(row=row, column=1, value=key)
        kc.font = Font(bold=True, size=10, color="333333")
        kc.border = box
        vc = ws.cell(row=row, column=2, value=value)
        vc.border = box
        vc.font = Font(bold=bool(color), size=10, color=color or "000000")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        return row + 1

    r = 6
    r = section(r, "가격 위치")
    r = line(r, [("저점 대비(%)", price.get("저점대비")), ("이평배열", price.get("이평배열"))])
    r = line(r, [("거래량 배수", price.get("거래량배수")), ("52주 저점", price.get("저점가"))])
    r += 1

    if val:
        r = section(r, "밸류에이션   " + str(val.get("기준", "")))
        r = line(r, [("PER(배)", val.get("PER")), ("업종 중앙값", val.get("업종PER"))])
        r = line(r, [("PBR(배)", val.get("PBR")), ("ROE(%)", val.get("ROE"))])
        if val.get("위치"):
            tone = RED if "할증" in str(val["위치"]) else BLUE
            r = wide(r, "업종 대비", val["위치"], tone)
        r += 1

    if con:
        r = section(r, "증권가 전망")
        r = line(r, [("목표주가", con.get("목표주가")), ("현재가 대비(%)", con.get("괴리율"))])
        if con.get("의견"):
            r = wide(r, "투자의견", con["의견"])
        if con.get("서프라이즈"):
            r = wide(r, "실적 서프라이즈", con["서프라이즈"])
        r += 1

    if flow:
        r = section(r, "수급 (최근 20거래일 누적)")
        fr = r
        r = line(r, [("기관(주)", flow.get("기관")), ("외국인(주)", flow.get("외국인"))])
        for col in (2, 4):
            c = ws.cell(row=fr, column=col)
            if isinstance(c.value, (int, float)):
                c.number_format = _SIGNED_INT_FMT
                c.font = Font(bold=True, color=RED if c.value > 0 else BLUE)
        r += 1

    fin = data.get("stability") or {}
    if fin:
        r = section(r, "재무 안정성 · 배당")
        r = line(r, [("부채비율(%)", fin.get("부채비율")), ("당좌비율(%)", fin.get("당좌비율"))])
        r = line(r, [("주당배당금(원)", fin.get("주당배당금")), ("시가배당률(%)", fin.get("시가배당률"))])
        r += 1

    fc_ = data.get("forecast") or {}
    if fc_:
        r = section(r, "컨센서스 전망 (증권사 추정)")
        r = line(r, [("올해 매출(억)", fc_.get("매출액")), ("올해 영업이익(억)", fc_.get("영업이익"))])
        if fc_.get("기간"):
            r = wide(r, "추정 기준", str(fc_["기간"]) + " — 확정 실적이 아닙니다")
        r += 1

    peers = data.get("peers") or []
    if peers:
        r = section(r, "같은 업종 주요 종목")
        hdr = ("종목", "등락률(%)", "현재가", "")
        for i, h in enumerate(hdr):
            hc = ws.cell(row=r, column=1 + i, value=h)
            hc.font = Font(bold=True, size=9, color=GREY)
            hc.fill = PatternFill("solid", fgColor=SKY)
            hc.border = box
        r += 1
        for pr in peers[:6]:
            nc = ws.cell(row=r, column=1, value=str(pr.get("종목", ""))[:12])
            nc.border = box
            nc.font = Font(bold=bool(pr.get("본인")), size=10)
            for i, key in enumerate(("등락률", "현재가")):
                vc = ws.cell(row=r, column=2 + i, value=pr.get(key))
                vc.border = box
                if isinstance(pr.get(key), (int, float)):
                    vc.number_format = _SIGNED_FMT if key == "등락률" else "#,##0"
            r += 1
        r += 1

    news = data.get("news") or []
    if news:
        r = section(r, "최근 공시·리포트")
        for item in news[:6]:
            day = str(item.get("date", "")).replace(".", "-")
            if len(day) == 8 and day[2] == "-":
                day = "20" + day
            dc = ws.cell(row=r, column=1, value=day[:10])
            dc.font = Font(size=9, color=GREY)
            dc.border = box
            title = str(item.get("title", ""))
            if title.startswith(sheet_title):
                title = title[len(sheet_title):]
                for pre in ("(주)", "\u321c", ")"):
                    if title.startswith(pre):
                        title = title[len(pre):]
                        break
                title = title.lstrip()
            c = ws.cell(row=r, column=2, value=title[:56])
            c.border = box
            c.alignment = Alignment(horizontal="left")
            # 원문 링크가 있으면 셀에서 바로 열리게 한다 — 제목만 보고는
            # 무슨 내용인지 알 수 없어 결국 다시 찾아야 한다.
            url = item.get("url")
            if url:
                c.hyperlink = url
                c.font = Font(size=10, color="0563C1", underline="single")
            else:
                c.font = Font(size=10)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            r += 1
            brief = str(item.get("brief") or "").strip()
            if brief:
                bc = ws.cell(row=r, column=2, value="  " + brief[:90])
                bc.font = Font(size=9, color=GREY)
                bc.alignment = Alignment(horizontal="left")
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
                r += 1

    # 다른 렌즈(TelegramLens·DartLens 등)에서 가져온 것들. 형식을 고정하지 않고
    # 표 형태로 받아 그대로 카드로 만든다 — 무엇을 넣을지는 부르는 쪽이 정한다.
    for card in (data.get("extras") or []):
        rows_ = card.get("rows") or []
        if not rows_:
            continue
        r += 1
        r = section(r, str(card.get("title", "참고")))
        cols_ = card.get("cols") or []
        if cols_:
            for i, cname in enumerate(cols_[:6]):
                hc = ws.cell(row=r, column=1 + i, value=str(cname))
                hc.font = Font(bold=True, size=9, color=GREY)
                hc.fill = PatternFill("solid", fgColor=SKY)
                hc.border = box
            r += 1
        for line_ in rows_[:8]:
            # 행은 리스트로도, {"cells": [...], "url": ...} 로도 받는다.
            # 텔레그램 원문처럼 링크가 있으면 마지막 칸에서 바로 열리게 한다.
            link = None
            if isinstance(line_, dict):
                cells = line_.get("cells") or []
                link = line_.get("url")
            elif isinstance(line_, (list, tuple)):
                cells = list(line_)
            else:
                cells = [line_]
            span = len(cols_) if cols_ else min(len(cells), 6)
            last_i = min(len(cells), 6) - 1
            for i, v in enumerate(cells[:6]):
                vc = ws.cell(row=r, column=1 + i, value=v)
                vc.border = box
                if link and i == last_i:
                    vc.hyperlink = link
                    vc.font = Font(size=10, color="0563C1", underline="single")
                else:
                    vc.font = Font(size=9 if i == 0 else 10,
                                   color=GREY if i == 0 else "000000")
                vc.alignment = Alignment(horizontal="left", vertical="center")
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    vc.number_format = "#,##0.00" if float(v) != int(v) else "#,##0"
            # 마지막 칸이 본문이면 남은 열까지 늘려 준다
            if span < 6:
                ws.merge_cells(start_row=r, start_column=span,
                               end_row=r, end_column=6)
            r += 1

    qs = data.get("quarters") or []
    if len(qs) >= 2:
        top = 3
        for i, cname in enumerate(("분기", "매출액(억)", "영업이익(억)")):
            c = ws.cell(row=top, column=16 + i, value=cname)
            c.font = Font(bold=True, size=10)
            c.fill = PatternFill("solid", fgColor=SKY)
            c.border = box
            c.alignment = Alignment(horizontal="center")
        for i, q in enumerate(qs):
            rr = top + 1 + i
            ws.cell(row=rr, column=16, value=q.get("기간")).border = box
            a = ws.cell(row=rr, column=17, value=q.get("매출액"))
            a.number_format = "#,##0"
            a.border = box
            b = ws.cell(row=rr, column=18, value=q.get("영업이익"))
            b.number_format = _SIGNED_INT_FMT
            b.border = box
        last = top + len(qs)
        ws.conditional_formatting.add(
            "Q" + str(top + 1) + ":Q" + str(last),
            DataBarRule(start_type="min", end_type="max", color="8EA9DB"))
        ch = BarChart()
        ch.type = "col"
        ch.title = "분기 매출·영업이익 (억원)"
        ch.height, ch.width = 7.2, 14.5
        ch.add_data(Reference(ws, min_col=17, max_col=18, min_row=top, max_row=last),
                    titles_from_data=True)
        ch.set_categories(Reference(ws, min_col=16, min_row=top + 1, max_row=last))
        ch.y_axis.majorGridlines = None
        for s_, color in zip(ch.series, ("4472C4", "ED7D31")):
            s_.graphicalProperties.solidFill = color
            s_.graphicalProperties.line.noFill = True
        ws.add_chart(ch, "H3")

    daily = data.get("flow_daily") or []
    if len(daily) >= 5:
        base = 28
        for i, cname in enumerate(("날짜", "기관", "외국인")):
            c = ws.cell(row=base, column=16 + i, value=cname)
            c.font = Font(bold=True, size=10)
            c.fill = PatternFill("solid", fgColor=SKY)
            c.border = box
            c.alignment = Alignment(horizontal="center")
        for i, d in enumerate(daily):
            rr = base + 1 + i
            ws.cell(row=rr, column=16, value=d.get("날짜")).border = box
            for j, key in enumerate(("기관", "외국인")):
                v = ws.cell(row=rr, column=17 + j, value=d.get(key))
                v.number_format = _SIGNED_INT_FMT
                v.border = box
        lastd = base + len(daily)
        fc = BarChart()
        fc.type = "col"
        fc.grouping = "clustered"
        fc.overlap = -10
        fc.gapWidth = 50
        fc.title = "일별 순매매 — 기관 vs 외국인 (주)"
        fc.height, fc.width = 7.2, 14.5
        fc.add_data(Reference(ws, min_col=17, max_col=18, min_row=base, max_row=lastd),
                    titles_from_data=True)
        fc.set_categories(Reference(ws, min_col=16, min_row=base + 1, max_row=lastd))
        fc.y_axis.majorGridlines = None
        fc.x_axis.tickLblSkip = max(1, len(daily) // 6)
        fc.x_axis.tickMarkSkip = max(1, len(daily) // 6)
        for s_, color in zip(fc.series, (RED, BLUE)):
            s_.graphicalProperties.solidFill = color
            s_.graphicalProperties.line.noFill = True
        ws.add_chart(fc, "H18")
        target_base = lastd + 2
    else:
        target_base = 28

    tg = data.get("target") or []
    if len(tg) >= 2:
        base = target_base
        for i, cname in enumerate(("기준월", "목표주가")):
            c = ws.cell(row=base, column=16 + i, value=cname)
            c.font = Font(bold=True, size=10)
            c.fill = PatternFill("solid", fgColor=SKY)
            c.border = box
            c.alignment = Alignment(horizontal="center")
        for i, tp in enumerate(tg):
            rr = base + 1 + i
            ws.cell(row=rr, column=16, value=tp.get("월")).border = box
            v = ws.cell(row=rr, column=17, value=tp.get("목표주가"))
            v.number_format = "#,##0"
            v.border = box
        last2 = base + len(tg)
        lc = LineChart()
        lc.title = "증권사 목표주가 추이 (원)"
        lc.height, lc.width = 7.2, 14.5
        lc.add_data(Reference(ws, min_col=17, min_row=base, max_row=last2),
                    titles_from_data=True)
        lc.set_categories(Reference(ws, min_col=16, min_row=base + 1, max_row=last2))
        lc.legend = None
        lc.y_axis.majorGridlines = None
        s0 = lc.series[0]
        s0.smooth = False
        s0.graphicalProperties.line.solidFill = RED
        s0.graphicalProperties.line.width = 22000
        ws.add_chart(lc, "H33")


def save_dataframe_to_excel(
    df: pd.DataFrame,
    file_path: Path | str,
    sheet_name: str = "Data",
    metadata: dict | None = None,
    source: str = "네이버 증권",
) -> str:
    """DataFrame을 Excel로 저장. 메타데이터 시트도 함께 기록.

    종목코드 같은 선행 0이 있는 문자열은 엑셀이 자동으로 숫자 변환하지 않도록
    강제로 문자열 타입을 유지.

    Args:
        df: 저장할 DataFrame
        file_path: 저장 경로
        sheet_name: 메인 시트 이름
        metadata: 추가 정보 (수집 시간, 소스 등)
        source: 데이터 소스 이름 (KR: "네이버 증권", US: "Yahoo Finance" 등)

    Returns:
        저장된 파일의 절대 경로 (문자열)
    """
    file_path = Path(file_path)
    file_path.parent.mkdir(parents=True, exist_ok=True)

    df = df.copy()
    # 종목코드 컬럼은 문자열로 강제 (선행 0 보존)
    if "code" in df.columns:
        df["code"] = df["code"].astype(str).str.zfill(6)

    # 영문 열 이름을 한글로 바꿔 저장한다(파일을 여는 사람 기준).
    # query_excel 은 영문 키를 계속 받으므로 필터 호환은 별칭으로 처리한다.
    df = df.rename(columns={c: COLUMN_LABELS_KO.get(c, c) for c in df.columns})

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        _polish_sheet(writer.sheets[sheet_name], df)
        try:
            _add_charts(writer.book, writer.sheets[sheet_name], df, sheet_name)
        except Exception:
            pass    # 그래프는 부가 기능이다 — 실패해도 데이터 저장은 살린다

        # 메타데이터 시트
        meta_rows = [
            {"key": "수집 시간", "value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
            {"key": "소스", "value": source},
            {"key": "행 수", "value": len(df)},
            {"key": "컬럼 수", "value": len(df.columns)},
        ]
        if metadata:
            for k, v in metadata.items():
                meta_rows.append({"key": k, "value": str(v)})

        meta_df = pd.DataFrame(meta_rows)
        meta_df.to_excel(writer, sheet_name="Metadata", index=False)

    return str(file_path.resolve())


def load_excel(file_path: Path | str, sheet_name: str | int = 0) -> pd.DataFrame:
    """Excel 파일을 DataFrame으로 로드.

    종목코드는 문자열로 강제 로드 (선행 0 보존).
    """
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {file_path}")
    # 파일은 한글 헤더로 저장되지만 내부 처리는 영문 키로 통일한다.
    df = pd.read_excel(file_path, sheet_name=sheet_name,
                       dtype={"code": str, "종목코드": str})
    df = df.rename(columns={c: LABEL_TO_KEY.get(c, c) for c in df.columns})
    if "code" in df.columns:
        df["code"] = df["code"].astype(str).str.zfill(6)
    return df


def unknown_filter_columns(df: pd.DataFrame, filters: dict | None) -> list[str]:
    """필터에 적혔지만 이 파일에 없는 컬럼 목록.

    없는 컬럼은 조용히 무시되므로, 알려주지 않으면 "조건이 걸린 결과"로 읽힌다.
    PER 컬럼이 없는 스냅샷에 per_max=10 을 걸면 전 종목이 그대로 나오는데,
    사용자는 그게 저PER 목록인 줄 안다.
    """
    cols = set(df.columns)
    unknown: list[str] = []
    for key, cond in (filters or {}).items():
        base = key
        if not isinstance(cond, dict) and (key.endswith("_max") or key.endswith("_min")):
            base = key[:-4]
        if base not in cols:
            unknown.append(key)
    return unknown


def apply_filters(df: pd.DataFrame, filters: dict) -> pd.DataFrame:
    """DataFrame에 필터 조건을 적용.

    filters 형식:
        {"per": {"max": 10, "min": 0}, "drawdown_pct": {"max": -30}}
        {"foreign_net_5d": {"min": 0}}

    또는 간단한 형태:
        {"per_max": 10, "pbr_max": 1.5}  ← 이것도 지원
    """
    result = df.copy()

    for key, cond in filters.items():
        if isinstance(cond, dict):
            # {"per": {"max": 10, "min": 0}} 형태
            if key not in result.columns:
                continue
            if "min" in cond and cond["min"] is not None:
                result = result[result[key] >= cond["min"]]
            if "max" in cond and cond["max"] is not None:
                result = result[result[key] <= cond["max"]]
            if "equals" in cond:
                result = result[result[key] == cond["equals"]]
        else:
            # 간단한 형태: "per_max": 10
            if key.endswith("_max"):
                col = key[:-4]
                if col in result.columns:
                    result = result[result[col] <= cond]
            elif key.endswith("_min"):
                col = key[:-4]
                if col in result.columns:
                    result = result[result[col] >= cond]
            elif key in result.columns:
                result = result[result[key] == cond]

    return result
